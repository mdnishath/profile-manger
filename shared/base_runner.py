"""
Base worker class for all step runners.

Extracts the ~60-70% duplicated code (validation, backup-code merge,
proxy retry loop, browser lifecycle, login, forced password change,
operation loop, signout, cleanup, error handling) into a Template Method
base class.

Each step runner inherits BaseGmailBotWorker and only implements:
  - _dispatch_operation(op, page, account, ctx)  [REQUIRED]
  - Hook overrides for step-specific behaviour (optional)

External contract preserved:
  GmailBotWorker(worker_id, excel_processor) + process_account(account)
"""

import asyncio
import random
import traceback

import pandas as pd
from playwright.async_api import async_playwright

from src.screen_detector import ScreenDetector
from src.utils import ConfigManager, TOTPGenerator
from src.login_flow import execute_login_flow

from shared.logger import print, _log
from shared.browser import launch_browser, create_context
from shared.signout import perform_signout
from shared import proxy_manager


# ── Constants shared across all steps ────────────────────────────────────────

NETWORK_ERRORS = (
    'net::ERR_', 'NS_ERROR_', 'Connection refused',
    'Connection reset', 'Connection timed out',
    'ERR_PROXY', 'ERR_TUNNEL', 'ERR_SOCKS',
    'SOCKS', 'Proxy connection', 'Network is unreachable',
    'Connection to Google failed', 'ECONNREFUSED', 'ETIMEDOUT',
    'socket hang up', 'Target closed', 'browser has been closed',
)

AUTH_FAILURES = [
    'WRONG_PASSWORD', 'PASSWORD_CHANGED', 'WRONG_2FA_CODE',
    'CAPTCHA_REQUIRED', 'RECOVERY_EMAIL_VERIFICATION',
    'PHONE_VERIFICATION', 'ACCOUNT_LOCKED', 'TOO_MANY_ATTEMPTS',
    'EMAIL_NOT_FOUND', 'ACCOUNT_RECOVERY_REDIRECT',
    'SUSPICIOUS_ACTIVITY', 'UNUSUAL_LOCATION',
    'VERIFY_PHONE_CODE', 'NO_2FA_CREDENTIALS',
    'SENSITIVE_ACTION_BLOCKED', 'DEVICE_APPROVAL_FAILED',
]


class BaseGmailBotWorker:
    """Template Method base for all step runners.

    Subclasses MUST implement:
        _dispatch_operation(op, page, account, ctx) -> str|bool|tuple|list

    Subclasses MAY override:
        _get_proxy()              — default: get_healthy_proxy()
        _use_proxy_retry          — default: True  (Step 4 sets False)
        _max_proxy_retries        — default: 5
        _get_login_url()          — default: config.get_url('login')
        _get_login_kwargs()       — default: {}
        _launch_browser(pw, proxy) — default: launch_browser(pw, proxy=proxy)
        _create_context(browser)  — default: create_context(browser)
        _post_page_setup(page, context) — default: no-op
        _parse_operations(ops_str) — default: split by comma/space, uppercase
        _build_success_kwargs(operations_done, operations_failed, ctx) — default: {}
        _on_operations_complete(page, account, row_index, ctx) — default: no-op
    """

    _use_proxy_retry = True
    _max_proxy_retries = 5

    def __init__(self, worker_id, excel_processor):
        self.worker_id = worker_id
        self.excel_processor = excel_processor
        self.config = ConfigManager()

    # ── Credential extraction & validation ───────────────────────────────

    def _extract_credentials(self, account):
        """Extract and validate core credentials from account dict.

        Returns (email, password, totp_secret, operations_str, row_index).
        Raises ValueError on invalid email/password.
        """
        email = account.get('Email', '')
        password = account.get('Password', '')
        totp_secret = account.get('TOTP Secret', '')
        row_index = account.get('row_index', 0)

        # Validate email
        if pd.isna(email) or not str(email).strip() or str(email).strip().lower() == 'nan':
            raise ValueError('Invalid email address')

        # Validate password
        if pd.isna(password) or not str(password).strip():
            raise ValueError('Invalid password')

        email = str(email).strip()
        password = str(password).strip()
        totp_secret = str(totp_secret).strip() if not pd.isna(totp_secret) else ''

        return email, password, totp_secret, row_index

    def _merge_backup_codes(self, account):
        """Merge Backup Code 1-10 columns into account['Backup Code']."""
        bc_list = []
        for i in range(1, 11):
            val = account.get(f'Backup Code {i}', '')
            if not pd.isna(val) and str(val).strip():
                bc_list.append(str(val).strip())
        account['Backup Code'] = bc_list[0] if bc_list else ''
        return bc_list

    # ── Operations parsing (overridable) ─────────────────────────────────

    def _parse_operations(self, operations_str):
        """Parse operations string into a list of op codes.

        Default: split by comma or space, uppercase.
        Step 2 overrides for legacy mapping.
        """
        ops_str = str(operations_str).upper().replace(',', ' ')
        return [op.strip() for op in ops_str.split() if op.strip()]

    def _get_default_operations(self):
        """Default operations string when Excel column is empty.

        Override per step (e.g. Step 1 returns 'L1', Step 4 returns 'A1').
        """
        return ''

    # ── Proxy (overridable) ──────────────────────────────────────────────

    def _get_proxy(self, exclude=None):
        """Get a proxy for this account. Override for different strategy."""
        return proxy_manager.get_healthy_proxy(exclude=exclude)

    # ── Browser lifecycle (overridable) ──────────────────────────────────

    async def _launch_browser(self, pw, proxy):
        """Launch browser. Override if fingerprint or extra args needed."""
        return await launch_browser(pw, proxy=proxy)

    async def _create_context(self, browser):
        """Create browser context. Override to pass fingerprint."""
        return await create_context(browser)

    async def _post_page_setup(self, page, context):
        """Hook called after page creation, before login.

        Override in Step 3 to set English language headers/cookies.
        """
        pass

    # ── Login (overridable) ──────────────────────────────────────────────

    def _get_login_url(self):
        """Login URL. Step 2 overrides with recovery URL."""
        return self.config.get_url("login")

    def _get_login_kwargs(self):
        """Extra kwargs for execute_login_flow. Step 2 overrides."""
        return {}

    # ── Forced password change (shared) ──────────────────────────────────

    async def _handle_forced_password_change(self, login_result, email, row_index):
        """Save forced password change to Excel. Identical across all steps."""
        forced_new_pw = login_result.get('forced_new_password', '')
        if not forced_new_pw:
            return

        _log(self.worker_id, f"*** FORCED PASSWORD CHANGE: new pw = {forced_new_pw}")
        try:
            from openpyxl import load_workbook as _lwb
            with self.excel_processor.lock:
                _wb = _lwb(self.excel_processor.excel_file)
                _ws = _wb.active
                _headers = [c.value for c in _ws[1]]
                if 'Password' in _headers:
                    _ws.cell(row_index, _headers.index('Password') + 1, forced_new_pw)
                if 'New Password' in _headers:
                    _ws.cell(row_index, _headers.index('New Password') + 1, forced_new_pw)
                _wb.save(self.excel_processor.excel_file)
                _wb.close()
            _log(self.worker_id, f"Password updated in Excel for {email}")
        except Exception as pw_err:
            _log(self.worker_id, f"WARNING: Could not save new password: {pw_err}")

    # ── Operation dispatch (ABSTRACT — must override) ────────────────────

    async def _dispatch_operation(self, op, page, account, ctx):
        """Dispatch a single operation. MUST be overridden by each step.

        Args:
            op: Operation code string (e.g. 'L1', '2a', 'R1', 'A1')
            page: Playwright page
            account: Account dict from Excel
            ctx: Step-specific context dict (e.g. password_url for Step 2).
                 Mutations to this dict are visible to the caller (used by
                 Step 2 to pass back authenticator_key / backup_codes).

        Returns:
            True/False for simple ops, tuple for authenticator, list for backup codes,
            "SKIP - reason" string to skip.
        """
        raise NotImplementedError(
            f"{self.__class__.__name__} must implement _dispatch_operation()"
        )

    # ── Success kwargs builder (overridable) ─────────────────────────────

    def _build_success_kwargs(self, operations_done, operations_failed, ctx):
        """Build extra kwargs for update_row_status on SUCCESS.

        Override in Step 2 to add authenticator_key, backup_codes, op_status.
        Override in Step 1 to save L6/L7 results.
        """
        return {}

    async def _on_operations_complete(self, page, account, row_index, ctx):
        """Hook called after all operations complete, before signout.

        Override in Step 1 to save L6/L7 results to Excel columns.
        """
        pass

    # ── Cleanup (shared) ─────────────────────────────────────────────────

    async def _cleanup_browser(self, context=None, browser=None,
                               socks_bridge=None, pw=None):
        """Close browser resources safely."""
        try:
            if context:
                await asyncio.wait_for(context.close(), timeout=5)
        except Exception:
            pass
        try:
            if browser:
                await asyncio.wait_for(browser.close(), timeout=5)
        except Exception:
            pass
        if socks_bridge:
            try:
                await socks_bridge.stop()
            except Exception:
                pass
        if pw:
            try:
                await pw.stop()
            except Exception:
                pass

    # ── Main flow (Template Method) ──────────────────────────────────────

    async def process_account(self, account):
        """Full orchestration: validate → proxy → login → ops → signout.

        This is the Template Method — subclasses customize via hooks,
        NOT by overriding this method.
        """
        # ── Extract & validate credentials ────────────────────────────
        self._merge_backup_codes(account)
        row_index = account.get('row_index', 0)

        try:
            email, password, totp_secret, row_index = self._extract_credentials(account)
        except ValueError as ve:
            _log(self.worker_id, f"VALIDATE: SKIP - {ve}")
            self.excel_processor.update_row(row_index, status='FAILED', error=str(ve))
            return

        # ── Parse operations ──────────────────────────────────────────
        operations_raw = account.get('Operations', self._get_default_operations())
        operations_list = self._parse_operations(operations_raw)
        total_ops = len(operations_list)

        _log(self.worker_id, "=" * 60)
        _log(self.worker_id, f"ACCOUNT START: {email} (Row {row_index})")
        _log(self.worker_id, f"  TOTP: {'YES' if totp_secret else 'NO'} | Backup: {'YES' if account.get('Backup Code') else 'NO'}")
        _log(self.worker_id, f"  Operations ({total_ops}): {operations_list}")
        _log(self.worker_id, "=" * 60)

        operations_done = []
        operations_failed = []
        ctx = {}  # Step-specific context (populated by hooks)

        # ── Proxy retry loop ──────────────────────────────────────────
        if self._use_proxy_retry:
            await self._process_with_proxy_retry(
                account, email, row_index,
                operations_list, total_ops,
                operations_done, operations_failed, ctx,
            )
        else:
            await self._process_single_attempt(
                account, email, row_index,
                operations_list, total_ops,
                operations_done, operations_failed, ctx,
            )

    # ── Proxy-retry flow (Steps 1, 2, 3) ─────────────────────────────────

    async def _process_with_proxy_retry(self, account, email, row_index,
                                        operations_list, total_ops,
                                        operations_done, operations_failed, ctx):
        """Process with proxy retry on network errors."""
        tried_proxies = []
        using_local_ip = False

        for proxy_attempt in range(self._max_proxy_retries):
            account_proxy = self._get_proxy(exclude=tried_proxies)
            if account_proxy:
                tried_proxies.append(account_proxy)
                _log(self.worker_id, f"[PROXY] {email} → {account_proxy.get('server', '')} (attempt {proxy_attempt+1}/{self._max_proxy_retries})")
            else:
                if using_local_ip:
                    _log(self.worker_id, f"[PROXY] {email} → All proxies exhausted + local IP tried. Giving up.")
                    break
                using_local_ip = True
                _log(self.worker_id, f"[PROXY] {email} → Local IP (no healthy proxy)")

            pw = None
            context = None
            browser = None
            socks_bridge = None

            try:
                pw = await async_playwright().start()
                _log(self.worker_id, "[BROWSER] Launching Chromium...")
                browser, socks_bridge = await self._launch_browser(pw, account_proxy)
                context = await self._create_context(browser)
                page = await context.new_page()
                _log(self.worker_id, "[BROWSER] Browser ready")

                await self._post_page_setup(page, context)

                # ── Login ─────────────────────────────────────────────
                await self._do_login(page, account, email, row_index, ctx)

                # ── Operations ────────────────────────────────────────
                await self._run_operations(
                    page, account, operations_list, total_ops,
                    operations_done, operations_failed, ctx,
                )

                # Hook: post-operations (e.g. Step 1 saves L6/L7)
                await self._on_operations_complete(page, account, row_index, ctx)

                # ── Signout & cleanup ─────────────────────────────────
                await perform_signout(page, self.worker_id)
                await self._cleanup_browser(context, browser, socks_bridge, pw)
                _log(self.worker_id, "[BROWSER] Closed")

                # ── Excel → SUCCESS ───────────────────────────────────
                self._write_success(row_index, operations_done, operations_failed, ctx)

                _log(self.worker_id, f"ACCOUNT DONE: {email} = SUCCESS")
                proxy_manager.mark_alive(account_proxy)
                break  # exit proxy retry loop

            except Exception as e:
                error_str = str(e)
                _log(self.worker_id, "=" * 60)
                _log(self.worker_id, f"ACCOUNT ERROR: {email} (attempt {proxy_attempt+1}/{self._max_proxy_retries})")
                _log(self.worker_id, f"  Error: {error_str}")
                _log(self.worker_id, "=" * 60)

                await self._cleanup_browser(context, browser, socks_bridge, pw)

                is_network = any(kw.lower() in error_str.lower() for kw in NETWORK_ERRORS)

                if is_network and proxy_attempt < self._max_proxy_retries - 1:
                    proxy_manager.mark_dead(account_proxy)
                    delay = min(2 ** proxy_attempt + random.random(), 30)
                    _log(self.worker_id, f"[PROXY] Network error — switching proxy in {delay:.1f}s...")
                    await asyncio.sleep(delay)
                    continue

                if any(f in error_str for f in AUTH_FAILURES):
                    _log(self.worker_id, "  Type: AUTH FAILURE (no retry)")

                if is_network:
                    proxy_manager.mark_dead(account_proxy)

                traceback.print_exc()

                self.excel_processor.update_row_status(
                    row_index=row_index,
                    status='FAILED',
                    operations_done='',
                    error_message=error_str[:500],
                )
                _log(self.worker_id, f"[EXCEL] Row {row_index} -> FAILED")
                break

    # ── Single-attempt flow (Step 4) ──────────────────────────────────────

    async def _process_single_attempt(self, account, email, row_index,
                                      operations_list, total_ops,
                                      operations_done, operations_failed, ctx):
        """Process without proxy retry (Step 4 pattern)."""
        account_proxy = self._get_proxy()
        if account_proxy:
            _log(self.worker_id, f"[PROXY] {email} → {account_proxy.get('server', '')}")
        else:
            _log(self.worker_id, f"[PROXY] {email} → Local IP (no proxy)")

        try:
            async with async_playwright() as p:
                _log(self.worker_id, "[BROWSER] Launching Chromium...")
                browser, socks_bridge = await self._launch_browser(p, account_proxy)
                context = await self._create_context(browser)
                page = await context.new_page()
                _log(self.worker_id, "[BROWSER] Browser ready")

                await self._post_page_setup(page, context)

                # ── Login ─────────────────────────────────────────────
                await self._do_login(page, account, email, row_index, ctx)

                # ── Operations ────────────────────────────────────────
                await self._run_operations(
                    page, account, operations_list, total_ops,
                    operations_done, operations_failed, ctx,
                )

                await self._on_operations_complete(page, account, row_index, ctx)

                # ── Signout & cleanup ─────────────────────────────────
                await perform_signout(page, self.worker_id)

                try:
                    await asyncio.wait_for(context.close(), timeout=5)
                except Exception:
                    pass
                try:
                    await asyncio.wait_for(browser.close(), timeout=5)
                except Exception:
                    pass
                if socks_bridge:
                    await socks_bridge.stop()
                _log(self.worker_id, "[BROWSER] Closed")

                # ── Excel → SUCCESS ───────────────────────────────────
                self._write_success(row_index, operations_done, operations_failed, ctx)
                _log(self.worker_id, f"ACCOUNT DONE: {email} = SUCCESS")

        except Exception as e:
            error_str = str(e)
            _log(self.worker_id, "=" * 60)
            _log(self.worker_id, f"ACCOUNT FAILED: {email}")
            _log(self.worker_id, f"  Error: {error_str}")
            _log(self.worker_id, "=" * 60)

            if any(f in error_str for f in AUTH_FAILURES):
                _log(self.worker_id, "  Type: AUTH FAILURE (no retry)")

            traceback.print_exc()

            self.excel_processor.update_row_status(
                row_index=row_index,
                status='FAILED',
                operations_done='',
                error_message=error_str[:500],
            )
            _log(self.worker_id, f"[EXCEL] Row {row_index} -> FAILED")

            try:
                _locals = locals()
                if _locals.get('browser'):
                    await asyncio.wait_for(browser.close(), timeout=5)
                if _locals.get('socks_bridge'):
                    await socks_bridge.stop()
                _log(self.worker_id, "[BROWSER] Closed")
            except Exception as browser_err:
                _log(self.worker_id, f"[BROWSER] Close error: {browser_err}")

    # ── Login helper (shared) ─────────────────────────────────────────────

    async def _do_login(self, page, account, email, row_index, ctx):
        """Execute login flow. Raises on failure."""
        login_url = self._get_login_url()
        _log(self.worker_id, f"[LOGIN] Starting login for {email}")
        _log(self.worker_id, f"[LOGIN] URL = {login_url}")

        detector = ScreenDetector(page)
        totp_gen = TOTPGenerator()

        login_result = await execute_login_flow(
            page=page,
            account=account,
            worker_id=self.worker_id,
            login_url=login_url,
            detector=detector,
            totp_gen=totp_gen,
            **self._get_login_kwargs(),
        )

        if not login_result.get('success'):
            _log(self.worker_id, f"[LOGIN] FAILED - {login_result.get('error', 'Unknown')}")
            raise Exception(login_result.get('error', 'Unknown login failure'))

        await self._handle_forced_password_change(login_result, email, row_index)

        _log(self.worker_id, f"[LOGIN] SUCCESS - URL = {page.url[:100]}")

        # Store login_result in ctx for step-specific post-login (e.g. Step 2 URL capture)
        ctx['login_result'] = login_result
        ctx['detector'] = detector

        # Hook for post-login setup (e.g. Step 2 URL capture)
        await self._post_login(page, account, ctx)

    async def _post_login(self, page, account, ctx):
        """Hook called after successful login, before operations.

        Override in Step 2 for URL capture + phone confirmation.
        """
        pass

    # ── Operations loop (shared) ──────────────────────────────────────────

    async def _run_operations(self, page, account, operations_list, total_ops,
                              operations_done, operations_failed, ctx):
        """Run all operations with try/except per op."""
        _log(self.worker_id, f"Starting {total_ops} operations: {operations_list}")

        for op_idx, op in enumerate(operations_list, 1):
            _log(self.worker_id, "")
            _log(self.worker_id, f"-- Operation {op} ({op_idx}/{total_ops}) --")
            try:
                result = await self._dispatch_operation(op, page, account, ctx)

                # Handle result
                if isinstance(result, str) and 'SKIP' in result:
                    _log(self.worker_id, f"[OP] {op}: SKIPPED - {result}")
                elif result is False:
                    raise Exception(f"Operation {op} returned False")
                elif result is True:
                    operations_done.append(op)
                    _log(self.worker_id, f"[OP] {op}: SUCCESS")
                elif result is None:
                    # Some ops don't return explicitly — treat as success
                    operations_done.append(op)
                    _log(self.worker_id, f"[OP] {op}: SUCCESS")
                else:
                    # Custom return (string description, tuple, list, etc.)
                    # Let _handle_operation_result process it
                    self._handle_operation_result(op, result, operations_done, ctx)

            except Exception as op_err:
                err_msg = str(op_err)[:100]
                _log(self.worker_id, f"[OP] {op} FAILED: {err_msg}")
                operations_failed.append(f"{op}: {err_msg}")
                remaining = total_ops - op_idx
                _log(self.worker_id, f"[OP] SKIP to next operation ({remaining} remaining)...")
                continue

        # Summary
        _log(self.worker_id, "")
        _log(self.worker_id, "=" * 60)
        _log(self.worker_id, f"ALL OPERATIONS DONE")
        _log(self.worker_id, f"  Done:   {len(operations_done)}/{total_ops} -> {', '.join(str(d) for d in operations_done) or 'none'}")
        if operations_failed:
            _log(self.worker_id, f"  Failed: {len(operations_failed)}/{total_ops} -> {', '.join(operations_failed)}")
        _log(self.worker_id, "=" * 60)

    def _handle_operation_result(self, op, result, operations_done, ctx):
        """Handle non-standard operation results.

        Override in Step 1 for L6/L7 tuple results.
        Override in Step 2 for authenticator tuple / backup codes list.
        Default: treat truthy as success.
        """
        if result:
            operations_done.append(str(op))
            _log(self.worker_id, f"[OP] {op}: SUCCESS")
        else:
            raise Exception(f"Operation {op} returned falsy: {result}")

    # ── Write SUCCESS to Excel ────────────────────────────────────────────

    def _write_success(self, row_index, operations_done, operations_failed, ctx):
        """Write SUCCESS status to Excel."""
        failed_msg = ' | '.join(operations_failed) if operations_failed else ''
        ops_str = ', '.join(str(d) for d in operations_done) if operations_done else 'None'

        extra_kwargs = self._build_success_kwargs(operations_done, operations_failed, ctx)

        self.excel_processor.update_row_status(
            row_index=row_index,
            status='SUCCESS',
            operations_done=ops_str,
            error_message=failed_msg,
            **extra_kwargs,
        )
        _log(self.worker_id, f"[EXCEL] Row {row_index} -> SUCCESS")
