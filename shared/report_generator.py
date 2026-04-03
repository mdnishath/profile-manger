"""
MailNexus Pro Report Generator
Creates professional XLS reports with multiple tabs, charts, and meaningful error analysis.
Auto-detects columns from actual data — works with both old and new operation formats.
"""

import re
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers,
)
## Charts removed — clean text-only reports
from openpyxl.utils import get_column_letter

from shared.logger import print


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# COLOUR PALETTE — dark navy theme
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
_NAVY       = '1B2A4A'
_WHITE      = 'FFFFFF'
_LIGHT_GREY = 'F5F6FA'
_MID_GREY   = 'E8E8E8'
_DARK_TEXT  = '2D3436'
_SUB_TEXT   = '636E72'
_GREEN      = '00B894'
_GREEN_BG   = 'DFF5EF'
_RED        = 'D63031'
_RED_BG     = 'FFEAEA'
_YELLOW     = 'FDCB6E'
_YELLOW_BG  = 'FFF8E1'
_BLUE       = '0984E3'
_BLUE_BG    = 'E3F2FD'
_ORANGE     = 'E17055'
_LINK_BLUE  = '0563C1'

# ── Style objects ────────────────────────────────────────────────────────────
HEADER_FILL  = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type='solid')
HEADER_FONT  = Font(name='Segoe UI', bold=True, color=_WHITE, size=10)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

SUCCESS_ROW  = PatternFill(start_color=_GREEN_BG, end_color=_GREEN_BG, fill_type='solid')
FAILED_ROW   = PatternFill(start_color=_RED_BG,   end_color=_RED_BG,   fill_type='solid')
SKIP_ROW     = PatternFill(start_color=_YELLOW_BG, end_color=_YELLOW_BG, fill_type='solid')
PENDING_ROW  = PatternFill(start_color=_BLUE_BG,  end_color=_BLUE_BG,  fill_type='solid')
ALT_ROW      = PatternFill(start_color=_LIGHT_GREY, end_color=_LIGHT_GREY, fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin', color=_MID_GREY),
    right=Side(style='thin', color=_MID_GREY),
    top=Side(style='thin', color=_MID_GREY),
    bottom=Side(style='thin', color=_MID_GREY),
)
THICK_BOTTOM = Border(bottom=Side(style='medium', color=_NAVY))

TITLE_FONT    = Font(name='Segoe UI', bold=True, size=20, color=_NAVY)
SUBTITLE_FONT = Font(name='Segoe UI', bold=True, size=13, color=_DARK_TEXT)
SECTION_FONT  = Font(name='Segoe UI', bold=True, size=11, color=_NAVY)
METRIC_VAL    = Font(name='Segoe UI', bold=True, size=26, color=_NAVY)
METRIC_LABEL  = Font(name='Segoe UI', size=10, color=_SUB_TEXT)
BODY_FONT     = Font(name='Segoe UI', size=10, color=_DARK_TEXT)
BOLD_BODY     = Font(name='Segoe UI', bold=True, size=10, color=_DARK_TEXT)
LINK_FONT     = Font(name='Segoe UI', size=10, color=_LINK_BLUE, underline='single')
STATUS_SUCCESS = Font(name='Segoe UI', bold=True, size=10, color=_GREEN)
STATUS_FAILED  = Font(name='Segoe UI', bold=True, size=10, color=_RED)
STATUS_SKIP    = Font(name='Segoe UI', bold=True, size=10, color=_YELLOW)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ERROR MESSAGE TRANSLATION
# Maps raw error codes → (Short Title, Explanation, Suggestion)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ERROR_TRANSLATIONS = {
    'LOGIN_TIMEOUT': (
        'Login Timed Out',
        'Could not complete sign-in within the allowed attempts.',
        'Check password & TOTP secret. Google may require manual verification.'
    ),
    'PASSWORD_INCORRECT': (
        'Wrong Password',
        'Google rejected the password entered for this account.',
        'Verify the password in your input file is correct.'
    ),
    'TOTP_FAILED': (
        'TOTP Verification Failed',
        'The 2FA code generated from TOTP secret was rejected.',
        'Check that the TOTP secret is valid and time is synced.'
    ),
    'ACCOUNT_LOCKED': (
        'Account Locked',
        'Google has temporarily locked this account due to suspicious activity.',
        'Wait 24-48 hours and try again, or verify via recovery options.'
    ),
    'ACCOUNT_DISABLED': (
        'Account Disabled',
        'This Google account has been disabled or suspended.',
        'Account cannot be processed. Contact Google support if needed.'
    ),
    'CHALLENGE_REQUIRED': (
        'Security Challenge',
        'Google requires additional identity verification (phone, email, etc.).',
        'Complete the challenge manually first, then retry.'
    ),
    'CHALLENGE_UNRESOLVABLE': (
        'Challenge — No Credentials',
        'Google shows a verification challenge but no matching credentials are in your Excel.',
        'Add Recovery Email or Recovery Phone to the Excel for this account. '
        'If 2FA is enabled, add TOTP Secret or Backup Code.'
    ),
    'NO_2FA_OPTION': (
        '2FA Selection — No Matching Option',
        'Google shows 2FA options but the bot could not find/click any of them.',
        'Add Recovery Email or Recovery Phone to the Excel. '
        'If the account has Authenticator enabled, add the TOTP Secret.'
    ),
    'NO_2FA_CREDENTIALS': (
        '2FA Required — No Credentials',
        'Google requires 2FA but no TOTP Secret, Backup Code, or Recovery data is in your Excel.',
        'Add TOTP Secret, Backup Code, Recovery Email, or Recovery Phone to the Excel.'
    ),
    'NO_RECOVERY_EMAIL': (
        'Recovery Email Required',
        'Google asks to confirm your recovery email but the Recovery Email column is empty.',
        'Add the Recovery Email for this account in the Excel.'
    ),
    'NO_RECOVERY_PHONE': (
        'Recovery Phone Required',
        'Google asks to confirm your recovery phone but the Recovery Phone column is empty.',
        'Add the Recovery Phone for this account in the Excel.'
    ),
    'VERIFY_CODE_SENT': (
        'Verification Code Sent',
        'Google sent a verification code to recovery email/phone. Bot cannot read it automatically.',
        'Login manually to clear the challenge, or add TOTP Secret to avoid this.'
    ),
    'SUSPICIOUS_ACTIVITY': (
        'Suspicious Activity',
        'Google flagged suspicious activity on this account and requires a password change.',
        'Re-run the account — the bot will auto-change the password and continue.'
    ),
    'CAPTCHA': (
        'CAPTCHA Required',
        'Google is showing a CAPTCHA that cannot be solved automatically.',
        'Try using a different proxy or wait before retrying.'
    ),
    'NETWORK_ERROR': (
        'Network Error',
        'Connection to Google failed or timed out.',
        'Check internet connection and proxy settings.'
    ),
    'BROWSER_CRASH': (
        'Browser Crashed',
        'The automation browser crashed unexpectedly.',
        'Reduce number of workers or check system resources.'
    ),
    'PROXY_ERROR': (
        'Proxy Connection Failed',
        'Could not connect through the assigned proxy server.',
        'Verify proxy is working. Try a different proxy or local IP.'
    ),
    'PAGE_LOAD_TIMEOUT': (
        'Page Load Timeout',
        'Google page took too long to load.',
        'Check internet speed and proxy latency.'
    ),
    'ELEMENT_NOT_FOUND': (
        'UI Element Not Found',
        'Expected button or form field was not found on the page.',
        'Google may have changed their UI. Check for bot updates.'
    ),
    'OPERATION_FAILED': (
        'Operation Failed',
        'The requested operation could not be completed.',
        'Check the screenshot folder for visual details.'
    ),
    'UNEXPECTED_PAGE': (
        'Unexpected Page',
        'Landed on an unexpected Google page during processing.',
        'Account may need manual attention. Check screenshots.'
    ),
    'SESSION_EXPIRED': (
        'Session Expired',
        'Login session expired before operations could complete.',
        'Try again — the bot will re-login automatically.'
    ),
    '400': (
        'Bad Request (400)',
        'Google returned a 400 error page.',
        'Session may be invalid. Bot will auto-retry.'
    ),
}

# Regex to extract error code from raw message like "LOGIN_TIMEOUT - blah blah"
_ERROR_CODE_PAT = re.compile(r'^([A-Z_0-9]+)\s*[-:]?\s*')
_URL_PAT = re.compile(r'https?://[^\s|,]+')


def _translate_error(raw_msg: str, account: dict = None) -> dict:
    """
    Context-aware error translation.
    Analyses the raw error message AND the account data to produce a
    meaningful, human-friendly explanation of WHY the error happened.
    """
    if not raw_msg or not isinstance(raw_msg, str) or raw_msg.strip() == '':
        return {'code': '', 'title': '', 'explanation': '', 'suggestion': '', 'raw': ''}

    raw = str(raw_msg).strip()
    account = account or {}

    # Try to match error code
    m = _ERROR_CODE_PAT.match(raw)
    code = m.group(1) if m else ''

    if code in ERROR_TRANSLATIONS:
        title, explanation, suggestion = ERROR_TRANSLATIONS[code]
    else:
        title, explanation, suggestion = '', '', ''
        for known_code, (t, e, s) in ERROR_TRANSLATIONS.items():
            if known_code in raw.upper():
                code = known_code
                title, explanation, suggestion = t, e, s
                break

    if not title:
        title = 'Processing Error'
        explanation = raw[:120] if len(raw) > 120 else raw
        suggestion = 'Check the screenshot folder for more details.'

    # ── SMART CONTEXT ANALYSIS ────────────────────────────────────────────
    # Cross-reference error with account data to give precise explanation
    new_pw    = str(account.get('New Password', '')).strip()
    old_pw    = str(account.get('Password', '')).strip()
    proc_at   = str(account.get('Processed At', '')).strip()

    if new_pw and new_pw not in ('nan', 'None', ''):
        _has_new_pw = True
    else:
        _has_new_pw = False

    # LOGIN_TIMEOUT at PASSWORD_INPUT + account has New Password value
    # → means password was already changed in a prior run
    if code == 'LOGIN_TIMEOUT' and 'PASSWORD_INPUT' in raw:
        if _has_new_pw:
            title = 'Password Already Changed'
            explanation = (
                f'Login failed because the password was already changed '
                f'(to "{new_pw}") in a previous run. '
                f'The old password in your input file no longer works.'
            )
            suggestion = (
                f'Update the Password column in your input file to: {new_pw} '
                f'then re-run. Or remove this account from the input file.'
            )
        else:
            title = 'Wrong Password / Login Blocked'
            explanation = (
                'Could not login — Google rejected the password. '
                'Either the password is wrong or Google blocked the sign-in.'
            )
            suggestion = (
                'Verify the password is correct. If the account has 2FA, '
                'check the TOTP secret. Try again after some time.'
            )

    # LOGIN_TIMEOUT at TOTP/challenge screens
    elif code == 'LOGIN_TIMEOUT' and ('CHALLENGE' in raw.upper() or 'SELECTION' in raw.upper()):
        title = 'Stuck at Verification'
        explanation = (
            'Login got stuck at a Google verification/challenge screen. '
            'Google is requesting additional identity proof.'
        )
        suggestion = (
            'Add Recovery Email or Recovery Phone to the Excel for this account. '
            'If 2FA is enabled, add TOTP Secret. Or login manually to clear the challenge.'
        )

    # LOGIN_TIMEOUT generic (not at PASSWORD_INPUT)
    elif code == 'LOGIN_TIMEOUT' and 'PASSWORD_INPUT' not in raw:
        title = 'Login Did Not Complete'
        explanation = (
            'The login process did not finish within the allowed time. '
            'Google may be slow or showing unexpected pages.'
        )
        suggestion = 'Check proxy/internet speed. Try again with fewer workers.'

    # OPERATION_FAILED with context
    elif code == 'OPERATION_FAILED':
        ops = str(account.get('Operations', '')).strip()
        ops_done = str(account.get('Operations Done', '')).strip()
        if ops_done and ops_done not in ('nan', 'None', ''):
            title = 'Partially Completed'
            explanation = (
                f'Some operations completed ({ops_done}) but not all. '
                f'Requested: {ops}.'
            )
            suggestion = 'Check screenshots for which operation failed and why.'
        else:
            explanation = 'No operations could be completed after login.'
            suggestion = 'Check screenshots. The account page may have changed.'

    return {
        'code': code,
        'title': title,
        'explanation': explanation,
        'suggestion': suggestion,
        'raw': raw,
    }


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# COLUMN DETECTION — auto-detect from actual data
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Priority order for data sheets — columns appear in this order if present.
# Any column NOT in this list appears at the end alphabetically.
_PRIORITY_COLUMNS = [
    # ── Input columns (exact user-specified order) ──
    'First Name',
    'Email', 'Password', 'TOTP Secret',
    'Recovery Email', 'Recovery Phone', 'Year',
    'Backup Code 1', 'Backup Code 2', 'Backup Code 3', 'Backup Code 4',
    'Backup Code 5', 'Backup Code 6', 'Backup Code 7', 'Backup Code 8',
    'Backup Code 9', 'Backup Code 10',
    'Operations',
    'New Password', 'New Recovery Phone', 'New Recovery Email',
    'New 2FA Phone', 'Last Name',
    'GMB Name', 'gmb_name',
    'GMB URL', 'review_place_url',
    'Review Text', 'review_text',
    'Review Stars', 'review_stars',
    'Appeal Message', 'appeal_message',
    # ── Processing output columns (after input columns) ──
    'Status', 'Error Message', 'Operations Done',
    'Share Link',
    'Processed At',
]

# Columns to HIDE from the report (internal / redundant)
_HIDDEN_COLUMNS = {
    'Backup Code', 'Backup Codes', 'row_index', 'Screenshots Folder',
    'New Authenticator Key',  # redundant — TOTP Secret holds effective value
    'Password (Current)', 'TOTP (Current)',  # legacy — originals are overwritten
    'Recovery Email (Current)', 'Recovery Phone (Current)',
    'Proxy',  # internal
}

# Display-friendly header names for columns with internal/snake_case names
_DISPLAY_NAMES = {
    'gmb_name':         'GMB Name',
    'review_place_url': 'GMB URL',
    'review_text':      'Review Text',
    'review_stars':     'Review Stars',
    'appeal_message':   'Appeal Message',
    'Processed At':     'Date',
}

# "Result" columns — values that were APPLIED to accounts (highlight these)
_RESULT_COLUMNS = {
    'New Password', 'New Recovery Phone', 'New Recovery Email',
    'New 2FA Phone', 'Share Link',
}

# Highlight fill for result cells that have actual values
_RESULT_FILL = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
_RESULT_FONT = Font(name='Segoe UI', bold=True, size=10, color='1B5E20')

# Detect operation columns — both old (Op1-Op8) and new (Op1, Op2a-Op6b, Op7-Op9)
_OP_COL_PAT = re.compile(r'^Op\d+[ab]?\s*:', re.IGNORECASE)

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# STEP-AWARE COLUMN DEFINITIONS (per-step, not grouped)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

_BACKUP_CODE_COLS = [f'Backup Code {i}' for i in range(1, 11)]

# ── Step 1: Language / Activity ───────────────────────────────────────────────
_STEP1_ALL_COLUMNS = [
    'Email', 'Password', 'TOTP Secret', 'Operations',
    'Status', 'Operations Done', 'Error Message',
    'Map Used', 'Gmail Year', 'Processed At',
]
_STEP1_SUCCESS_COLUMNS = [
    'Email', 'Operations Done', 'Map Used', 'Gmail Year', 'Processed At',
]
_STEP1_FAILED_COLUMNS = [
    'Email', 'Password', 'TOTP Secret',
    'Status', 'Error Message', 'Operations', 'Processed At',
]

# ── Step 2: Security / Modification ──────────────────────────────────────────
_STEP2_ALL_COLUMNS = [
    'Email', 'Password', 'TOTP Secret', 'Operations',
    'New Password', 'New Recovery Phone', 'New Recovery Email', 'New 2FA Phone',
    'First Name', 'Last Name',
    'Status', 'Operations Done', 'Error Message',
] + _BACKUP_CODE_COLS + ['Processed At']

# Success sheet uses original column names — _apply_effective_values() overwrites
# them with updated data if changed, so no need for (Current) suffix or
# New Authenticator Key (TOTP Secret already holds the effective value).
_STEP2_SUCCESS_COLUMNS = [
    'Email', 'Password', 'TOTP Secret',
    'Recovery Email', 'Recovery Phone',
    'First Name', 'Last Name', 'Operations Done',
] + _BACKUP_CODE_COLS + ['Processed At']

_STEP2_FAILED_COLUMNS = [
    'Email', 'Password', 'TOTP Secret',
    'Status', 'Error Message', 'Operations', 'Processed At',
]

# ── Step 3: Maps Reviews ─────────────────────────────────────────────────────
_STEP3_ALL_COLUMNS = [
    'First Name', 'Email', 'Password', 'TOTP Secret',
    'GMB Name', 'gmb_name', 'GMB URL', 'review_place_url',
    'Review Text', 'review_text', 'Review Stars', 'review_stars',
    'Status', 'Operations Done', 'Error Message',
    'Share Link', 'Processed At',
]
_STEP3_SUCCESS_COLUMNS = [
    'First Name', 'Email', 'Password', 'TOTP Secret',
    'GMB Name', 'gmb_name', 'GMB URL', 'review_place_url',
    'Review Text', 'review_text',
    'Share Link', 'Processed At',
]
_STEP3_FAILED_COLUMNS = [
    'Email', 'Status', 'Error Message', 'GMB URL', 'review_place_url', 'Processed At',
]

# ── Step 4: Account Appeals ──────────────────────────────────────────────────
_STEP4_ALL_COLUMNS = [
    'Email', 'Appeal Message', 'appeal_message',
    'Status', 'Operations Done', 'Error Message', 'Processed At',
]
_STEP4_SUCCESS_COLUMNS = [
    'Email', 'Appeal Message', 'appeal_message', 'Operations Done', 'Processed At',
]
_STEP4_FAILED_COLUMNS = [
    'Email', 'Status', 'Error Message', 'Appeal Message', 'appeal_message', 'Processed At',
]

# ── Per-step lookup table ────────────────────────────────────────────────────
_STEP_COLUMNS = {
    'step1': {'all': _STEP1_ALL_COLUMNS, 'success': _STEP1_SUCCESS_COLUMNS, 'failed': _STEP1_FAILED_COLUMNS},
    'step2': {'all': _STEP2_ALL_COLUMNS, 'success': _STEP2_SUCCESS_COLUMNS, 'failed': _STEP2_FAILED_COLUMNS},
    'step3': {'all': _STEP3_ALL_COLUMNS, 'success': _STEP3_SUCCESS_COLUMNS, 'failed': _STEP3_FAILED_COLUMNS},
    'step4': {'all': _STEP4_ALL_COLUMNS, 'success': _STEP4_SUCCESS_COLUMNS, 'failed': _STEP4_FAILED_COLUMNS},
}

# ── Input template columns (for XLS template generator) ──────────────────────
_STEP_TEMPLATE_COLUMNS = {
    'step1': ['Email', 'Password', 'TOTP Secret'] + _BACKUP_CODE_COLS + ['Operations'],
    'step2': ['Email', 'Password', 'TOTP Secret'] + _BACKUP_CODE_COLS + [
        'Operations', 'New Password', 'New Recovery Phone', 'New Recovery Email',
        'New 2FA Phone', 'First Name', 'Last Name',
    ],
    'step3': ['Email', 'Password', 'TOTP Secret'] + _BACKUP_CODE_COLS + [
        'Operations', 'First Name', 'GMB Name', 'GMB URL', 'Review Text', 'Review Stars',
    ],
    'step4': ['Email', 'Password', 'TOTP Secret'] + _BACKUP_CODE_COLS + [
        'Operations', 'Appeal Message',
    ],
}

# Example/placeholder values for template row 2
_TEMPLATE_EXAMPLES = {
    'Email': 'example@gmail.com',
    'Password': 'YourPassword123',
    'TOTP Secret': 'BASE32SECRET',
    'Operations': 'see docs',
    'New Password': 'NewPass456',
    'New Recovery Phone': '+1234567890',
    'New Recovery Email': 'recovery@email.com',
    'New 2FA Phone': '+1234567890',
    'First Name': 'John',
    'Last Name': 'Doe',
    'GMB Name': 'My Business Name',
    'GMB URL': 'https://maps.google.com/...',
    'Review Text': 'Great service!',
    'Review Stars': 5,
    'Appeal Message': 'Your appeal text here',
}
for i in range(1, 11):
    _TEMPLATE_EXAMPLES[f'Backup Code {i}'] = f'CODE{i}'


def _get_step_columns(all_columns: list, step_name: str, sheet_type: str, accounts: list = None) -> list:
    """Return step-specific columns for a given sheet type.

    Args:
        all_columns: All columns present in the data.
        step_name:   'step1', 'step2', 'step3', 'step4', or '' (auto-detect).
        sheet_type:  'all', 'success', or 'failed'.
        accounts:    Account data — used to filter out empty columns.

    Returns:
        Ordered list of columns to include, filtered to only those present and non-empty.
    """
    _NAN = {'nan', 'None', 'NaT', '', None}

    # Build set of columns that have at least one non-empty value
    non_empty = set(all_columns)
    if accounts:
        non_empty = set()
        for a in accounts:
            for k, v in a.items():
                if str(v).strip() not in _NAN:
                    non_empty.add(k)

    step_def = _STEP_COLUMNS.get(step_name)
    if step_def:
        template = step_def.get(sheet_type, step_def['all'])
        return [c for c in template if c in all_columns and c in non_empty]

    # Fallback: use full classify (also filters empty)
    return _classify_columns(all_columns, accounts)


def _apply_effective_values(accounts: list, step_name: str) -> list:
    """Overwrite original columns with effective (current) values for success reports.

    Password       → New Password if changed, else keep original
    TOTP Secret    → New Authenticator Key if generated, else keep original
    Recovery Email  → New Recovery Email if changed, else keep original
    Recovery Phone → New Recovery Phone if changed, else keep original

    Works for ALL steps — if no "New X" data exists the original is kept as-is.
    Returns a NEW list of dicts (does not mutate originals).
    """
    _NAN = ('nan', 'None', 'NaT', '')

    result = []
    for a in accounts:
        row = dict(a)  # shallow copy
        ops_done = str(row.get('Operations Done', '') or '').strip()

        # Password
        new_pw = str(row.get('New Password', '') or '').strip()
        if new_pw and new_pw not in _NAN:
            pw_changed = any(k in ops_done for k in ['Change Password', 'S2: 1'])
            if pw_changed:
                row['Password'] = new_pw

        # TOTP Secret ← New Authenticator Key (no separate column needed)
        new_key = str(row.get('New Authenticator Key', '') or '').strip()
        if new_key and new_key not in _NAN:
            row['TOTP Secret'] = new_key

        # Recovery Email
        new_rec_email = str(row.get('New Recovery Email', '') or '').strip()
        if new_rec_email and new_rec_email not in _NAN:
            rec_changed = any(k in ops_done for k in ['Recovery Email', 'S2: 3a', 'S2: 3A'])
            if rec_changed:
                row['Recovery Email'] = new_rec_email

        # Recovery Phone
        new_rec_phone = str(row.get('New Recovery Phone', '') or '').strip()
        if new_rec_phone and new_rec_phone not in _NAN:
            rec_changed = any(k in ops_done for k in ['Recovery Phone', 'S2: 2a', 'S2: 2A'])
            if rec_changed:
                row['Recovery Phone'] = new_rec_phone

        result.append(row)
    return result


def _classify_columns(all_columns: list, accounts: list = None) -> list:
    """Return columns in display order, excluding hidden and empty ones.

    If accounts is provided, also drops columns where ALL rows are empty/nan.
    """
    _NAN = {'nan', 'None', 'NaT', '', None}

    # Build set of columns that have at least one non-empty value
    if accounts:
        non_empty = set()
        for a in accounts:
            for k, v in a.items():
                if str(v).strip() not in _NAN:
                    non_empty.add(k)
    else:
        non_empty = set(all_columns)

    priority = []
    op_cols = []
    other = []

    for col in all_columns:
        if col in _HIDDEN_COLUMNS:
            continue
        # Skip columns with NO data in any row
        if col not in non_empty:
            continue
        if col in _PRIORITY_COLUMNS:
            priority.append(col)
        elif _OP_COL_PAT.match(col):
            op_cols.append(col)
        else:
            other.append(col)

    # Sort priority columns by their defined order
    priority.sort(key=lambda c: _PRIORITY_COLUMNS.index(c) if c in _PRIORITY_COLUMNS else 999)
    op_cols.sort()
    other.sort()

    return priority + op_cols + other


def _detect_op_columns(all_columns: list) -> list:
    """Find all operation status columns."""
    return [c for c in all_columns if _OP_COL_PAT.match(c)]


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# UTILITY HELPERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _status_of(account: dict) -> str:
    """Normalised status string."""
    return str(account.get('Status', '')).strip().upper()


def _auto_fit(ws, min_w=10, max_w=55):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        best = min_w
        for cell in col_cells:
            try:
                length = len(str(cell.value or ''))
                if length > best:
                    best = length
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(best + 3, max_w)


def _style_row(ws, row, fill, num_cols):
    """Apply fill, border, alignment to a whole row."""
    for c in range(1, num_cols + 1):
        cell = ws.cell(row, c)
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        cell.font = BODY_FONT


def _set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
    """Write a cell with optional styling."""
    cell = ws.cell(row, col, value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    return cell


def _make_hyperlink(ws, row, col, url, display_text=None):
    """Make a cell a clickable hyperlink."""
    cell = ws.cell(row, col, display_text or url)
    cell.hyperlink = url
    cell.font = LINK_FONT
    return cell


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 1: DASHBOARD (Summary + Charts)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _write_dashboard(ws, accounts):
    """Professional dashboard with KPIs, pie chart, op breakdown + bar chart."""
    total = len(accounts)
    success = sum(1 for a in accounts if 'SUCCESS' in _status_of(a))
    failed = sum(1 for a in accounts if 'FAILED' in _status_of(a))
    skipped = sum(1 for a in accounts if 'SKIP' in _status_of(a))
    pending = total - success - failed - skipped
    rate = f'{success * 100 / total:.1f}%' if total > 0 else '0%'
    fail_rate = f'{failed * 100 / total:.1f}%' if total > 0 else '0%'
    ts = datetime.now().strftime('%d %b %Y | %H:%M')

    # ── Title Banner ──────────────────────────────────────────────────────
    ws.merge_cells('A1:H1')
    c = _set_cell(ws, 1, 1, '  MAILNEXUS PRO  ', TITLE_FONT,
                  PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type='solid'))
    c.font = Font(name='Segoe UI', bold=True, size=20, color=_WHITE)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 48

    ws.merge_cells('A2:H2')
    c2 = _set_cell(ws, 2, 1, f'Processing Report | {ts}',
                   Font(name='Segoe UI', italic=True, size=10, color=_SUB_TEXT))
    c2.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 24

    # ── KPI Cards (row 4-6) ──────────────────────────────────────────────
    row = 4
    _set_cell(ws, row, 1, 'KEY METRICS', SECTION_FONT)
    ws.merge_cells(f'A{row}:H{row}')
    ws.cell(row, 1).alignment = Alignment(horizontal='left')
    row += 1

    # Draw KPI boxes across columns A-H (4 metrics)
    kpis = [
        (str(total),   'Total Accounts', PatternFill(start_color=_BLUE_BG, end_color=_BLUE_BG, fill_type='solid'),
         Font(name='Segoe UI', bold=True, size=26, color=_BLUE)),
        (str(success), 'Successful', PatternFill(start_color=_GREEN_BG, end_color=_GREEN_BG, fill_type='solid'),
         Font(name='Segoe UI', bold=True, size=26, color=_GREEN)),
        (str(failed),  'Failed', PatternFill(start_color=_RED_BG, end_color=_RED_BG, fill_type='solid'),
         Font(name='Segoe UI', bold=True, size=26, color=_RED)),
        (rate,         'Success Rate', PatternFill(start_color=_YELLOW_BG, end_color=_YELLOW_BG, fill_type='solid'),
         Font(name='Segoe UI', bold=True, size=26, color='F39C12')),
    ]

    for idx, (val, label, bg, val_font) in enumerate(kpis):
        col_start = 1 + idx * 2
        col_end   = col_start + 1

        # Value row
        ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
        vc = _set_cell(ws, row, col_start, val, val_font, bg)
        vc.alignment = Alignment(horizontal='center', vertical='center')
        vc.border = THIN_BORDER
        # Also style the merged cell's partner
        ws.cell(row, col_end).fill = bg
        ws.cell(row, col_end).border = THIN_BORDER

        # Label row
        ws.merge_cells(start_row=row + 1, start_column=col_start, end_row=row + 1, end_column=col_end)
        lc = _set_cell(ws, row + 1, col_start, label, METRIC_LABEL, bg)
        lc.alignment = Alignment(horizontal='center', vertical='center')
        lc.border = THIN_BORDER
        ws.cell(row + 1, col_end).fill = bg
        ws.cell(row + 1, col_end).border = THIN_BORDER

    ws.row_dimensions[row].height = 44
    ws.row_dimensions[row + 1].height = 22
    row += 3

    # ── Extra metrics row ─────────────────────────────────────────────────
    extras = [
        ('Pending', str(pending)),
        ('Skipped', str(skipped)),
        ('Failure Rate', fail_rate),
    ]
    _set_cell(ws, row, 1, 'ADDITIONAL', SECTION_FONT)
    row += 1
    for eidx, (elabel, evalue) in enumerate(extras):
        _set_cell(ws, row, 1 + eidx * 3, f'{elabel}:', BOLD_BODY)
        _set_cell(ws, row, 2 + eidx * 3, evalue, BODY_FONT)
    row += 2

    # ── Operation Breakdown (clean text table — no charts) ───────────────
    op_cols = _detect_op_columns([k for a in accounts for k in a.keys()])
    seen = set()
    op_cols = [oc for oc in op_cols if not (oc in seen or seen.add(oc))]
    # Filter out operation columns where ALL values are empty/N/A
    _NAN_VALS = {'', 'NONE', 'NAN', 'NAT', 'NOT REQUESTED'}
    op_cols = [oc for oc in op_cols if any(
        str(a.get(oc, '')).strip().upper() not in _NAN_VALS for a in accounts
    )]

    if op_cols:
        _set_cell(ws, row, 1, 'OPERATION BREAKDOWN', SECTION_FONT)
        row += 1

        op_hdrs = ['Operation', 'Success', 'Failed', 'Skipped']
        for ci, h in enumerate(op_hdrs, 1):
            _set_cell(ws, row, ci, h, HEADER_FONT, HEADER_FILL, HEADER_ALIGN, THIN_BORDER)
        row += 1

        for op_col in op_cols:
            s, f, sk = 0, 0, 0
            for a in accounts:
                v = str(a.get(op_col, '')).strip().upper()
                if 'SUCCESS' in v: s += 1
                elif 'FAILED' in v or 'FAIL' in v: f += 1
                elif 'SKIP' in v: sk += 1

            display_name = op_col.split(':', 1)[1].strip() if ':' in op_col else op_col
            _set_cell(ws, row, 1, display_name, BOLD_BODY, border=THIN_BORDER)
            _set_cell(ws, row, 2, s, BODY_FONT,
                      PatternFill(start_color=_GREEN_BG, end_color=_GREEN_BG, fill_type='solid') if s > 0 else None,
                      Alignment(horizontal='center'), THIN_BORDER)
            _set_cell(ws, row, 3, f, BODY_FONT,
                      PatternFill(start_color=_RED_BG, end_color=_RED_BG, fill_type='solid') if f > 0 else None,
                      Alignment(horizontal='center'), THIN_BORDER)
            _set_cell(ws, row, 4, sk, BODY_FONT,
                      PatternFill(start_color=_YELLOW_BG, end_color=_YELLOW_BG, fill_type='solid') if sk > 0 else None,
                      Alignment(horizontal='center'), THIN_BORDER)
            row += 1

    # ── Changes Applied (per account) ────────────────────────────────────
    # Collect which accounts have any new/applied values
    change_rows = []
    for a in accounts:
        changes = []
        for col_name, display_name in [
            ('New Password', 'Password'),
            ('New Recovery Phone', 'Recovery Phone'),
            ('New Recovery Email', 'Recovery Email'),
            ('New 2FA Phone', '2FA Phone'),
            ('New Authenticator Key', 'Auth Key'),
        ]:
            val = str(a.get(col_name, '')).strip()
            if val and val not in ('nan', 'None', 'NaT', ''):
                changes.append((display_name, val))
        if changes:
            change_rows.append((str(a.get('Email', '')), _status_of(a), changes))

    if change_rows:
        row += 2
        _set_cell(ws, row, 1, 'APPLIED VALUES PER ACCOUNT', SECTION_FONT)
        row += 1

        # Headers
        chg_hdrs = ['Email', 'Status', 'Field Changed', 'New Value']
        for ci, h in enumerate(chg_hdrs, 1):
            _set_cell(ws, row, ci, h, HEADER_FONT, HEADER_FILL, HEADER_ALIGN, THIN_BORDER)
        row += 1

        for email, status, changes in change_rows:
            for cidx, (field_name, value) in enumerate(changes):
                _set_cell(ws, row, 1, email if cidx == 0 else '', BOLD_BODY, border=THIN_BORDER)
                st_cell = _set_cell(ws, row, 2, status if cidx == 0 else '', BODY_FONT, border=THIN_BORDER)
                if cidx == 0:
                    if 'SUCCESS' in status:
                        st_cell.font = STATUS_SUCCESS
                    elif 'FAILED' in status:
                        st_cell.font = STATUS_FAILED
                _set_cell(ws, row, 3, field_name, BODY_FONT, border=THIN_BORDER)
                val_cell = _set_cell(ws, row, 4, value,
                                     Font(name='Segoe UI', bold=True, size=11, color='1B5E20'),
                                     PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid'),
                                     border=THIN_BORDER)
                row += 1

    # ── Column widths ─────────────────────────────────────────────────────
    for col_idx in range(1, 9):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.column_dimensions['D'].width = 30  # New Value column wider
    ws.sheet_properties.tabColor = _NAVY


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 2/3/4: DATA SHEETS (All / Success / Failed)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _write_data_sheet(ws, accounts, columns, sheet_label='All'):
    """Write a data sheet with auto-detected columns, styled rows."""
    if not accounts:
        _set_cell(ws, 1, 1, f'No {sheet_label.lower()} accounts found.',
                  Font(name='Segoe UI', italic=True, size=11, color=_SUB_TEXT))
        return

    num_cols = len(columns)

    # ── Header row ────────────────────────────────────────────────────────
    for ci, header in enumerate(columns, 1):
        display = _DISPLAY_NAMES.get(header, header)
        c = _set_cell(ws, 1, ci, display, HEADER_FONT, HEADER_FILL, HEADER_ALIGN, THIN_BORDER)
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────
    for ri, account in enumerate(accounts, 2):
        status = _status_of(account)

        for ci, col in enumerate(columns, 1):
            raw_val = account.get(col, '')
            if raw_val is None:
                raw_val = ''
            # NaN handling
            val = str(raw_val)
            if val in ('nan', 'None', 'NaT'):
                val = ''

            cell = ws.cell(ri, ci, val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)

            # Status column colouring
            if col == 'Status':
                if 'SUCCESS' in status:
                    cell.font = STATUS_SUCCESS
                elif 'FAILED' in status:
                    cell.font = STATUS_FAILED
                elif 'SKIP' in status:
                    cell.font = STATUS_SKIP

            # Error Message — show context-aware translated message
            if col == 'Error Message' and val:
                translated = _translate_error(val, account)
                if translated['title']:
                    cell.value = f"{translated['title']} -- {translated['explanation']}"

            # Highlight RESULT columns (New Password, etc.) when they have values
            if col in _RESULT_COLUMNS and val:
                cell.font = _RESULT_FONT
                cell.fill = _RESULT_FILL

            # Make URLs clickable
            if val.startswith('http'):
                try:
                    cell.hyperlink = val
                    cell.font = LINK_FONT
                except Exception:
                    pass
            elif _URL_PAT.search(val):
                url = _URL_PAT.search(val).group(0).rstrip('|').strip()
                try:
                    cell.hyperlink = url
                    cell.font = LINK_FONT
                except Exception:
                    pass

        # Row colouring
        if 'SUCCESS' in status:
            _style_row(ws, ri, SUCCESS_ROW, num_cols)
        elif 'FAILED' in status:
            _style_row(ws, ri, FAILED_ROW, num_cols)
        elif 'SKIP' in status:
            _style_row(ws, ri, SKIP_ROW, num_cols)
        elif ri % 2 == 0:
            _style_row(ws, ri, ALT_ROW, num_cols)

    _auto_fit(ws)

    # Tab colour
    if sheet_label == 'Success':
        ws.sheet_properties.tabColor = _GREEN
    elif sheet_label == 'Failed':
        ws.sheet_properties.tabColor = _RED


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SHEET 5: ERROR ANALYSIS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _write_error_analysis(ws, accounts):
    """Detailed error analysis — grouped by error type with explanations."""
    failed = [a for a in accounts if 'FAILED' in _status_of(a)]

    # Title
    ws.merge_cells('A1:F1')
    _set_cell(ws, 1, 1, 'Error Analysis & Troubleshooting',
              Font(name='Segoe UI', bold=True, size=16, color=_RED),
              align=Alignment(horizontal='center', vertical='center'))
    ws.row_dimensions[1].height = 36

    if not failed:
        _set_cell(ws, 3, 1, 'No errors found — all accounts processed successfully!',
                  Font(name='Segoe UI', italic=True, size=12, color=_GREEN))
        ws.sheet_properties.tabColor = _GREEN
        return

    # Translate each account's error WITH its context
    per_account = []
    for a in failed:
        raw = str(a.get('Error Message', '')).strip()
        translated = _translate_error(raw, a)  # context-aware!
        per_account.append((a, translated))

    # Also group by title for the summary chart
    title_counts = {}
    for a, t in per_account:
        key = t['title'] or t['code'] or 'Unknown'
        title_counts[key] = title_counts.get(key, 0) + 1

    row = 3

    # ── Per-Account Error Details (most important — shows WHY each failed) ─
    _set_cell(ws, row, 1, 'DETAILED ERROR ANALYSIS', SECTION_FONT)
    row += 1

    detail_hdrs = ['Email', 'Error', 'Why It Failed', 'How To Fix']
    for ci, h in enumerate(detail_hdrs, 1):
        _set_cell(ws, row, ci, h, HEADER_FONT, HEADER_FILL, HEADER_ALIGN, THIN_BORDER)
    row += 1

    for a, info in per_account:
        _set_cell(ws, row, 1, str(a.get('Email', '')), BOLD_BODY, border=THIN_BORDER)

        _set_cell(ws, row, 2, info['title'] or info['code'], BODY_FONT,
                  PatternFill(start_color=_RED_BG, end_color=_RED_BG, fill_type='solid'),
                  border=THIN_BORDER)

        expl_cell = _set_cell(ws, row, 3, info['explanation'], BODY_FONT, border=THIN_BORDER)
        expl_cell.alignment = Alignment(wrap_text=True, vertical='center')

        fix_cell = _set_cell(ws, row, 4, info['suggestion'],
                             Font(name='Segoe UI', bold=True, size=10, color=_BLUE),
                             border=THIN_BORDER)
        fix_cell.alignment = Alignment(wrap_text=True, vertical='center')

        # Alternating row colour
        if row % 2 == 0:
            for ci in range(1, 5):
                if not ws.cell(row, ci).fill or ws.cell(row, ci).fill.start_color.rgb == '00000000':
                    ws.cell(row, ci).fill = PatternFill(
                        start_color='FFF5F5', end_color='FFF5F5', fill_type='solid')
        row += 1

    row += 1

    # ── Error Summary (grouped) ───────────────────────────────────────────
    _set_cell(ws, row, 1, 'ERROR SUMMARY', SECTION_FONT)
    row += 1
    summary_hdrs = ['Error Type', 'Count']
    for ci, h in enumerate(summary_hdrs, 1):
        _set_cell(ws, row, ci, h, HEADER_FONT, HEADER_FILL, HEADER_ALIGN, THIN_BORDER)
    row += 1

    chart_start = row
    sorted_titles = sorted(title_counts.items(), key=lambda x: -x[1])
    for title, count in sorted_titles:
        _set_cell(ws, row, 1, title, BOLD_BODY, border=THIN_BORDER)
        _set_cell(ws, row, 2, count, BODY_FONT,
                  PatternFill(start_color=_RED_BG, end_color=_RED_BG, fill_type='solid'),
                  Alignment(horizontal='center'), THIN_BORDER)
        row += 1

    # No chart — clean text table is enough

    _auto_fit(ws, min_w=14, max_w=60)
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 45
    ws.sheet_properties.tabColor = _RED


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MAIN ENTRY POINTS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def generate_report(output_dir, accounts_data, session_id=None, step_name=''):
    """
    Generate professional MailNexus Pro XLS report.

    Args:
        output_dir:    Path to output directory.
        accounts_data: list[dict] — each dict is one row from the Excel file.
        session_id:    Optional session ID (used to remove old reports).
        step_name:     Step identifier ('step1'-'step4') for column filtering.

    Returns:
        str: Path to generated report file.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    total = len(accounts_data)
    failed = sum(1 for a in accounts_data if 'FAILED' in _status_of(a))
    now = datetime.now()
    time_str = now.strftime('%H%M')
    date_str = now.strftime('%d_%m_%Y')

    filename = f"{total}_{failed}-Mailnexus-{time_str}-{date_str}.xlsx"
    filepath = output_dir / filename

    # Remove old MailNexus reports from same session
    if session_id:
        for old in output_dir.glob('*-Mailnexus-*.xlsx'):
            try:
                old.unlink()
            except Exception:
                pass

    # Auto-detect columns from actual data
    all_cols_set = set()
    for a in accounts_data:
        all_cols_set.update(a.keys())
    all_cols_list = list(all_cols_set)

    # Step-aware column selection — also filters out empty columns
    all_accounts_cols = _get_step_columns(all_cols_list, step_name, 'all', accounts_data)
    success_list_pre = [a for a in accounts_data if 'SUCCESS' in _status_of(a)]
    failed_list_pre  = [a for a in accounts_data if 'FAILED' in _status_of(a)]
    success_cols = _get_step_columns(all_cols_list, step_name, 'success', success_list_pre)
    failed_cols  = _get_step_columns(all_cols_list, step_name, 'failed', failed_list_pre)

    # Fallback: if step-specific filtering returned nothing, use full classify
    if not all_accounts_cols:
        all_accounts_cols = _classify_columns(all_cols_list, accounts_data)
    if not success_cols:
        success_cols = _classify_columns(all_cols_list, success_list_pre)
    if not failed_cols:
        failed_cols = all_accounts_cols

    wb = Workbook()

    # ── Sheet 1: Dashboard ────────────────────────────────────────────────
    ws_dash = wb.active
    ws_dash.title = 'Dashboard'
    _write_dashboard(ws_dash, accounts_data)

    # ── Sheet 2: All Accounts ─────────────────────────────────────────────
    ws_all = wb.create_sheet('All Accounts')
    _write_data_sheet(ws_all, accounts_data, all_accounts_cols, 'All')

    # ── Sheet 3: Successful ───────────────────────────────────────────────
    success_list = [a for a in accounts_data if 'SUCCESS' in _status_of(a)]
    # Overwrite original columns with effective values (updated if changed, else old)
    success_list = _apply_effective_values(success_list, step_name)
    ws_ok = wb.create_sheet('Successful')
    _write_data_sheet(ws_ok, success_list, success_cols, 'Success')

    # ── Sheet 4: Failed (step-specific minimal columns) ───────────────────
    failed_list = [a for a in accounts_data if 'FAILED' in _status_of(a)]
    ws_fail = wb.create_sheet('Failed')
    _write_data_sheet(ws_fail, failed_list, failed_cols, 'Failed')

    # ── Sheet 5: Error Analysis ───────────────────────────────────────────
    ws_err = wb.create_sheet('Error Analysis')
    _write_error_analysis(ws_err, accounts_data)

    # ── Sheet 6 & 7: Live / Missing (only when Live Check Status exists) ──
    has_live_check = any(
        str(a.get('Live Check Status', '')).strip()
        for a in accounts_data
        if str(a.get('Live Check Status', '')).strip()
        and str(a.get('Live Check Status', '')) not in ('nan', 'None', 'NaT')
    )
    if has_live_check:
        live_cols = [c for c in all_accounts_cols if c not in (
            'Error Message', 'New Authenticator Key', 'Processed At',
        )]
        if 'Live Check Status' not in live_cols:
            live_cols.append('Live Check Status')

        live_list = [
            a for a in accounts_data
            if str(a.get('Live Check Status', '')).strip().lower() == 'live'
        ]
        ws_live = wb.create_sheet('Live')
        _write_data_sheet(ws_live, live_list, live_cols, 'Live')
        ws_live.sheet_properties.tabColor = _GREEN

        missing_list = [
            a for a in accounts_data
            if str(a.get('Live Check Status', '')).strip().lower() != 'live'
            and str(a.get('Live Check Status', '')).strip()
            and str(a.get('Live Check Status', '')) not in ('nan', 'None', 'NaT')
        ]
        ws_missing = wb.create_sheet('Missing')
        _write_data_sheet(ws_missing, missing_list, live_cols, 'Missing')
        ws_missing.sheet_properties.tabColor = _RED

    try:
        wb.save(str(filepath))
    except PermissionError:
        # File likely open in Excel — save with alternate name
        alt = filepath.with_stem(filepath.stem + '_new')
        wb.save(str(alt))
        wb.close()
        print(f"[REPORT] Original file locked, saved as: {alt.name}")
        return str(alt)
    wb.close()

    print(f"[REPORT] MailNexus Pro report generated: {filepath}")
    return str(filepath)


def generate_partial_report(output_dir, accounts_data, session_id=None, step_name=''):
    """Generate partial report (crash recovery / mid-process)."""
    return generate_report(output_dir, accounts_data, session_id, step_name=step_name)


def generate_from_excel(excel_path, output_dir=None, step_name=''):
    """
    Convenience: generate a MailNexus report directly from any ExcelProcessor
    output file.  Called by Electron UI or command-line.

    Args:
        excel_path: Path to the Excel file with account data.
        output_dir: Where to save the report (defaults to same dir).
        step_name:  Step identifier ('step1'-'step4') for column filtering.

    Returns:
        str: Path to generated report.
    """
    import pandas as pd

    excel_path = Path(excel_path)
    if not excel_path.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    if output_dir is None:
        output_dir = str(excel_path.parent)

    # Auto-detect step from file path if not provided (e.g. .../step3/output/...)
    if not step_name:
        path_str = str(excel_path).replace('\\', '/').lower()
        for s in ('step1', 'step2', 'step3', 'step4'):
            if f'/{s}/' in path_str or f'\\{s}\\' in path_str:
                step_name = s
                break

    df = pd.read_excel(str(excel_path), engine='openpyxl')

    # Auto-extract Share Link from 'Operations Done' if not already a column
    if 'Share Link' not in df.columns and 'Operations Done' in df.columns:
        import re
        _url_re = re.compile(r'https?://[^\s|,]+')
        df['Share Link'] = df['Operations Done'].apply(
            lambda x: (m.group(0).rstrip('|').strip() if (m := _url_re.search(str(x))) else '')
        )

    # No (Current) columns needed — _apply_effective_values() in generate_report()
    # overwrites the original columns (Password, TOTP Secret, etc.) directly
    # on the success sheet with updated values if changed.

    accounts = df.to_dict('records')
    return generate_report(output_dir, accounts, session_id='manual', step_name=step_name)


def generate_template(step_name: str, output_dir: str = None) -> str:
    """Generate a blank XLS template with the required input columns for a step.

    Creates a styled Excel file with headers and one example row so users
    know exactly what data to put in the spreadsheet for that step.

    Args:
        step_name: 'step1', 'step2', 'step3', or 'step4'.
        output_dir: Where to save (defaults to ./templates/).

    Returns:
        str: Path to generated template file.
    """
    template_cols = _STEP_TEMPLATE_COLUMNS.get(step_name)
    if not template_cols:
        raise ValueError(f"Unknown step: {step_name}. Must be step1-step4.")

    if output_dir is None:
        output_dir = str(Path('templates'))
    out = Path(output_dir)
    out.mkdir(parents=True, exist_ok=True)

    step_num = step_name.replace('step', '')
    filepath = out / f'step{step_num}_template.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = f'Step {step_num} Template'

    # ── Header row (styled) ──────────────────────────────────────────────
    for col_idx, col_name in enumerate(template_cols, 1):
        display = _DISPLAY_NAMES.get(col_name, col_name)
        cell = ws.cell(1, col_idx, display)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    # ── Example row (row 2) ──────────────────────────────────────────────
    example_fill = PatternFill(start_color=_LIGHT_GREY, end_color=_LIGHT_GREY, fill_type='solid')
    example_font = Font(name='Segoe UI', italic=True, size=10, color=_SUB_TEXT)
    for col_idx, col_name in enumerate(template_cols, 1):
        val = _TEMPLATE_EXAMPLES.get(col_name, '')
        cell = ws.cell(2, col_idx, val)
        cell.font = example_font
        cell.fill = example_fill
        cell.border = THIN_BORDER

    # ── Freeze top row + auto-fit widths ─────────────────────────────────
    ws.freeze_panes = 'A2'
    _auto_fit(ws, min_w=12, max_w=40)

    # ── Auto-filter on header row ────────────────────────────────────────
    from openpyxl.utils import get_column_letter as _gcl
    last_col = _gcl(len(template_cols))
    ws.auto_filter.ref = f'A1:{last_col}1'

    wb.save(str(filepath))
    wb.close()

    print(f"[TEMPLATE] Step {step_num} template generated: {filepath}")
    return str(filepath)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# WRITE REVIEW REPORT
# Professional client-ready report for Google Review campaigns
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def generate_review_report(output_dir: str, results: list, campaign_name: str = '') -> str:
    """
    Generate a professional client-ready Excel report for a Write Review campaign.

    Args:
        output_dir:     Directory to save the report.
        results:        List of result dicts from _review_worker.
                        Each dict has: email, gmb_url, stars, review_text,
                        success, review_status, share_link, summary.
        campaign_name:  Optional label for the report title.

    Returns:
        str: Path to the saved .xlsx file.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    now = datetime.now()
    date_str  = now.strftime('%d_%m_%Y')
    time_str  = now.strftime('%H%M')
    total     = len(results)
    live      = sum(1 for r in results if str(r.get('review_status', '')).lower() in ('live', 'posted'))
    pending   = sum(1 for r in results if str(r.get('review_status', '')).lower() in ('pending',))
    failed    = sum(1 for r in results if not r.get('success'))
    success   = total - failed

    filename  = f"ReviewReport_{total}_{live}live-Mailnexus-{time_str}-{date_str}.xlsx"
    filepath  = output_dir / filename

    wb = Workbook()

    # ── helpers ──────────────────────────────────────────────────────────────
    def _hdr(ws, col, row, val, width=None):
        c = ws.cell(row, col, val)
        c.font = HEADER_FONT; c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGN; c.border = THIN_BORDER
        if width:
            ws.column_dimensions[get_column_letter(col)].width = width
        return c

    def _cell(ws, col, row, val, font=None, fill=None, align=None, hyperlink=None):
        c = ws.cell(row, col, val)
        c.font  = font  or BODY_FONT
        c.fill  = fill  or PatternFill()
        c.alignment = align or Alignment(vertical='center', wrap_text=False)
        c.border = THIN_BORDER
        if hyperlink and val:
            c.hyperlink = str(hyperlink)
            c.font = LINK_FONT
        return c

    def _stars_str(n):
        try: n = int(n)
        except Exception: n = 0
        return '★' * n + '☆' * (5 - n)

    def _status_fill(status):
        s = str(status).lower()
        if s in ('live', 'posted'):   return SUCCESS_ROW
        if s in ('pending',):         return PENDING_ROW
        if s in ('failed', 'error'):  return FAILED_ROW
        return PatternFill()

    def _status_font(status):
        s = str(status).lower()
        if s in ('live', 'posted'):  return STATUS_SUCCESS
        if s in ('failed', 'error'): return STATUS_FAILED
        return BODY_FONT

    CENTRE = Alignment(horizontal='center', vertical='center')
    WRAP   = Alignment(vertical='center', wrap_text=True)

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — DASHBOARD
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = 'Dashboard'
    ws.sheet_properties.tabColor = _NAVY
    ws.column_dimensions['A'].width = 2
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 22
    ws.row_dimensions[1].height  = 10
    ws.row_dimensions[2].height  = 46
    ws.row_dimensions[3].height  = 22
    ws.row_dimensions[4].height  = 18
    ws.row_dimensions[5].height  = 60

    # Title
    ws.merge_cells('B2:F2')
    title_cell = ws.cell(2, 2, campaign_name or 'Google Review Campaign Report')
    title_cell.font      = TITLE_FONT
    title_cell.alignment = Alignment(horizontal='left', vertical='center')

    ws.merge_cells('B3:F3')
    sub = ws.cell(3, 2, f'Generated: {now.strftime("%d %B %Y  %H:%M")}   ·   Total Accounts: {total}')
    sub.font = Font(name='Segoe UI', size=11, color=_SUB_TEXT)

    # Metric cards row (row 5)
    ws.row_dimensions[6].height = 50
    ws.row_dimensions[7].height = 24
    metrics = [
        ('TOTAL SENT',   total,   _NAVY,   _WHITE),
        ('✓ LIVE',        live,    _GREEN,  _WHITE),
        ('⏳ PENDING',    pending, _BLUE,   _WHITE),
        ('✗ FAILED',      failed,  _RED,    _WHITE),
    ]
    for i, (label, value, bg, fg) in enumerate(metrics):
        col = i + 2
        fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
        # Value
        vc = ws.cell(6, col, value)
        vc.font      = Font(name='Segoe UI', bold=True, size=28, color=fg)
        vc.fill      = fill
        vc.alignment = CENTRE
        vc.border    = THIN_BORDER
        # Label
        lc = ws.cell(7, col, label)
        lc.font      = Font(name='Segoe UI', bold=True, size=10, color=fg)
        lc.fill      = fill
        lc.alignment = CENTRE
        lc.border    = THIN_BORDER

    # Success rate bar (text)
    ws.row_dimensions[9].height = 28
    ws.merge_cells('B9:F9')
    rate = round(success / total * 100) if total else 0
    bar_filled  = int(rate / 5)   # 0–20 chars
    bar_empty   = 20 - bar_filled
    bar_str     = '█' * bar_filled + '░' * bar_empty
    rate_cell   = ws.cell(9, 2, f'  SUCCESS RATE  {bar_str}  {rate}%')
    rate_cell.font      = Font(name='Courier New', bold=True, size=11,
                                color=_WHITE if rate >= 50 else _RED)
    rate_cell.fill      = PatternFill(start_color=_NAVY, end_color=_NAVY, fill_type='solid')
    rate_cell.alignment = Alignment(vertical='center')
    rate_cell.border    = THIN_BORDER

    # Notes
    ws.row_dimensions[11].height = 20
    ws.merge_cells('B11:F11')
    note = ws.cell(11, 2, 'REVIEW STATUS LEGEND:   ✓ Live = Confirmed live on Google Maps   '
                            '⏳ Pending = Posted, awaiting approval   ✗ Failed = Error / not posted')
    note.font      = Font(name='Segoe UI', size=9, color=_SUB_TEXT, italic=True)
    note.alignment = Alignment(vertical='center')

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — ALL REVIEWS
    # ══════════════════════════════════════════════════════════════════════════
    ws_all = wb.create_sheet('All Reviews')
    ws_all.sheet_properties.tabColor = '6366F1'
    ws_all.freeze_panes = 'A2'

    cols_all = ['#', 'Email', 'Stars', 'Review Text', 'Review Status', 'Share Link', 'Summary / Error']
    widths   = [5,   28,       10,      40,             16,              50,           40]
    for ci, (h, w) in enumerate(zip(cols_all, widths), 1):
        _hdr(ws_all, ci, 1, h, w)

    for ri, r in enumerate(results, 2):
        status = str(r.get('review_status', 'unknown')).lower()
        row_fill = _status_fill(status)
        vals = [
            ri - 1,
            r.get('email', ''),
            _stars_str(r.get('stars', 5)),
            (r.get('review_text', '') or '')[:200],
            str(r.get('review_status', '')).upper(),
            r.get('share_link', ''),
            r.get('summary', ''),
        ]
        for ci, val in enumerate(vals, 1):
            al = CENTRE if ci in (1, 3, 5) else Alignment(vertical='center', wrap_text=(ci == 4))
            is_link = (ci == 6)
            _cell(ws_all, ci, ri, val,
                  font=_status_font(status) if ci == 5 else (LINK_FONT if is_link and val else BODY_FONT),
                  fill=row_fill,
                  align=al,
                  hyperlink=val if is_link and val else None)
        ws_all.row_dimensions[ri].height = 20

    last_col_letter = get_column_letter(len(cols_all))
    ws_all.auto_filter.ref = f'A1:{last_col_letter}1'

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — LIVE REVIEWS  (client proof sheet)
    # ══════════════════════════════════════════════════════════════════════════
    live_results = [r for r in results if str(r.get('review_status', '')).lower() in ('live', 'posted')]
    ws_live = wb.create_sheet('Live Reviews ✓')
    ws_live.sheet_properties.tabColor = _GREEN
    ws_live.freeze_panes = 'A2'

    cols_live = ['#', 'Email', 'Stars', 'Review Text', 'Share Link (Click to Verify)']
    widths_l  = [5,   28,       10,      45,             55]
    for ci, (h, w) in enumerate(zip(cols_live, widths_l), 1):
        _hdr(ws_live, ci, 1, h, w)

    for ri, r in enumerate(live_results, 2):
        vals = [
            ri - 1,
            r.get('email', ''),
            _stars_str(r.get('stars', 5)),
            (r.get('review_text', '') or '')[:200],
            r.get('share_link', '') or 'View on Google Maps',
        ]
        link = r.get('share_link', '')
        for ci, val in enumerate(vals, 1):
            al = CENTRE if ci in (1, 3) else Alignment(vertical='center', wrap_text=(ci == 4))
            _cell(ws_live, ci, ri, val,
                  fill=SUCCESS_ROW,
                  align=al,
                  hyperlink=link if ci == 5 and link else None)
        ws_live.row_dimensions[ri].height = 20

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — PENDING
    # ══════════════════════════════════════════════════════════════════════════
    pending_results = [r for r in results if str(r.get('review_status', '')).lower() == 'pending']
    if pending_results:
        ws_pend = wb.create_sheet('Pending ⏳')
        ws_pend.sheet_properties.tabColor = _BLUE
        ws_pend.freeze_panes = 'A2'
        cols_p = ['#', 'Email', 'Stars', 'Review Text', 'Note']
        widths_p = [5, 28, 10, 45, 40]
        for ci, (h, w) in enumerate(zip(cols_p, widths_p), 1):
            _hdr(ws_pend, ci, 1, h, w)
        for ri, r in enumerate(pending_results, 2):
            vals = [ri - 1, r.get('email', ''), _stars_str(r.get('stars', 5)),
                    (r.get('review_text', '') or '')[:200],
                    'Posted — awaiting Google approval (usually 1-3 days)']
            for ci, val in enumerate(vals, 1):
                al = CENTRE if ci in (1, 3) else Alignment(vertical='center', wrap_text=(ci == 4))
                _cell(ws_pend, ci, ri, val, fill=PENDING_ROW, align=al)
            ws_pend.row_dimensions[ri].height = 20

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 5 — FAILED
    # ══════════════════════════════════════════════════════════════════════════
    failed_results = [r for r in results if not r.get('success')]
    if failed_results:
        ws_fail = wb.create_sheet('Failed ✗')
        ws_fail.sheet_properties.tabColor = _RED
        ws_fail.freeze_panes = 'A2'
        cols_f = ['#', 'Email', 'Stars', 'Error / Reason']
        widths_f = [5, 28, 10, 60]
        for ci, (h, w) in enumerate(zip(cols_f, widths_f), 1):
            _hdr(ws_fail, ci, 1, h, w)
        for ri, r in enumerate(failed_results, 2):
            vals = [ri - 1, r.get('email', ''), _stars_str(r.get('stars', 5)),
                    r.get('summary', 'Unknown error')]
            for ci, val in enumerate(vals, 1):
                al = CENTRE if ci in (1, 3) else Alignment(vertical='center', wrap_text=(ci == 4))
                _cell(ws_fail, ci, ri, val, fill=FAILED_ROW, align=al)
            ws_fail.row_dimensions[ri].height = 20

    # ── Save ─────────────────────────────────────────────────────────────────
    try:
        wb.save(str(filepath))
    except PermissionError:
        alt = filepath.with_stem(filepath.stem + '_new')
        wb.save(str(alt))
        wb.close()
        print(f"[REVIEW REPORT] File locked, saved as: {alt.name}")
        return str(alt)
    wb.close()

    print(f"[REVIEW REPORT] Report saved: {filepath}")
    return str(filepath)
