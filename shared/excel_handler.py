"""
Shared Excel handler: thread-safe row locking and Excel read/write.
Used by both Step 1 and Step 2 workers.

Performance: accounts are loaded into memory ONCE at init.
get_next_account() uses the in-memory cache + row locking —
no disk I/O on the hot path.  Only update_row_status() writes to disk.
"""

import threading
import time
import shutil
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from shared.logger import print


class ExcelRowLock:
    """Thread-safe row locking mechanism."""

    def __init__(self):
        self.locked_rows = set()
        self.processed_rows = set()  # Track ALL processed rows to prevent retry
        self.lock = threading.Lock()

    def try_lock_row(self, row_index):
        """Try to lock a row. Returns True if successful, False if already locked or processed."""
        with self.lock:
            if row_index in self.processed_rows:
                return False
            if row_index in self.locked_rows:
                return False
            self.locked_rows.add(row_index)
            return True

    def unlock_row(self, row_index):
        """Unlock a row when processing is complete."""
        with self.lock:
            if row_index in self.locked_rows:
                self.locked_rows.remove(row_index)

    def mark_processed(self, row_index):
        """Mark a row as fully processed (success or failed) to prevent retry."""
        with self.lock:
            self.processed_rows.add(row_index)
            if row_index in self.locked_rows:
                self.locked_rows.remove(row_index)


class ExcelProcessor:
    """Handles Excel file reading and writing with row locking.

    Accounts are loaded into memory once.  get_next_account() is O(n) over the
    cached list (fast — no disk I/O).  Only update_row_status() touches disk.
    """

    def __init__(self, excel_file_path):
        self.input_excel_file = excel_file_path  # Original input file (read-only)
        self.lock = threading.Lock()
        self.row_lock = ExcelRowLock()

        # Create output folder
        self.base_dir = Path(excel_file_path).parent.parent
        self.output_dir = self.base_dir / "output"
        self.output_dir.mkdir(exist_ok=True)

        # Create output Excel file (copy of input with timestamp)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        input_filename = Path(excel_file_path).stem
        output_filename = f"{input_filename}_output_{timestamp}.xlsx"
        self.excel_file = self.output_dir / output_filename

        shutil.copy2(excel_file_path, self.excel_file)
        print(f"[SYSTEM] Created output Excel file: {self.excel_file}")
        print(f"[SYSTEM] Input file will remain unchanged: {excel_file_path}")

        # Real-time progress tracking
        self.progress_lock = threading.Lock()
        self.total_accounts = 0
        self.success_count = 0
        self.failed_count = 0
        self.processed_count = 0

        # ── In-memory account cache (loaded once) ──────────────────────────
        self._cached_accounts = None   # list[dict] — populated on first access
        self._cache_lock = threading.Lock()

    # ──────────────────────────────────────────────────────────────────────
    # Account cache
    # ──────────────────────────────────────────────────────────────────────

    def _ensure_cache(self):
        """Load accounts from Excel into memory (once). Thread-safe."""
        if self._cached_accounts is not None:
            return

        with self._cache_lock:
            # Double-check after acquiring lock
            if self._cached_accounts is not None:
                return

            max_retries = 5
            retry_delay = 2
            for attempt in range(max_retries):
                try:
                    # dtype=str prevents empty columns being read as float64,
                    # which would give workers NaN instead of '' for blank cells
                    # (e.g. typing "nan" into a password field).
                    df = pd.read_excel(self.excel_file, engine='openpyxl', dtype=str)
                    df.fillna('', inplace=True)

                    if 'Status' not in df.columns:
                        df['Status'] = 'Pending'

                    pending = df[
                        (df['Status'] == '') |
                        (df['Status'].str.upper() == 'PENDING')
                    ]

                    accounts = []
                    for idx, row in pending.iterrows():
                        record = row.to_dict()
                        record['_original_index'] = idx
                        record['row_index'] = idx + 2   # +2: Excel 1-indexed + header
                        accounts.append(record)

                    self._cached_accounts = accounts
                    print(f"[EXCEL] Cached {len(accounts)} pending accounts into memory")
                    return

                except (PermissionError, OSError) as e:
                    if attempt < max_retries - 1:
                        print(f"[EXCEL] File access error (attempt {attempt + 1}/{max_retries}): {e}")
                        time.sleep(retry_delay)
                    else:
                        print(f"[EXCEL] CRITICAL: Cannot read Excel file after {max_retries} attempts")
                        raise Exception(f"Excel file access failed: {e}")

    # ──────────────────────────────────────────────────────────────────────
    # Public API — used by workers
    # ──────────────────────────────────────────────────────────────────────

    def read_pending_accounts(self):
        """Read all accounts that haven't been processed yet.

        Returns (list[dict], DataFrame) for backward compatibility with
        report generators that expect the full DataFrame.
        """
        with self.lock:
            max_retries = 5
            retry_delay = 2

            for attempt in range(max_retries):
                try:
                    df = pd.read_excel(self.excel_file, engine='openpyxl', dtype=str)
                    df.fillna('', inplace=True)

                    if 'Status' not in df.columns:
                        df['Status'] = 'Pending'

                    pending = df[
                        (df['Status'] == '') |
                        (df['Status'].str.upper() == 'PENDING')
                    ]

                    pending = pending.copy()
                    pending['_original_index'] = pending.index

                    return pending.to_dict('records'), df

                except (PermissionError, OSError) as e:
                    if attempt < max_retries - 1:
                        print(f"[EXCEL] File access error (attempt {attempt + 1}/{max_retries}): {e}")
                        time.sleep(retry_delay)
                    else:
                        print(f"[EXCEL] CRITICAL: Cannot read Excel file after {max_retries} attempts")
                        raise Exception(f"Excel file access failed: {e}")

    def get_next_account(self):
        """Get next available account and lock its row.

        Uses the in-memory cache — NO disk I/O on the hot path.
        """
        self._ensure_cache()

        for account in self._cached_accounts:
            row_index = account['row_index']

            if self.row_lock.try_lock_row(row_index):
                email = account.get('Email', 'unknown')
                print(f"[EXCEL] Locked row {row_index} for {email}")
                return account

        return None

    def update_progress(self, status):
        """Update real-time progress counters."""
        with self.progress_lock:
            self.processed_count += 1
            if status == 'SUCCESS':
                self.success_count += 1
            else:
                self.failed_count += 1

            pending = self.total_accounts - self.processed_count
            progress_pct = (self.processed_count / self.total_accounts * 100) if self.total_accounts > 0 else 0
            print(f"\n{'='*70}")
            print(f"PROGRESS: {self.processed_count}/{self.total_accounts} ({progress_pct:.1f}%)")
            print(f"[+] Success: {self.success_count}  |  [-] Failed: {self.failed_count}  |  [~] Pending: {pending}")
            print(f"{'='*70}\n", flush=True)

    def update_row(self, row_index, status, error=''):
        """Simple wrapper for early validation failures."""
        self.update_row_status(row_index=row_index, status=status, error_message=error)

    def update_row_status(self, row_index, status, operations_done='', error_message='',
                          screenshots_folder='', authenticator_key='', backup_codes='',
                          op1_status='', op2_status='', op3_status='', op4_status='',
                          op5_status='', op6_status='', op7_status='', op8_status='',
                          live_check_status='',
                          **extra_kwargs):
        """Update Excel row with results (thread-safe with retry)."""
        with self.lock:
            max_retries = 5
            retry_delay = 2

            for attempt in range(max_retries):
                try:
                    wb = load_workbook(self.excel_file)
                    ws = wb.active

                    headers = [cell.value for cell in ws[1]]
                    status_col       = headers.index('Status') + 1             if 'Status'                in headers else len(headers) + 1
                    operations_col   = headers.index('Operations Done') + 1    if 'Operations Done'       in headers else len(headers) + 2
                    error_col        = headers.index('Error Message') + 1      if 'Error Message'         in headers else len(headers) + 3
                    auth_key_col     = headers.index('New Authenticator Key') + 1 if 'New Authenticator Key' in headers else len(headers) + 4
                    # Backup Code 1-10 (10 separate columns for individual codes)
                    backup_code_cols = []
                    for bc_i in range(1, 11):
                        bc_name = f'Backup Code {bc_i}'
                        backup_code_cols.append(
                            headers.index(bc_name) + 1 if bc_name in headers else len(headers) + 4 + bc_i
                        )
                    timestamp_col    = headers.index('Processed At') + 1       if 'Processed At'          in headers else len(headers) + 15

                    op1_col = headers.index('Op1: Change Password') + 1  if 'Op1: Change Password' in headers else len(headers) + 16
                    op2_col = headers.index('Op2: Recovery Phone') + 1   if 'Op2: Recovery Phone'  in headers else len(headers) + 17
                    op3_col = headers.index('Op3: Recovery Email') + 1   if 'Op3: Recovery Email'  in headers else len(headers) + 18
                    op4_col = headers.index('Op4: Authenticator') + 1    if 'Op4: Authenticator'   in headers else len(headers) + 19
                    op5_col = headers.index('Op5: Backup Codes') + 1     if 'Op5: Backup Codes'    in headers else len(headers) + 20
                    op6_col = headers.index('Op6: 2FA Phone') + 1        if 'Op6: 2FA Phone'       in headers else len(headers) + 21
                    op7_col = headers.index('Op7: Remove Devices') + 1   if 'Op7: Remove Devices'  in headers else len(headers) + 22
                    op8_col = headers.index('Op8: Change Name') + 1      if 'Op8: Change Name'     in headers else len(headers) + 23
                    live_check_col = headers.index('Live Check Status') + 1 if 'Live Check Status'   in headers else len(headers) + 24

                    # Add headers if missing
                    for bc_i in range(10):
                        bc_name = f'Backup Code {bc_i + 1}'
                        if bc_name not in headers:
                            ws.cell(1, backup_code_cols[bc_i], bc_name)

                    for col_idx, col_name in [
                        (status_col, 'Status'), (operations_col, 'Operations Done'),
                        (error_col, 'Error Message'),
                        (auth_key_col, 'New Authenticator Key'),
                        (timestamp_col, 'Processed At'), (op1_col, 'Op1: Change Password'),
                        (op2_col, 'Op2: Recovery Phone'), (op3_col, 'Op3: Recovery Email'),
                        (op4_col, 'Op4: Authenticator'), (op5_col, 'Op5: Backup Codes'),
                        (op6_col, 'Op6: 2FA Phone'), (op7_col, 'Op7: Remove Devices'),
                        (op8_col, 'Op8: Change Name'),
                        (live_check_col, 'Live Check Status'),
                    ]:
                        if col_name not in headers:
                            ws.cell(1, col_idx, col_name)

                    # Update row data
                    ws.cell(row_index, status_col,       status)
                    ws.cell(row_index, operations_col,   operations_done)
                    ws.cell(row_index, error_col,        error_message)

                    # Only overwrite authenticator key when a NEW key is provided
                    if authenticator_key:
                        ws.cell(row_index, auth_key_col, authenticator_key)

                    # Write backup codes to 10 separate columns
                    if backup_codes:
                        codes_list = [c.strip() for c in backup_codes.split(',') if c.strip()]
                        for bc_i in range(10):
                            ws.cell(row_index, backup_code_cols[bc_i],
                                    codes_list[bc_i] if bc_i < len(codes_list) else '')

                    ws.cell(row_index, timestamp_col,    datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

                    for col, val in [
                        (op1_col, op1_status), (op2_col, op2_status),
                        (op3_col, op3_status), (op4_col, op4_status),
                        (op5_col, op5_status), (op6_col, op6_status),
                        (op7_col, op7_status), (op8_col, op8_status),
                    ]:
                        if val:
                            ws.cell(row_index, col, val)

                    # Write Live Check Status
                    if live_check_status:
                        ws.cell(row_index, live_check_col, live_check_status)

                    fill = (
                        PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                        if status == 'SUCCESS'
                        else PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    )
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row_index, col).fill = fill

                    wb.save(self.excel_file)
                    wb.close()

                    print(f"[EXCEL] Updated row {row_index}: {status}")
                    self.update_progress(status)
                    break

                except (PermissionError, OSError) as e:
                    if attempt < max_retries - 1:
                        print(f"[EXCEL] File access error (attempt {attempt + 1}/{max_retries}): {e}")
                        time.sleep(retry_delay)
                    else:
                        print(f"[ERROR] Failed to update Excel row {row_index} after {max_retries} attempts: {e}")
                        self.row_lock.unlock_row(row_index)
                        return  # Don't mark processed — allow retry
                except Exception as e:
                    print(f"[ERROR] Unexpected error updating row {row_index}: {e}")
                    import traceback
                    traceback.print_exc()
                    self.row_lock.unlock_row(row_index)
                    return

            self.row_lock.mark_processed(row_index)
