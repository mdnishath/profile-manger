"""
Shared worker runner.
Handles threading, Excel setup, worker loop, and report generation.
Both Step 1 and Step 2 delegate to start_production_processing() here.
"""

import asyncio
import re
import threading
import time
import traceback
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

from shared.logger import print
from shared.excel_handler import ExcelProcessor
from shared import proxy_manager, fingerprint_manager


# ── Hyperlink helpers ─────────────────────────────────────────────────────────

_URL_PAT = re.compile(r'https?://[^\s|,]+')
_LINK_FONT = Font(color='0563C1', underline='single')


def _extract_url(text) -> str:
    """Extract the first URL from a cell string (handles 'Link: https://...')."""
    if not text or not isinstance(text, str):
        return ''
    m = _URL_PAT.search(text)
    return m.group(0).rstrip('|').strip() if m else ''


def _apply_hyperlinks(filepath):
    """
    Post-process an Excel file so that any URL-containing cell becomes a
    clickable hyperlink (blue, underlined).  Handles:
      - cells whose value IS a URL
      - cells whose value CONTAINS a URL (e.g. 'Share Link' column)
    """
    wb = None
    try:
        wb = load_workbook(filepath)
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if not cell.value or not isinstance(cell.value, str):
                        continue
                    val = cell.value.strip()
                    url = val if val.startswith('http') else _extract_url(val)
                    if url:
                        cell.hyperlink = url
                        cell.font = _LINK_FONT
        wb.save(filepath)
    except Exception as e:
        print(f"[REPORT] Hyperlink post-process failed for {filepath}: {e}")
    finally:
        if wb:
            wb.close()


async def worker_main(worker_id, excel_processor, stop_event, worker_class,
                      per_account_timeout=600):
    """Main async loop for one worker.

    Args:
        per_account_timeout: Max seconds per account before giving up (default 10 min).
    """
    print(f"[WORKER {worker_id}] Worker starting...")
    worker = worker_class(worker_id, excel_processor)

    # ── Batch mode: if worker supports it and operations match, use batch ──
    if hasattr(worker, 'process_batch_live_check') and hasattr(worker, 'should_use_batch'):
        if worker.should_use_batch(excel_processor):
            print(f"[WORKER {worker_id}] Using BATCH mode (A3 Live Check)")
            await worker.process_batch_live_check(excel_processor, stop_event)
            print(f"[WORKER {worker_id}] Batch mode complete. Stopped.")
            return

    empty_retries = 0
    max_empty_retries = 6  # 6 retries × 5s = 30s max wait for locked accounts

    while not stop_event.is_set():
        print(f"[WORKER {worker_id}] Getting next account...")
        account = excel_processor.get_next_account()

        if account is None:
            empty_retries += 1
            if empty_retries > max_empty_retries:
                print(f"[WORKER {worker_id}] No more work after {max_empty_retries} retries. Shutting down.")
                break
            print(f"[WORKER {worker_id}] No accounts available (retry {empty_retries}/{max_empty_retries}). Waiting 5s...")
            await asyncio.sleep(5)
            continue

        empty_retries = 0  # Reset on successful account pick

        try:
            await asyncio.wait_for(
                worker.process_account(account),
                timeout=per_account_timeout,
            )
        except asyncio.TimeoutError:
            email = account.get('Email', 'unknown')
            print(f"[WORKER {worker_id}] TIMEOUT: {email} exceeded {per_account_timeout}s — skipping")
            row_index = account.get('row_index')
            if row_index is not None:
                excel_processor.update_row(row_index, 'FAILED',
                                           error=f'Timeout after {per_account_timeout}s')
        except Exception as e:
            print(f"[WORKER {worker_id}] Critical error: {e}")
            traceback.print_exc()
            await asyncio.sleep(5)

        await asyncio.sleep(2)

    print(f"[WORKER {worker_id}] Stopped.")


def run_worker(worker_id, excel_processor, stop_event, worker_class):
    """Run worker_main in a dedicated thread via asyncio.run()."""
    print(f"[WORKER {worker_id}] run_worker() called", flush=True)
    print(f"[WORKER {worker_id}] Starting asyncio.run()", flush=True)
    asyncio.run(worker_main(worker_id, excel_processor, stop_event, worker_class))


def generate_reports(excel_processor, step_name=''):
    """Print a console summary of processing results.

    The actual Excel report (MailNexus Pro) is generated separately by
    report_generator.generate_from_excel(), which is called from
    start_production_processing() below.
    """
    print("\n" + "="*70)
    print("FINAL SUMMARY")
    print("="*70)

    try:
        df = pd.read_excel(excel_processor.excel_file)

        total = len(df)
        success_df = df[df['Status'] == 'SUCCESS']
        failed_df  = df[df['Status'] == 'FAILED']
        success = len(success_df)
        failed  = len(failed_df)
        pending = total - success - failed

        print(f"\nProcessing Results:")
        print(f"   Total Accounts: {total}")
        print(f"   [+] SUCCESS: {success} ({success*100//total if total > 0 else 0}%)")
        print(f"   [-] FAILED:  {failed}  ({failed*100//total if total > 0 else 0}%)")
        print(f"   [~] PENDING: {pending} ({pending*100//total if total > 0 else 0}%)")

        if failed > 0:
            print(f"\n[-] Failed Accounts:")
            for _, row in failed_df.iterrows():
                err = str(row.get('Error Message', ''))[:80] if pd.notna(row.get('Error Message')) else 'Unknown error'
                print(f"   - {row['Email']}: {err}")

        if success > 0:
            print(f"\n[+] Successful Accounts:")
            for _, row in success_df.iterrows():
                ops = str(row.get('Operations Done', '')) if pd.notna(row.get('Operations Done')) else 'None'
                print(f"   + {row['Email']}: {ops}")

        print(f"\nMain File: {excel_processor.excel_file}")
        print(f"Reports: output/ folder")

    except Exception as e:
        print(f"\n[ERROR] Could not generate summary: {e}")
        traceback.print_exc()

    print("\n" + "="*70)


def start_production_processing(excel_file, num_workers=10, worker_class=None, step_banner='GMAIL BOT', step_name=''):
    """
    Start multi-worker processing.

    Args:
        excel_file:   Path to input Excel file.
        num_workers:  Number of parallel browser workers.
        worker_class: GmailBotWorker class (from step1.runner or step2.runner).
        step_banner:  Banner text printed at startup.
        step_name:    Step identifier ('step1', 'step2', 'step3', 'step4') for reports.
    """
    print("="*70)
    print(step_banner)
    print("="*70)
    print(f"Excel File: {excel_file}")
    print(f"Workers:    {num_workers}")
    print("="*70)

    if not Path(excel_file).exists():
        print(f"[ERROR] Excel file not found: {excel_file}")
        return

    # ── Proxy setup ───────────────────────────────────────────────────────────
    proxy_manager.load()
    print(f"[PROXY] Loaded {proxy_manager.proxy_count()} proxies")

    # Run health check on all proxies before starting workers
    if proxy_manager.is_enabled() and proxy_manager.proxy_count() > 0:
        proxy_manager.run_health_check(max_workers=50, timeout=8.0)
        proxy_manager.start_auto_refresh()
        stats = proxy_manager.get_health_stats()
        print(f"[PROXY] Health: {stats['alive']} alive / {stats['dead']} dead / {stats['total']} total")
        if stats['alive'] == 0:
            print(f"[PROXY] WARNING: All proxies are dead! Workers will use local IP.")

    proxy_manager.assign(num_workers)
    print(f"[PROXY] {proxy_manager.summary()}")
    if proxy_manager.is_enabled() and proxy_manager.proxy_count() > 0:
        p_count  = proxy_manager.proxy_count()
        assigned = min(p_count, num_workers)
        local_ip = num_workers - assigned
        print(f"[PROXY] Assignment: Workers 1-{assigned} → unique proxy each"
              + (f" | Workers {assigned+1}-{num_workers} → Local IP (no proxy available)" if local_ip > 0 else ""))

    # ── Fingerprint setup ─────────────────────────────────────────────────────
    fingerprint_manager.load()
    fingerprint_manager.assign(num_workers)
    print(f"[FINGERPRINT] {fingerprint_manager.summary()}")

    stop_event = threading.Event()
    excel_processor = ExcelProcessor(excel_file)
    print(f"[SYSTEM] Created shared ExcelProcessor with shared row locking")
    print(f"[OUTPUT_FILE] {excel_processor.excel_file}")

    pending_accounts, _ = excel_processor.read_pending_accounts()
    excel_processor.total_accounts = len(pending_accounts)
    print(f"[SYSTEM] Total pending accounts: {excel_processor.total_accounts}")
    print(f"\n{'='*70}")
    print(f"STARTING PROCESSING: 0/{excel_processor.total_accounts} (0.0%)")
    print(f"[+] Success: 0  |  [-] Failed: 0  |  [~] Pending: {excel_processor.total_accounts}")
    print(f"{'='*70}\n")

    threads = []
    for worker_id in range(1, num_workers + 1):
        thread = threading.Thread(
            target=run_worker,
            args=(worker_id, excel_processor, stop_event, worker_class),
            daemon=True
        )
        thread.start()
        threads.append(thread)
        print(f"[SYSTEM] Started Worker {worker_id}")
        time.sleep(1)  # Stagger starts to avoid thundering herd

    print(f"\n[SYSTEM] All {num_workers} workers started!")
    print("[SYSTEM] Press Ctrl+C to stop all workers...")

    try:
        for thread in threads:
            thread.join()
    except KeyboardInterrupt:
        print("\n[SYSTEM] Stopping all workers...")
        stop_event.set()
        for thread in threads:
            thread.join(timeout=10)
        # Generate partial report on crash/stop
        print("[SYSTEM] Generating partial report before exit...")
        try:
            from shared.report_generator import generate_from_excel
            output_dir = Path("output")
            output_dir.mkdir(exist_ok=True)
            report_path = generate_from_excel(excel_processor.excel_file, str(output_dir), step_name=step_name)
            print(f"[REPORT] Partial MailNexus Pro report saved: {report_path}")
        except Exception as partial_err:
            print(f"[REPORT] Partial report failed: {partial_err}")
    except Exception as unexpected_err:
        print(f"\n[SYSTEM] Unexpected error: {unexpected_err}")
        stop_event.set()
        for thread in threads:
            thread.join(timeout=10)
        # Generate partial report on unexpected crash
        print("[SYSTEM] Generating partial report on crash...")
        try:
            from shared.report_generator import generate_from_excel
            output_dir = Path("output")
            output_dir.mkdir(exist_ok=True)
            report_path = generate_from_excel(excel_processor.excel_file, str(output_dir), step_name=step_name)
            print(f"[REPORT] Crash recovery report saved: {report_path}")
        except Exception as crash_err:
            print(f"[REPORT] Crash report failed: {crash_err}")

    print("\n[SYSTEM] All workers stopped. Processing complete!")
    generate_reports(excel_processor, step_name=step_name)

    # ── MailNexus Pro Report ──────────────────────────────────────────
    try:
        from shared.report_generator import generate_from_excel
        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)
        report_path = generate_from_excel(excel_processor.excel_file, str(output_dir), step_name=step_name)
        print(f"[REPORT] MailNexus Pro report: {report_path}")
    except Exception as report_err:
        print(f"[REPORT] MailNexus Pro report generation failed: {report_err}")
        import traceback
        traceback.print_exc()
