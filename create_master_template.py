"""Create master Excel template for MailNexus Pro."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

# ── Color definitions ────────────────────────────────────────────────
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill_common = PatternFill('solid', fgColor='2C3E50')
header_fill_step2  = PatternFill('solid', fgColor='8E44AD')
header_fill_step3  = PatternFill('solid', fgColor='27AE60')
header_fill_step4  = PatternFill('solid', fgColor='E67E22')
example_font = Font(name='Calibri', size=10, color='666666', italic=True)
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ── Headers ──────────────────────────────────────────────────────────
backup_code_headers = [f'Backup Code {i}' for i in range(1, 11)]
common_headers = ['First Name', 'Email', 'Password', 'TOTP Secret', 'Recovery Email', 'Recovery Phone', 'Year'] + backup_code_headers + ['Operations']
step2_headers = ['New Password', 'New Recovery Phone', 'New Recovery Email', 'New 2FA Phone', 'Last Name']
step3_headers = ['GMB Name', 'GMB URL', 'Review Text', 'Review Stars']
step4_headers = ['Appeal Message']
all_headers = common_headers + step2_headers + step3_headers + step4_headers

# ── Write headers ────────────────────────────────────────────────────
ws = wb.active
ws.title = 'Accounts'

for col_idx, header in enumerate(all_headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = header_font
    cell.alignment = center
    cell.border = thin_border
    if header in common_headers:
        cell.fill = header_fill_common
    elif header in step2_headers:
        cell.fill = header_fill_step2
    elif header in step3_headers:
        cell.fill = header_fill_step3
    elif header in step4_headers:
        cell.fill = header_fill_step4

# ── Example rows ─────────────────────────────────────────────────────
examples = [
    {
        'Email': 'user1@gmail.com', 'Password': 'MyPass123!',
        'TOTP Secret': 'JBSWY3DPEHPK3PXP',
        'Backup Code 1': '1234 5678', 'Backup Code 2': '2345 6789',
        'Backup Code 3': '3456 7890', 'Backup Code 4': '4567 8901',
        'Backup Code 5': '5678 9012', 'Backup Code 6': '6789 0123',
        'Backup Code 7': '7890 1234', 'Backup Code 8': '8901 2345',
        'Backup Code 9': '9012 3456', 'Backup Code 10': '0123 4567',
        'Operations': 'L1,L2',
    },
    {
        'Email': 'user2@gmail.com', 'Password': 'OldPass456!',
        'TOTP Secret': 'KRSXG5CTMVRXEZLUKN',
        'Recovery Email': 'old_recovery@yahoo.com', 'Recovery Phone': '+8801700000000',
        'Year': 2018,
        'Backup Code 1': '9876 5432', 'Backup Code 2': '8765 4321',
        'Operations': '1,2a,3a,4a,5a,7,8',
        'New Password': 'NewSecure789!', 'New Recovery Phone': '+8801712345678',
        'New Recovery Email': 'recovery@outlook.com', 'New 2FA Phone': '+8801798765432',
        'First Name': 'John', 'Last Name': 'Doe',
    },
    {
        'Email': 'user3@gmail.com', 'Password': 'Pass789!',
        'Recovery Email': 'backup@gmail.com', 'Year': 2020,
        'Backup Code 1': '5555 6666', 'Operations': '2b,3b,4b,5b,6b',
    },
    {
        'Email': 'user4@gmail.com', 'Password': 'ReviewPass!',
        'TOTP Secret': 'MFZWIZB2MFZWIZB2', 'Operations': 'R3',
        'GMB Name': 'Example Business',
        'GMB URL': 'https://www.google.com/maps/place/Example+Business/@23.8,90.4',
        'Review Text': 'Great service and friendly staff!', 'Review Stars': 5,
    },
    {
        'Email': 'user5@gmail.com', 'Password': 'StarOnly!',
        'Backup Code 1': '1111 2222', 'Operations': 'R3',
        'GMB Name': 'Another Place',
        'GMB URL': 'https://www.google.com/maps/place/Another+Place/@23.8,90.4',
        'Review Stars': 4,
    },
    {
        'Email': 'user6@gmail.com', 'Password': 'MultiOp!',
        'TOTP Secret': 'GEZDGNBVGY3TQOJQ', 'Operations': 'R1,R3,R4',
        'GMB Name': 'Third Place Restaurant',
        'GMB URL': 'https://www.google.com/maps/place/Third+Place/@23.8,90.4',
        'Review Text': 'Amazing food!', 'Review Stars': 5,
    },
    {
        'Email': 'user7@gmail.com', 'Password': 'AppealPass!',
        'Backup Code 1': '3333 4444', 'Operations': 'A1',
        'Appeal Message': 'My account was incorrectly flagged. Please review.',
    },
    {
        'Email': 'user8@gmail.com', 'Password': 'CheckPass!',
        'TOTP Secret': 'JBSWY3DPEHPK3PXP', 'Operations': '9',
    },
]

for row_idx, example in enumerate(examples, 2):
    for col_idx, header in enumerate(all_headers, 1):
        val = example.get(header, '')
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.font = example_font
        cell.alignment = left_align
        cell.border = thin_border

# ── Column widths ────────────────────────────────────────────────────
widths = {
    'Email': 25, 'Password': 18, 'TOTP Secret': 24,
    'Recovery Email': 25, 'Recovery Phone': 20, 'Year': 10,
    **{f'Backup Code {i}': 16 for i in range(1, 11)},
    'Operations': 22, 'New Password': 18, 'New Recovery Phone': 20,
    'New Recovery Email': 24, 'New 2FA Phone': 20, 'First Name': 14,
    'Last Name': 14, 'GMB Name': 25, 'GMB URL': 55, 'Review Text': 35,
    'Review Stars': 14, 'Appeal Message': 45,
}
for col_idx, header in enumerate(all_headers, 1):
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = widths.get(header, 18)

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:{openpyxl.utils.get_column_letter(len(all_headers))}1'

# ── GUIDE sheet ──────────────────────────────────────────────────────
guide = wb.create_sheet('Guide')
guide_data = [
    ['Column', 'Required', 'Used By', 'Description'],
    ['Email', 'YES', 'All Steps', 'Gmail address (e.g. user@gmail.com)'],
    ['Password', 'YES', 'All Steps', 'Account password'],
    ['TOTP Secret', 'If 2FA', 'All Steps', 'Authenticator secret key (base32). Leave empty if no 2FA.'],
    ['Recovery Email', 'Optional', 'All Steps', 'Current recovery email on account. Used when Google asks to confirm email during login.'],
    ['Recovery Phone', 'Optional', 'All Steps', 'Current recovery phone on account. Used when Google asks to confirm phone during login.'],
    ['Year', 'Optional', 'All Steps', 'Year the Gmail account was created. Used when Google asks "When did you create this account?" during recovery.'],
    ['Backup Code 1-10', 'If 2FA', 'All Steps', 'Google backup codes (10 columns: Backup Code 1 to 10). Each code is 8 digits (e.g. 1234 5678). Used if TOTP fails.'],
    ['Operations', 'YES', 'All Steps', 'Comma-separated operation codes (see Operations tab)'],
    ['New Password', 'Op 1', 'Step 2', 'New password for Op 1 (Change Password)'],
    ['New Recovery Phone', 'Op 2a', 'Step 2', 'Phone number with country code (e.g. +8801712345678)'],
    ['New Recovery Email', 'Op 3a', 'Step 2', 'New recovery email address'],
    ['New 2FA Phone', 'Op 6a', 'Step 2', 'Phone for 2FA SMS verification'],
    ['First Name', 'Op 8', 'Step 2', 'New first name for account'],
    ['Last Name', 'Op 8', 'Step 2', 'New last name for account'],
    ['GMB Name', 'R3 (optional)', 'Step 3', 'Google My Business name (for reporting/tracking)'],
    ['GMB URL', 'R3', 'Step 3', 'Full Google Maps place URL'],
    ['Review Text', 'R3 (optional)', 'Step 3', 'Review text. Leave empty for star-only review.'],
    ['Review Stars', 'R3', 'Step 3', 'Star rating 1-5 (default: 5)'],
    ['Appeal Message', 'A1', 'Step 4', 'Appeal message text for suspended accounts'],
]

guide_header_fill = PatternFill('solid', fgColor='34495E')
for row_idx, row_data in enumerate(guide_data, 1):
    for col_idx, val in enumerate(row_data, 1):
        cell = guide.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = left_align
        if row_idx == 1:
            cell.font = header_font
            cell.fill = guide_header_fill
            cell.alignment = center

guide.column_dimensions['A'].width = 22
guide.column_dimensions['B'].width = 14
guide.column_dimensions['C'].width = 14
guide.column_dimensions['D'].width = 55
guide.freeze_panes = 'A2'

# ── OPERATIONS sheet ─────────────────────────────────────────────────
ops_sheet = wb.create_sheet('Operations')
ops_data = [
    ['Code', 'Step', 'Operation', 'Description'],
    ['L1', 'Step 1', 'Language Change', 'Set account language to English (US)'],
    ['L2', 'Step 1', 'Activity Fix', 'Clear notifications + security checkup'],
    ['L4', 'Step 1', 'Safe Browsing ON', 'Enable Google Safe Browsing'],
    ['L5', 'Step 1', 'Safe Browsing OFF', 'Disable Google Safe Browsing'],
    ['', '', '', ''],
    ['1', 'Step 2', 'Change Password', 'Set new password (needs New Password column)'],
    ['2a', 'Step 2', 'Add Recovery Phone', 'Add/update recovery phone (needs New Recovery Phone)'],
    ['2b', 'Step 2', 'Remove Recovery Phone', 'Remove recovery phone from account'],
    ['3a', 'Step 2', 'Add Recovery Email', 'Add/update recovery email (needs New Recovery Email)'],
    ['3b', 'Step 2', 'Remove Recovery Email', 'Remove recovery email from account'],
    ['4a', 'Step 2', 'Add Authenticator', 'Generate new authenticator key (replaces old)'],
    ['4b', 'Step 2', 'Remove Authenticator', 'Remove authenticator app from account'],
    ['5a', 'Step 2', 'Generate Backup Codes', 'Generate new set of backup codes'],
    ['5b', 'Step 2', 'Remove Backup Codes', 'Revoke all backup codes'],
    ['6a', 'Step 2', 'Add 2FA Phone', 'Add/replace 2FA phone (needs New 2FA Phone)'],
    ['6b', 'Step 2', 'Remove 2FA Phone', 'Remove all 2FA phone numbers'],
    ['7', 'Step 2', 'Remove All Devices', 'Remove all connected devices'],
    ['8', 'Step 2', 'Change Name', 'Change display name (needs First Name / Last Name)'],
    ['9', 'Step 2', 'Security Checkup', 'Expand sections, confirm recovery, remove 3rd party access'],
    ['', '', '', ''],
    ['R1', 'Step 3', 'Delete All Reviews', 'Delete every posted review'],
    ['R2', 'Step 3', 'Delete Draft Reviews', 'Delete pending/draft reviews only'],
    ['R3', 'Step 3', 'Write Review', 'Post review (needs GMB URL, Review Stars, optional Review Text)'],
    ['R4', 'Step 3', 'Profile Lock ON', 'Make Google Maps profile private'],
    ['R5', 'Step 3', 'Profile Lock OFF', 'Make Google Maps profile public'],
    ['', '', '', ''],
    ['A1', 'Step 4', 'Do All Appeal', 'Submit appeal for flagged/suspended account'],
    ['A2', 'Step 4', 'Delete Refused Appeals', 'Remove all refused appeal entries'],
    ['A3', 'Step 4', 'Live Check', 'Check current appeal status and write to Excel'],
]

ops_header_fill = PatternFill('solid', fgColor='2980B9')
step_fills = {
    'Step 1': PatternFill('solid', fgColor='F0F4C3'),
    'Step 2': PatternFill('solid', fgColor='E1BEE7'),
    'Step 3': PatternFill('solid', fgColor='C8E6C9'),
    'Step 4': PatternFill('solid', fgColor='FFE0B2'),
}

for row_idx, row_data in enumerate(ops_data, 1):
    for col_idx, val in enumerate(row_data, 1):
        cell = ops_sheet.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.alignment = left_align
        if row_idx == 1:
            cell.font = header_font
            cell.fill = ops_header_fill
            cell.alignment = center
        elif row_data[1] in step_fills:
            cell.fill = step_fills[row_data[1]]

ops_sheet.column_dimensions['A'].width = 10
ops_sheet.column_dimensions['B'].width = 10
ops_sheet.column_dimensions['C'].width = 28
ops_sheet.column_dimensions['D'].width = 60
ops_sheet.freeze_panes = 'A2'

# ── Save ─────────────────────────────────────────────────────────────
wb.save('master_template.xlsx')
print('master_template.xlsx created!')
print(f'Sheets: {wb.sheetnames}')
print(f'Columns: {len(all_headers)}')
print(f'Example rows: {len(examples)}')
