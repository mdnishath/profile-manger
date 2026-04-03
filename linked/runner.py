"""
Linked multi-step worker: TWO-PHASE login in a single browser session.

Phase 1 (General login):   Steps 1/3/4 ops  — standard login URL → inbox
Phase 2 (Recovery login):  Step 2 ops        — recovery URL → rapt token → security ops

Both phases share the SAME browser instance.  Signout happens once at the end.
"""

import asyncio
import traceback
from urllib.parse import urlparse, parse_qs

import pandas as pd
from playwright.async_api import async_playwright

from src.screen_detector import ScreenDetector, LoginScreen
from src.utils import ConfigManager, TOTPGenerator
from src.login_flow import execute_login_flow

from shared.logger import print, _log
from shared.browser import launch_browser, create_context
from shared.signout import perform_signout
from shared import proxy_manager, fingerprint_manager

# Step 1 operations
from step1.language_change import change_language_to_english_us
from step1.operations import fix_activity, set_safe_browsing, check_map_used, get_gmail_creation_year

# Step 2 operations
from step2.operations import (
    change_password, update_recovery_phone, remove_recovery_phone,
    update_recovery_email, remove_recovery_email,
    change_authenticator_app, remove_authenticator_app,
    generate_backup_codes, remove_backup_codes,
    add_and_replace_2fa_phone, remove_2fa_phone,
    remove_all_devices, change_name, security_checkup,
    enable_2fa, disable_2fa,
)

# Step 3 operations
from step3.operations import (
    delete_all_reviews, delete_not_posted_reviews,
    write_review, set_profile_lock, get_review_link,
)

# Step 4 operations
from step4.operations import (
    do_all_appeal, delete_refused_appeal, live_check,
)


class LinkedWorker:
    """Linked multi-step worker: TWO-PHASE login, single browser session per account.

    Phase 1 — General login  (Steps 1/3/4):  standard URL → inbox → ops → no signout
    Phase 2 — Recovery login (Step 2):        recovery URL → rapt token → security ops → signout
    """

    def __init__(self, worker_id, excel_processor, steps_config=None):
        self.worker_id = worker_id
        self.excel_processor = excel_processor
        self.config = ConfigManager()
        self.steps_config = steps_config or {}
        self.steps = sorted(self.steps_config.get('steps', [1]))
        self.ops_per_step = self.steps_config.get('ops_per_step', {})
        self._extra_data = {}  # L6 map_used, L7 gmail_year results

    def _parse_ops(self, ops_string):
        """Parse comma/space-separated operations string into list.
        Preserves original case (Step 1: L1/L2, Step 2: 1/2a/3a, Step 3: R1/R2, Step 4: A1/A2).
        """
        raw = str(ops_string).replace(',', ' ')
        return [op.strip() for op in raw.split() if op.strip()]

    async def process_account(self, account):
        """Process one account with TWO-PHASE login in a single browser session.

        Phase 1 (General login):   Steps 1/3/4 → standard login URL → inbox → ops → NO signout
        Phase 2 (Recovery login):  Step 2       → recovery URL → rapt token → security ops → signout

        If only step 2 selected:   skip Phase 1, go straight to Phase 2.
        If no step 2 selected:     skip Phase 2, Phase 1 ends with signout.
        """
        email       = account.get('Email', '')
        password    = account.get('Password', '')

        # Flexible TOTP column — try multiple common names
        totp_secret = ''
        for _totp_col in ['TOTP Secret', 'totp_secret', 'TOTP', 'totp', 'Totp Secret',
                          'TOTP Key', 'totp_key', 'Authenticator Key', 'authenticator_key',
                          'Secret Key', 'secret_key', 'OTP Secret', 'otp_secret']:
            _tv = account.get(_totp_col, '')
            if _tv and not pd.isna(_tv) and str(_tv).strip() and str(_tv).strip().lower() != 'nan':
                totp_secret = str(_tv).strip()
                break

        # Merge backup codes
        _bc_list = []
        for _bci in range(1, 11):
            _bc_val = account.get(f'Backup Code {_bci}', '')
            if not pd.isna(_bc_val) and str(_bc_val).strip():
                _bc_list.append(str(_bc_val).strip())
        backup_code = _bc_list[0] if _bc_list else ''
        account['Backup Code'] = backup_code
        row_index = account.get('row_index', 0)

        # Validate
        if pd.isna(email) or not str(email).strip() or str(email).strip().lower() == 'nan':
            _log(self.worker_id, "VALIDATE: SKIP - Invalid email (empty or nan)")
            self.excel_processor.update_row(row_index, status='FAILED', error='Invalid email address')
            return

        if pd.isna(password) or not str(password).strip():
            _log(self.worker_id, f"VALIDATE: SKIP {email} - Invalid password (empty)")
            self.excel_processor.update_row(row_index, status='FAILED', error='Invalid password')
            return

        email       = str(email).strip()
        password    = str(password).strip()
        totp_secret = str(totp_secret).strip() if not pd.isna(totp_secret) else ''

        # Build per-step operations lists
        step_ops = {}
        all_ops_flat = []
        for step_num in self.steps:
            key = str(step_num)
            ops_str = self.ops_per_step.get(key, account.get('Operations', ''))
            ops_list = self._parse_ops(ops_str)
            step_ops[step_num] = ops_list
            all_ops_flat.extend(ops_list)

        total_ops = len(all_ops_flat)
        step_label = '+'.join(str(s) for s in self.steps)

        # Separate steps into two phases
        non_step2_steps = [s for s in self.steps if s != 2]  # Steps 1, 3, 4
        has_step2 = 2 in self.steps

        _log(self.worker_id, "=" * 60)
        _log(self.worker_id, f"LINKED ACCOUNT START: {email} (Row {row_index})")
        _log(self.worker_id, f"  Steps: {self.steps} | TOTP: {'YES' if totp_secret else 'NO'} | Backup: {'YES' if backup_code else 'NO'}")
        if non_step2_steps:
            _log(self.worker_id, f"  Phase 1 (General login): Steps {non_step2_steps}")
        if has_step2:
            _log(self.worker_id, f"  Phase 2 (Recovery login): Step 2")
        for sn in self.steps:
            _log(self.worker_id, f"  Step {sn} ops: {step_ops[sn]}")
        _log(self.worker_id, f"  Total operations: {total_ops}")
        _log(self.worker_id, "=" * 60)

        operations_done   = []
        operations_failed = []

        # ── Per-account healthy proxy with retry ─────────────────────
        _NETWORK_ERRORS = (
            'net::ERR_', 'NS_ERROR_', 'Connection refused',
            'Connection reset', 'Connection timed out',
            'ERR_PROXY', 'ERR_TUNNEL', 'ERR_SOCKS',
            'SOCKS', 'Proxy connection', 'Network is unreachable',
            'Connection to Google failed', 'ECONNREFUSED', 'ETIMEDOUT',
            'socket hang up', 'Target closed', 'browser has been closed',
        )
        _MAX_PROXY_RETRIES = 5
        _tried_proxies: list[dict] = []
        _using_local_ip = False

        for _proxy_attempt in range(_MAX_PROXY_RETRIES):
            account_proxy = proxy_manager.get_healthy_proxy(exclude=_tried_proxies)
            if account_proxy:
                _tried_proxies.append(account_proxy)
                _log(self.worker_id, f"[PROXY] {email} → {account_proxy.get('server', '')} (attempt {_proxy_attempt+1}/{_MAX_PROXY_RETRIES})")
            else:
                if _using_local_ip:
                    _log(self.worker_id, f"[PROXY] {email} → All proxies exhausted + local IP tried. Giving up.")
                    break
                _using_local_ip = True
                _log(self.worker_id, f"[PROXY] {email} → Local IP (no healthy proxy)")

            _pw = None

            try:
                _pw = await async_playwright().start()
                _log(self.worker_id, "[BROWSER] Launching Chromium...")
                browser, _socks_bridge = await launch_browser(_pw, proxy=account_proxy)
                context = await create_context(browser)
                page    = await context.new_page()
                _log(self.worker_id, "[BROWSER] Browser ready")

                detector = ScreenDetector(page)
                totp_gen = TOTPGenerator()

                op_global_idx = 0
                authenticator_key = ''
                backup_codes_str = ''

                # ════════════════════════════════════════════════════════════
                # PHASE 1: General login → Step 1/3/4 ops → NO signout
                # ════════════════════════════════════════════════════════════
                if non_step2_steps:
                    general_login_url = self.config.get_url("login")
                    _log(self.worker_id, "")
                    _log(self.worker_id, "╔══════════════════════════════════════╗")
                    _log(self.worker_id, "║  PHASE 1: GENERAL LOGIN             ║")
                    _log(self.worker_id, "╚══════════════════════════════════════╝")
                    _log(self.worker_id, f"[LOGIN-1] Starting general login for {email}")
                    _log(self.worker_id, f"[LOGIN-1] URL = {general_login_url[:80]} | require_inbox=True")

                    login_result = await execute_login_flow(
                        page=page,
                        account=account,
                        worker_id=self.worker_id,
                        login_url=general_login_url,
                        detector=detector,
                        totp_gen=totp_gen,
                        require_inbox=True,
                    )

                    if not login_result.get('success'):
                        _log(self.worker_id, f"[LOGIN-1] FAILED - {login_result.get('error', 'Unknown')}")
                        raise Exception(login_result.get('error', 'Unknown login failure'))

                    # Handle forced password change
                    forced_new_pw = login_result.get('forced_new_password', '')
                    if forced_new_pw:
                        _log(self.worker_id, f"*** FORCED PASSWORD CHANGE: new pw = {forced_new_pw}")
                        try:
                            from openpyxl import load_workbook as _lwb
                            with self.excel_processor.lock:
                                _wb = _lwb(self.excel_processor.excel_file)
                                _ws = _wb.active
                                _headers = [c.value for c in _ws[1]]
                                if 'Password' in _headers:
                                    _ws.cell(row_index, _headers.index('Password') + 1, forced_new_pw)
                                if 'New Password' in _headers:
                                    _ws.cell(row_index, _headers.index('New Password') + 1, forced_new_pw)
                                _wb.save(self.excel_processor.excel_file)
                                _wb.close()
                            _log(self.worker_id, f"Password updated in Excel for {email}")
                        except Exception as pw_err:
                            _log(self.worker_id, f"WARNING: Could not save new password: {pw_err}")

                    _log(self.worker_id, f"[LOGIN-1] SUCCESS - Inbox loaded. URL = {page.url[:100]}")

                    # ── Run Step 1/3/4 operations ─────────────────────────
                    for step_num in non_step2_steps:
                        ops_list = step_ops.get(step_num, [])
                        if not ops_list:
                            continue

                        _log(self.worker_id, "")
                        _log(self.worker_id, f"{'='*40}")
                        _log(self.worker_id, f"STEP {step_num} OPERATIONS ({len(ops_list)} ops)")
                        _log(self.worker_id, f"{'='*40}")

                        for op_idx, op in enumerate(ops_list, 1):
                            op_global_idx += 1
                            _log(self.worker_id, "")
                            _log(self.worker_id, f"-- [S{step_num}] {op} ({op_global_idx}/{total_ops}) --")
                            try:
                                if step_num == 1:
                                    result = await self._run_step1_op(op, page)
                                elif step_num == 3:
                                    result = await self._run_step3_op(op, page, account)
                                elif step_num == 4:
                                    result = await self._run_step4_op(op, page, account)
                                else:
                                    _log(self.worker_id, f"[OP] Unknown step {step_num} — skipping")
                                    continue

                                # Step 1/3/4: result is True/False or dict or tuple
                                if isinstance(result, tuple) and len(result) == 2:
                                    # L6/L7 style: (True, "Fresh") or (False, "error")
                                    ok_flag, data_val = result
                                    if not ok_flag:
                                        raise Exception(f"{op} failed: {data_val}")
                                    # Store L6/L7 data for Excel saving
                                    if op == 'L6':
                                        self._extra_data['map_used'] = data_val
                                    elif op == 'L7':
                                        self._extra_data['gmail_year'] = data_val
                                    operations_done.append(f'S{step_num}: {op}={data_val}')
                                    _log(self.worker_id, f"[OP] {op}: SUCCESS → {data_val}")
                                elif isinstance(result, dict):
                                    if result.get('success') is False:
                                        raise Exception(result.get('summary', f'{op} failed'))
                                    summary = result.get('summary', op)
                                    operations_done.append(f'S{step_num}: {summary}')
                                    _log(self.worker_id, f"[OP] {op}: SUCCESS - {summary}")
                                elif result is True:
                                    operations_done.append(f'S{step_num}: {op}')
                                    _log(self.worker_id, f"[OP] {op}: SUCCESS")
                                elif result is False:
                                    raise Exception(f"Operation {op} returned False")
                                else:
                                    operations_done.append(f'S{step_num}: {op}')
                                    _log(self.worker_id, f"[OP] {op}: SUCCESS")

                            except Exception as op_err:
                                err_msg = str(op_err)[:100]
                                _log(self.worker_id, f"[OP] {op} FAILED: {err_msg}")
                                operations_failed.append(f"S{step_num}-{op}: {err_msg}")
                                remaining = total_ops - op_global_idx
                                _log(self.worker_id, f"[OP] SKIP to next ({remaining} remaining)...")
                                continue

                    _log(self.worker_id, "")
                    _log(self.worker_id, f"[PHASE-1] Step 1/3/4 ops complete. NO signout — staying logged in.")

                # ════════════════════════════════════════════════════════════
                # PHASE 2: Recovery login → rapt capture → Step 2 ops
                # ════════════════════════════════════════════════════════════
                if has_step2:
                    recovery_login_url = (
                        "https://accounts.google.com/signin/v2/recoveryidentifier"
                        "?flowName=GlifWebSignIn&flowEntry=AccountRecovery&ddm=0"
                    )
                    _log(self.worker_id, "")
                    _log(self.worker_id, "╔══════════════════════════════════════╗")
                    _log(self.worker_id, "║  PHASE 2: RECOVERY LOGIN            ║")
                    _log(self.worker_id, "╚══════════════════════════════════════╝")
                    _log(self.worker_id, f"[LOGIN-2] Starting recovery login for {email}")
                    _log(self.worker_id, f"[LOGIN-2] URL = {recovery_login_url[:80]} | require_inbox=False")

                    login_result2 = await execute_login_flow(
                        page=page,
                        account=account,
                        worker_id=self.worker_id,
                        login_url=recovery_login_url,
                        detector=detector,
                        totp_gen=totp_gen,
                        require_inbox=False,
                    )

                    if not login_result2.get('success'):
                        _log(self.worker_id, f"[LOGIN-2] FAILED - {login_result2.get('error', 'Unknown')}")
                        raise Exception(login_result2.get('error', 'Unknown recovery login failure'))

                    # Handle forced password change on recovery login
                    forced_new_pw2 = login_result2.get('forced_new_password', '')
                    if forced_new_pw2:
                        _log(self.worker_id, f"*** FORCED PASSWORD CHANGE (Phase 2): new pw = {forced_new_pw2}")
                        try:
                            from openpyxl import load_workbook as _lwb
                            with self.excel_processor.lock:
                                _wb = _lwb(self.excel_processor.excel_file)
                                _ws = _wb.active
                                _headers = [c.value for c in _ws[1]]
                                if 'Password' in _headers:
                                    _ws.cell(row_index, _headers.index('Password') + 1, forced_new_pw2)
                                if 'New Password' in _headers:
                                    _ws.cell(row_index, _headers.index('New Password') + 1, forced_new_pw2)
                                _wb.save(self.excel_processor.excel_file)
                                _wb.close()
                            _log(self.worker_id, f"Password updated in Excel for {email}")
                        except Exception as pw_err:
                            _log(self.worker_id, f"WARNING: Could not save new password: {pw_err}")

                    _log(self.worker_id, f"[LOGIN-2] SUCCESS - URL = {page.url[:100]}")

                    # ── Capture rapt token ────────────────────────────────
                    password_url = await self._capture_rapt_url(page, detector)

                    # ── Run Step 2 operations ─────────────────────────────
                    ops_list_s2 = step_ops.get(2, [])
                    if ops_list_s2:
                        _log(self.worker_id, "")
                        _log(self.worker_id, f"{'='*40}")
                        _log(self.worker_id, f"STEP 2 OPERATIONS ({len(ops_list_s2)} ops)")
                        _log(self.worker_id, f"{'='*40}")

                        for op_idx, op in enumerate(ops_list_s2, 1):
                            op_global_idx += 1
                            _log(self.worker_id, "")
                            _log(self.worker_id, f"-- [S2] {op} ({op_global_idx}/{total_ops}) --")
                            try:
                                result = await self._run_step2_op(op, page, account, password_url)

                                if op == '4a' and isinstance(result, tuple):
                                    success, new_key = result
                                    if success and new_key:
                                        authenticator_key = new_key
                                        operations_done.append(f'S2: Generate Authenticator')
                                        _log(self.worker_id, f"[OP] {op}: SUCCESS (key={new_key[:20]}...)")
                                        continue
                                    else:
                                        raise Exception("Could not generate authenticator key")
                                elif op == '5a' and isinstance(result, list):
                                    backup_codes_str = ', '.join(result[:10])
                                    operations_done.append(f'S2: Generate Backup Codes')
                                    _log(self.worker_id, f"[OP] {op}: SUCCESS ({len(result)} codes)")
                                    continue
                                elif isinstance(result, str) and 'SKIP' in result:
                                    _log(self.worker_id, f"[OP] {op}: SKIPPED - {result}")
                                    continue
                                elif result is False:
                                    raise Exception(f"Operation {op} returned False")
                                elif result is True or result:
                                    operations_done.append(f'S2: {op}')
                                    _log(self.worker_id, f"[OP] {op}: SUCCESS")
                                    continue
                                else:
                                    raise Exception(f"Operation {op} returned falsy: {result}")

                            except Exception as op_err:
                                err_msg = str(op_err)[:100]
                                _log(self.worker_id, f"[OP] {op} FAILED: {err_msg}")
                                operations_failed.append(f"S2-{op}: {err_msg}")
                                remaining = total_ops - op_global_idx
                                _log(self.worker_id, f"[OP] SKIP to next ({remaining} remaining)...")
                                continue

                # ── ALL OPERATIONS COMPLETE ────────────────────────────────
                _log(self.worker_id, "")
                _log(self.worker_id, "=" * 60)
                _log(self.worker_id, f"ALL LINKED OPERATIONS DONE for {email}")
                _log(self.worker_id, f"  Done:   {len(operations_done)}/{total_ops} -> {', '.join(operations_done) or 'none'}")
                if operations_failed:
                    _log(self.worker_id, f"  Failed: {len(operations_failed)}/{total_ops} -> {', '.join(operations_failed)}")
                _log(self.worker_id, "=" * 60)

                # ── SIGNOUT (once, at the very end) ───────────────────────
                await perform_signout(page, self.worker_id)

                try:
                    await asyncio.wait_for(context.close(), timeout=5)
                except Exception:
                    pass
                try:
                    if browser:
                        await asyncio.wait_for(browser.close(), timeout=5)
                except Exception:
                    pass
                if _socks_bridge:
                    await _socks_bridge.stop()
                if _pw:
                    try:
                        await _pw.stop()
                    except Exception:
                        pass
                _log(self.worker_id, "[BROWSER] Closed")

                # ── EXCEL → SUCCESS ──────────────────────────────────────
                failed_msg = ' | '.join(operations_failed) if operations_failed else ''

                extra = {}
                if authenticator_key:
                    extra['authenticator_key'] = authenticator_key
                if backup_codes_str:
                    extra['backup_codes'] = backup_codes_str

                self.excel_processor.update_row_status(
                    row_index=row_index,
                    status='SUCCESS',
                    operations_done=', '.join(operations_done) if operations_done else 'None',
                    error_message=failed_msg,
                    **extra,
                )

                # ── Save L6/L7 results to dedicated Excel columns ────────
                _map_used = self._extra_data.get('map_used', '')
                _gmail_year = self._extra_data.get('gmail_year', '')
                if _map_used or _gmail_year:
                    try:
                        from openpyxl import load_workbook as _lwb
                        with self.excel_processor.lock:
                            _wb = _lwb(self.excel_processor.excel_file)
                            _ws = _wb.active
                            _headers = [c.value for c in _ws[1]]
                            if _map_used:
                                _cn = 'Map Used'
                                _ci = _headers.index(_cn) + 1 if _cn in _headers else _ws.max_column + 1
                                if _cn not in _headers:
                                    _ws.cell(1, _ci, _cn)
                                _ws.cell(row_index, _ci, _map_used)
                            if _gmail_year:
                                _cn = 'Gmail Year'
                                _ci = _headers.index(_cn) + 1 if _cn in _headers else _ws.max_column + 1
                                if _cn not in _headers:
                                    _ws.cell(1, _ci, _cn)
                                _ws.cell(row_index, _ci, _gmail_year)
                            _wb.save(self.excel_processor.excel_file)
                            _wb.close()
                            _log(self.worker_id, f"[EXCEL] Saved Map Used={_map_used}, Gmail Year={_gmail_year}")
                    except Exception as _ee:
                        _log(self.worker_id, f"[EXCEL] WARNING: L6/L7 save failed: {_ee}")

                _log(self.worker_id, f"[EXCEL] Row {row_index} -> SUCCESS")
                _log(self.worker_id, f"ACCOUNT DONE: {email} = SUCCESS")

                proxy_manager.mark_alive(account_proxy)
                self._extra_data = {}  # Reset for next account
                break  # exit proxy retry loop

            except Exception as e:
                error_str = str(e)
                _log(self.worker_id, "=" * 60)
                _log(self.worker_id, f"ACCOUNT ERROR: {email} (attempt {_proxy_attempt+1}/{_MAX_PROXY_RETRIES})")
                _log(self.worker_id, f"  Error: {error_str}")
                _log(self.worker_id, "=" * 60)

                try:
                    if 'context' in locals() and context:
                        await asyncio.wait_for(context.close(), timeout=5)
                    if 'browser' in locals() and browser:
                        await asyncio.wait_for(browser.close(), timeout=5)
                    if '_socks_bridge' in locals() and _socks_bridge:
                        await _socks_bridge.stop()
                    if '_pw' in locals() and _pw:
                        await _pw.stop()
                    _log(self.worker_id, "[BROWSER] Closed")
                except Exception as browser_err:
                    _log(self.worker_id, f"[BROWSER] Close error: {browser_err}")

                is_network_error = any(kw.lower() in error_str.lower() for kw in _NETWORK_ERRORS)

                if is_network_error and _proxy_attempt < _MAX_PROXY_RETRIES - 1:
                    proxy_manager.mark_dead(account_proxy)
                    _log(self.worker_id, "[PROXY] Network error — switching to new proxy...")
                    await asyncio.sleep(2)
                    continue

                auth_failures = [
                    'WRONG_PASSWORD', 'PASSWORD_CHANGED', 'WRONG_2FA_CODE',
                    'CAPTCHA_REQUIRED', 'RECOVERY_EMAIL_VERIFICATION',
                    'PHONE_VERIFICATION', 'ACCOUNT_LOCKED', 'TOO_MANY_ATTEMPTS',
                    'EMAIL_NOT_FOUND', 'ACCOUNT_RECOVERY_REDIRECT',
                    'SUSPICIOUS_ACTIVITY', 'UNUSUAL_LOCATION',
                    'VERIFY_PHONE_CODE', 'NO_2FA_CREDENTIALS',
                    'SENSITIVE_ACTION_BLOCKED', 'DEVICE_APPROVAL_FAILED',
                ]
                if any(f in error_str for f in auth_failures):
                    _log(self.worker_id, "  Type: AUTH FAILURE (no retry)")

                if is_network_error:
                    proxy_manager.mark_dead(account_proxy)

                traceback.print_exc()

                self.excel_processor.update_row_status(
                    row_index=row_index,
                    status='FAILED',
                    operations_done='',
                    error_message=error_str[:500],
                )
                _log(self.worker_id, f"[EXCEL] Row {row_index} -> FAILED")
                break

    # ── Rapt URL capture (copied from step2/runner.py) ─────────────────

    async def _capture_rapt_url(self, page, detector):
        """Capture password URL with rapt token for step 2 operations."""
        _log(self.worker_id, "[RAPT] Capturing rapt token for Step 2 operations...")
        password_url = None

        current_screen = await detector.detect_current_screen()
        _log(self.worker_id, f"[RAPT] Current screen: {current_screen.name}")

        # Extract rapt token from current URL
        rapt_token = None
        try:
            qs = parse_qs(urlparse(page.url).query)
            rapt_token = qs.get('rapt', [None])[0]
            if rapt_token:
                _log(self.worker_id, f"[RAPT] Extracted rapt token from URL: {rapt_token[:30]}...")
        except Exception:
            pass

        # Handle success screen — click "Change password"
        if current_screen == LoginScreen.SUCCESS_SCREEN:
            _log(self.worker_id, "[RAPT] Success screen — clicking 'Change password'...")
            change_pwd_selectors = [
                'a[aria-label="Change password"]',
                'a[href*="signinoptions/password"]',
                '[jsname="hSRGPd"]',
                'a[aria-label*="password" i]',
                'a:has-text("Change password")',
                'button:has-text("Change password")',
            ]
            clicked = False
            url_before = page.url
            for sel in change_pwd_selectors:
                try:
                    elem = page.locator(sel).first
                    if await elem.count() > 0:
                        await elem.click()
                        _log(self.worker_id, f"[RAPT] Clicked: {sel}")
                        clicked = True
                        break
                except Exception:
                    continue

            if clicked:
                for wait_i in range(10):
                    await asyncio.sleep(1)
                    if page.url != url_before:
                        _log(self.worker_id, f"[RAPT] Navigation detected: {page.url[:100]}")
                        break

                captured_url = page.url
                if 'myaccount.google.com' in captured_url and 'signinoptions' in captured_url:
                    password_url = captured_url
                elif 'rapt=' in captured_url:
                    password_url = captured_url
                else:
                    _log(self.worker_id, "[RAPT] URL not password page, will construct manually")

        # Fallback: construct from rapt token
        if not password_url and rapt_token:
            _log(self.worker_id, "[RAPT] Building URL from rapt token...")
            password_url = f"https://myaccount.google.com/signinoptions/password?rapt={rapt_token}"
            try:
                await page.goto(password_url, wait_until="domcontentloaded", timeout=15000)
                await asyncio.sleep(3)
                password_url = page.url
            except Exception as nav_err:
                _log(self.worker_id, f"[RAPT] Navigation failed: {nav_err}")

        # Final fallback
        if not password_url:
            try:
                password_change_url = self.config.get_url("password_change")
                await page.goto(password_change_url, wait_until="networkidle", timeout=15000)
                await asyncio.sleep(2)
                password_url = page.url
            except Exception:
                pass

        if not password_url:
            _log(self.worker_id, "[RAPT] WARNING: Could not capture password URL")
            password_url = page.url
        else:
            has_rapt = 'rapt=' in password_url
            _log(self.worker_id, f"[RAPT] Password URL captured (rapt={'YES' if has_rapt else 'NO'}): {password_url[:100]}")

        return password_url

    # ── Step 1 operation dispatcher ────────────────────────────────────

    async def _run_step1_op(self, op, page):
        """Dispatch a Step 1 operation. Returns True/False or (True, data)."""
        if op == 'L1':
            return await change_language_to_english_us(page, self.worker_id)
        elif op == 'L2':
            return await fix_activity(page, self.worker_id)
        elif op == 'L4':
            return await set_safe_browsing(page, self.worker_id, enabled=True)
        elif op == 'L5':
            return await set_safe_browsing(page, self.worker_id, enabled=False)
        elif op == 'L6':
            return await check_map_used(page, self.worker_id)
        elif op == 'L7':
            return await get_gmail_creation_year(page, self.worker_id)
        else:
            _log(self.worker_id, f"[S1] Unknown op: {op}")
            return True

    # ── Step 2 operation dispatcher ────────────────────────────────────

    async def _run_step2_op(self, op, page, account, password_url):
        """Dispatch a Step 2 operation. Returns mixed types (True/False/tuple/list/str).
        Op codes are lowercase from UI: 1, 2a, 2b, 3a, 3b, 4a, 4b, 5a, 5b, 6a, 6b, 7, 8, 9, 10a, 10b.
        """
        password = account.get('Password', '')

        if op == '1':
            new_password = account.get('New Password', password + '123')
            if pd.isna(new_password):
                new_password = password + '123'
            return await change_password(page, self.config, new_password, password_url)
        elif op == '2a':
            new_phone = account.get('New Recovery Phone', '')
            if pd.isna(new_phone):
                new_phone = ''
            if not (new_phone and str(new_phone).strip()):
                return "SKIP - No recovery phone number"
            return await update_recovery_phone(page, self.config, str(new_phone), password_url)
        elif op == '2b':
            return await remove_recovery_phone(page, self.config, password_url)
        elif op == '3a':
            new_email = account.get('New Recovery Email', '')
            if pd.isna(new_email):
                new_email = ''
            if not (new_email and str(new_email).strip()):
                return "SKIP - No recovery email"
            return await update_recovery_email(page, self.config, str(new_email), password_url)
        elif op == '3b':
            return await remove_recovery_email(page, self.config, password_url)
        elif op == '4a':
            return await change_authenticator_app(page, self.config, password_url)
        elif op == '4b':
            return await remove_authenticator_app(page, self.config, password_url)
        elif op == '5a':
            return await generate_backup_codes(page, self.config, password_url)
        elif op == '5b':
            return await remove_backup_codes(page, self.config, password_url)
        elif op == '6a':
            new_2fa_phone = account.get('New 2FA Phone', '')
            if pd.isna(new_2fa_phone):
                new_2fa_phone = ''
            if not (new_2fa_phone and str(new_2fa_phone).strip()):
                return "SKIP - No 2FA phone number"
            return await add_and_replace_2fa_phone(page, self.config, str(new_2fa_phone), password_url)
        elif op == '6b':
            return await remove_2fa_phone(page, self.config, password_url)
        elif op == '7':
            return await remove_all_devices(page, self.config, password_url)
        elif op == '8':
            first_name = account.get('First Name', '')
            last_name = account.get('Last Name', '')
            if pd.isna(first_name):
                first_name = ''
            if pd.isna(last_name):
                last_name = ''
            if not ((first_name and str(first_name).strip()) or (last_name and str(last_name).strip())):
                return "SKIP - No first/last name"
            return await change_name(page, self.config, str(first_name), str(last_name), password_url)
        elif op == '9':
            return await security_checkup(page, self.config, password_url)
        elif op == '10a':
            return await enable_2fa(page, self.config, password_url)
        elif op == '10b':
            return await disable_2fa(page, self.config, password_url)
        else:
            _log(self.worker_id, f"[S2] Unknown op: {op}")
            return f"SKIP - Unknown op: {op}"

    # ── Step 3 operation dispatcher ────────────────────────────────────

    async def _run_step3_op(self, op, page, account):
        """Dispatch a Step 3 operation."""
        if op == 'R1':
            ok = await delete_all_reviews(page, self.worker_id)
            if not ok:
                raise Exception("delete_all_reviews returned False")
            return True
        elif op == 'R2':
            ok = await delete_not_posted_reviews(page, self.worker_id)
            if not ok:
                raise Exception("delete_not_posted_reviews returned False")
            return True
        elif op == 'R3':
            review_place = account.get('GMB URL', '') or account.get('review_place_url', '')
            review_text = account.get('Review Text', '') or account.get('review_text', '')
            review_stars = account.get('Review Stars', '') or account.get('review_stars', 5)
            place = str(review_place).strip() if review_place and not pd.isna(review_place) else ''
            text = str(review_text).strip() if review_text and not pd.isna(review_text) else ''
            stars = int(review_stars) if review_stars and not pd.isna(review_stars) else 5
            if not place:
                return {'success': True, 'summary': 'SKIPPED - no place URL'}
            return await write_review(page, self.worker_id, place_url=place, review_text=text, stars=stars)
        elif op == 'R4':
            ok = await set_profile_lock(page, self.worker_id, locked=True)
            if not ok:
                raise Exception("set_profile_lock(ON) returned False")
            return True
        elif op == 'R5':
            ok = await set_profile_lock(page, self.worker_id, locked=False)
            if not ok:
                raise Exception("set_profile_lock(OFF) returned False")
            return True
        elif op == 'R6':
            return await get_review_link(page, self.worker_id)
        else:
            _log(self.worker_id, f"[S3] Unknown op: {op}")
            return True

    # ── Step 4 operation dispatcher ────────────────────────────────────

    async def _run_step4_op(self, op, page, account):
        """Dispatch a Step 4 operation."""
        email = account.get('Email', '')
        appeal_message = account.get('Appeal Message', '') or account.get('appeal_message', '')
        if appeal_message and not pd.isna(appeal_message):
            appeal_message = str(appeal_message).strip()
        else:
            appeal_message = ''

        if op == 'A1':
            return await do_all_appeal(page, self.worker_id, email=email, appeal_message=appeal_message)
        elif op == 'A2':
            return await delete_refused_appeal(page, self.worker_id)
        elif op == 'A3':
            return await live_check(page, self.worker_id)
        else:
            _log(self.worker_id, f"[S4] Unknown op: {op}")
            return {'success': True, 'summary': f'Unknown op: {op}'}
