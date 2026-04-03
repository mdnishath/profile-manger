"""
Step 1 worker: login → operations → signout.

Operations (set via Excel 'Operations' column):
    L1 — Language Change     (set account language to English US)
    L2 — Activity Fix        (clear Google account activity/history)
    L4 — Safe Browsing ON    (enable Google Safe Browsing)
    L5 — Safe Browsing OFF   (disable Google Safe Browsing)
    L6 — Map Used Filter     (check Google Maps Local Guide → Fresh/Used)
    L7 — Gmail Creation Year (oldest email date → creation year)
"""

from shared.logger import _log
from shared.base_runner import BaseGmailBotWorker

from step1.language_change import change_language_to_english_us
from step1.operations import (
    fix_activity,
    set_safe_browsing,
    check_map_used,
    get_gmail_creation_year,
)


class GmailBotWorker(BaseGmailBotWorker):
    """Step 1 worker: login + language/activity/safe-browsing operations."""

    def _get_default_operations(self):
        return 'L1'

    async def _dispatch_operation(self, op, page, account, ctx):
        if op == 'L1':
            _log(self.worker_id, "[OP] L1: Language Change")
            ok = await change_language_to_english_us(page, self.worker_id)
            if not ok:
                raise Exception("change_language_to_english_us returned False")
            return 'Language Change'

        elif op == 'L2':
            _log(self.worker_id, "[OP] L2: Activity Fix")
            ok = await fix_activity(page, self.worker_id)
            if not ok:
                raise Exception("fix_activity returned False")
            return 'Activity Fix'

        elif op == 'L4':
            _log(self.worker_id, "[OP] L4: Safe Browsing ON")
            ok = await set_safe_browsing(page, self.worker_id, enabled=True)
            if not ok:
                raise Exception("set_safe_browsing(ON) returned False")
            return 'Safe Browsing ON'

        elif op == 'L5':
            _log(self.worker_id, "[OP] L5: Safe Browsing OFF")
            ok = await set_safe_browsing(page, self.worker_id, enabled=False)
            if not ok:
                raise Exception("set_safe_browsing(OFF) returned False")
            return 'Safe Browsing OFF'

        elif op == 'L6':
            _log(self.worker_id, "[OP] L6: Map Used Filter")
            ok, result_val = await check_map_used(page, self.worker_id)
            if not ok:
                raise Exception(f"check_map_used failed: {result_val}")
            ctx['map_used_result'] = result_val
            return f'Map Used: {result_val}'

        elif op == 'L7':
            _log(self.worker_id, "[OP] L7: Gmail Creation Year")
            ok, result_val = await get_gmail_creation_year(page, self.worker_id)
            if not ok:
                raise Exception(f"get_gmail_creation_year failed: {result_val}")
            ctx['gmail_year_result'] = result_val
            return f'Gmail Year: {result_val}'

        else:
            _log(self.worker_id, f"[OP] {op}: UNKNOWN operation - skipping")
            return f"SKIP - Unknown operation: {op}"

    def _handle_operation_result(self, op, result, operations_done, ctx):
        """L6/L7 return descriptive strings, all others handled by base."""
        if isinstance(result, str) and not result.startswith('SKIP'):
            operations_done.append(result)
            _log(self.worker_id, f"[OP] {op}: SUCCESS → {result}")
        else:
            super()._handle_operation_result(op, result, operations_done, ctx)

    async def _on_operations_complete(self, page, account, row_index, ctx):
        """Save L6/L7 results to dedicated Excel columns."""
        map_used = ctx.get('map_used_result', '')
        gmail_year = ctx.get('gmail_year_result', '')

        if not map_used and not gmail_year:
            return

        try:
            from openpyxl import load_workbook as _lwb
            with self.excel_processor.lock:
                _wb = _lwb(self.excel_processor.excel_file)
                _ws = _wb.active
                _headers = [c.value for c in _ws[1]]

                if map_used:
                    col_name = 'Map Used'
                    if col_name in _headers:
                        _col_idx = _headers.index(col_name) + 1
                    else:
                        _col_idx = _ws.max_column + 1
                        _ws.cell(1, _col_idx, col_name)
                    _ws.cell(row_index, _col_idx, map_used)

                if gmail_year:
                    col_name = 'Gmail Year'
                    if col_name in _headers:
                        _col_idx = _headers.index(col_name) + 1
                    else:
                        _col_idx = _ws.max_column + 1
                        _ws.cell(1, _col_idx, col_name)
                    _ws.cell(row_index, _col_idx, gmail_year)

                _wb.save(self.excel_processor.excel_file)
                _wb.close()
                _log(self.worker_id, f"[EXCEL] Saved Map Used={map_used}, Gmail Year={gmail_year}")
        except Exception as e:
            _log(self.worker_id, f"[EXCEL] WARNING: Could not save L6/L7 results: {e}")
