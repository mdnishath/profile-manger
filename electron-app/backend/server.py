"""
Flask API Server for Electron Frontend
Provides REST API for Gmail Bot operations
"""

from flask import Flask, Blueprint, request, jsonify, Response, send_from_directory
from flask_cors import CORS
import logging
import secrets
import threading
import sys
import os
from datetime import datetime

# Force stdout/stderr to UTF-8 so unicode log chars (✓ → etc.) never crash on Windows cp1252
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8', errors='replace')
import atexit
from pathlib import Path
import pandas as pd
import time
import json
import itertools

# ── Suppress Werkzeug HTTP access logs ────────────────────────────────────────
# Flask's dev server (Werkzeug) writes every request to stderr by default.
# Electron captures stderr as [ERR] and floods the UI log panel with noise like:
#   127.0.0.1 - - [08/Mar/2026 05:59:39] "GET /api/health HTTP/1.1" 200 -
# Only show genuine errors (500s, crashes), not routine 200-OK access logs.
logging.getLogger('werkzeug').setLevel(logging.ERROR)

# RESOURCES_PATH is set by main.js (Electron):
#   - In dev mode: the gmail_boat/ root folder (one above electron-app/)
#   - In packaged EXE: process.resourcesPath where bot scripts are bundled
# Fallback: go 3 levels up from this file (dev default)
RESOURCES_PATH = Path(os.environ.get(
    'RESOURCES_PATH',
    str(Path(__file__).parent.parent.parent)
))

sys.path.insert(0, str(RESOURCES_PATH))

# Screenshots / auth-key txt files directory
# Playwright saves screenshots with relative paths like "screenshots/...",
# which resolves to CWD/screenshots/.  In production (PyInstaller EXE),
# __file__ points to a temp _MEI* extraction dir, so we must use CWD instead.
if getattr(sys, 'frozen', False):
    # Production: backend.exe — CWD is the Electron app's install directory
    SCREENSHOTS_PATH = Path.cwd() / 'screenshots'
else:
    # Development: python server.py — CWD is typically electron-app/
    SCREENSHOTS_PATH = Path(__file__).parent.parent / 'screenshots'

# Lazy-loaded: prepare_excel_with_common_settings (heavy pandas/openpyxl imports)
_prepare_excel_fn = None
def prepare_excel_with_common_settings(*args, **kwargs):
    global _prepare_excel_fn
    if _prepare_excel_fn is None:
        from prepare_excel_with_common_settings import prepare_excel_with_common_settings as _fn
        _prepare_excel_fn = _fn
    return _prepare_excel_fn(*args, **kwargs)

app = Flask(__name__)
CORS(app, supports_credentials=False)

# ── Auth (licensing removed — open access) ───────────────────────────────────
# No token or license required. All endpoints are freely accessible.

# ── App Version & Update Check ───────────────────────────────────────────────
def _read_app_version():
    """Read version from package.json (stays in sync with build.bat)."""
    try:
        pkg = Path(__file__).resolve().parent.parent / 'package.json'
        if pkg.exists():
            return json.loads(pkg.read_text(encoding='utf-8')).get('version', '0.0.0')
    except Exception:
        pass
    return '0.0.0'


APP_VERSION = _read_app_version()
_version_cache = {'last_check': 0, 'data': None}


def _get_version_manifest_url():
    """Read version_manifest_url from tools/gist_config.json (written by admin panel)."""
    cfg_path = RESOURCES_PATH / 'tools' / 'gist_config.json'
    try:
        if cfg_path.exists():
            cfg = json.loads(cfg_path.read_text(encoding='utf-8'))
            return cfg.get('version_manifest_url', '')
    except Exception:
        pass
    return ''



# Auto-incrementing log ID (no collisions unlike time*1000)
_log_id = itertools.count(1)

# Global state
processing_state = {
    'status': 'idle',  # idle, processing, completed, stopped
    'current': 0,
    'total': 0,
    'current_account': '',
    'recent_logs': [],
    'log_clear_id': 0,   # highest log id at the time of last clear; SSE only replays logs after this
    'file_path': '',  # Input file path
    'output_file_path': '',  # Output file path (for progress tracking)
    'step_name': '',   # Current step(s) being processed (e.g. 'step1', 'step2', 'step1+step2')
    'step_label': '',  # Human-readable step label for UI (e.g. 'Step 2 - Security')
    'operations': '',  # Current operations string (e.g. '1,4,5' or 'A3')
}

processing_thread = None
_processing_lock = threading.Lock()  # Prevent double-start race condition
stop_flag = threading.Event()

# SMS code relay storage (received from phone app)
sms_codes = []


def _is_safe_child(child: Path, parent: Path) -> bool:
    """Return True if child is strictly inside parent (no traversal)."""
    try:
        child.resolve().relative_to(parent.resolve())
        return True
    except ValueError:
        return False


def _persist_output_dirs(output_file_path='', input_file_path=''):
    """Persist the latest output/input paths to disk so they survive server restarts.
    Called after processing completes (or whenever output_file_path is captured)."""
    try:
        state_file = RESOURCES_PATH / 'config' / 'last_output.json'
        state_file.parent.mkdir(parents=True, exist_ok=True)

        # Read existing state and merge (keep any previously saved dirs)
        existing = {}
        if state_file.exists():
            with open(state_file, 'r', encoding='utf-8') as f:
                existing = json.load(f)

        dirs = set(existing.get('output_dirs', []))
        if output_file_path:
            d = str(Path(output_file_path).parent.resolve())
            if Path(d).exists():
                dirs.add(d)
        if input_file_path:
            d = str((Path(input_file_path).parent.parent / 'output').resolve())
            if Path(d).exists():
                dirs.add(d)

        # Only keep dirs that actually exist
        dirs = [d for d in dirs if Path(d).exists()]

        with open(state_file, 'w', encoding='utf-8') as f:
            json.dump({'output_dirs': dirs}, f, indent=2)
    except Exception as e:
        print(f"[WARN] Could not persist output dirs: {e}", flush=True)


def _load_persisted_output_dirs():
    """Load previously saved output directories from disk (survives restarts)."""
    try:
        state_file = RESOURCES_PATH / 'config' / 'last_output.json'
        if state_file.exists():
            with open(state_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return [d for d in data.get('output_dirs', []) if Path(d).exists()]
    except Exception:
        pass
    return []


def _get_report_scan_dirs():
    """Return all directories where output/report .xlsx files may live.

    Sources (in order):
      1. electron-app/output — dev mode CWD writes
      2. RESOURCES_PATH/output — production EXE mode
      3. Directory of the current processing output file (from [OUTPUT_FILE] marker)
      4. Derived from the input Excel path — same logic as ExcelProcessor
      5. Persisted output dirs from last_output.json (survives server restarts)
      6. User home / output — common fallback for ExcelProcessor outputs"""
    scan_dirs = set()

    # 1. electron-app/output — where subprocess CWD writes in dev mode
    electron_output = Path(__file__).parent.parent / 'output'
    electron_output.mkdir(exist_ok=True)
    scan_dirs.add(str(electron_output.resolve()))

    # 2. RESOURCES_PATH/output — where files go in production EXE mode
    project_output = RESOURCES_PATH / 'output'
    if project_output.exists():
        scan_dirs.add(str(project_output.resolve()))

    # 3. Directory of the current/latest output file (set by subprocess via [OUTPUT_FILE])
    output_file = processing_state.get('output_file_path', '')
    if output_file:
        output_dir = Path(output_file).parent
        if output_dir.exists():
            scan_dirs.add(str(output_dir.resolve()))

    # 4. Derived from input Excel path — mirrors ExcelProcessor logic:
    #    base_dir = Path(excel).parent.parent  →  output_dir = base_dir / "output"
    input_file = processing_state.get('file_path', '')
    if input_file:
        derived_output = Path(input_file).parent.parent / 'output'
        if derived_output.exists():
            scan_dirs.add(str(derived_output.resolve()))

    # 5. Persisted output dirs from last_output.json (survive server restarts)
    for d in _load_persisted_output_dirs():
        scan_dirs.add(d)

    # 6. User home / output — common fallback (ExcelProcessor often writes here)
    user_home_output = Path.home() / 'output'
    if user_home_output.exists():
        scan_dirs.add(str(user_home_output.resolve()))

    # 7. Profile Manager reports — use resolved path (handles empty storage_path config)
    try:
        # _get_storage_path() always returns the real resolved path regardless of config
        pm_reports = profile_manager._get_storage_path() / 'reports'
        pm_reports.mkdir(parents=True, exist_ok=True)
        scan_dirs.add(str(pm_reports.resolve()))
    except Exception:
        pass

    # 7b. Also honour any custom storage_path set in config
    try:
        pm_storage = profile_manager.get_config().get('storage_path', '')
        if pm_storage:
            custom_reports = Path(pm_storage) / 'reports'
            custom_reports.mkdir(parents=True, exist_ok=True)
            scan_dirs.add(str(custom_reports.resolve()))
    except Exception:
        pass

    # 8. AppData/Roaming/MailNexusPro fallback (default on Windows)
    try:
        appdata_roaming = os.environ.get('APPDATA', '')
        if appdata_roaming:
            mnp = Path(appdata_roaming) / 'MailNexusPro' / 'profiles' / 'reports'
            mnp.mkdir(parents=True, exist_ok=True)
            scan_dirs.add(str(mnp.resolve()))
    except Exception:
        pass

    # 9. AppData/Local/GmailBotPro legacy fallback
    try:
        appdata_local = Path(os.environ.get('LOCALAPPDATA', '')) / 'GmailBotPro' / 'browser_profiles' / 'reports'
        if appdata_local.parent.exists():
            appdata_local.mkdir(parents=True, exist_ok=True)
            scan_dirs.add(str(appdata_local.resolve()))
    except Exception:
        pass

    return scan_dirs


def _find_latest_output_file():
    """Find the most recent *_output_*.xlsx file across all known directories.
    Used when processing_state has no output_file_path (e.g. after server restart)."""
    scan_dirs = _get_report_scan_dirs()
    best = None
    best_mtime = 0
    for d in scan_dirs:
        dp = Path(d)
        if not dp.exists():
            continue
        for f in dp.rglob('*_output_*.xlsx'):
            # Skip MailNexus reports (they are derived, not source)
            if 'mailnexus' in f.name.lower() or 'Mailnexus' in f.name:
                continue
            mt = f.stat().st_mtime
            if mt > best_mtime:
                best_mtime = mt
                best = f
    return str(best) if best else None


@app.route('/api/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({'success': True, 'message': 'Server is running'})


@app.route('/api/app/version', methods=['GET'])
def app_version():
    """Return current app version."""
    return jsonify({'version': APP_VERSION})


@app.route('/api/app/check-update', methods=['GET'])
def app_check_update():
    """Check for updates by fetching the version manifest from GitHub Gist."""
    import urllib.request

    manifest_url = _get_version_manifest_url()
    if not manifest_url:
        return jsonify({'update_available': False, 'current_version': APP_VERSION,
                        'message': 'Update check not configured'})

    # Cache for 1 hour
    now = time.time()
    if now - _version_cache['last_check'] < 3600 and _version_cache['data'] is not None:
        manifest = _version_cache['data']
    else:
        try:
            req = urllib.request.Request(manifest_url, headers={
                'User-Agent': 'MailNexus-Pro/' + APP_VERSION,
                'Cache-Control': 'no-cache',
            })
            with urllib.request.urlopen(req, timeout=5) as resp:
                manifest = json.loads(resp.read().decode('utf-8'))
            _version_cache['data'] = manifest
            _version_cache['last_check'] = now
        except Exception:
            return jsonify({'update_available': False, 'current_version': APP_VERSION,
                            'message': 'Could not reach update server'})

    latest = manifest.get('latest_version', APP_VERSION)

    # Simple semver compare: split by '.', compare each part as int
    def ver_tuple(v):
        try:
            return tuple(int(x) for x in v.split('.'))
        except Exception:
            return (0, 0, 0)

    update_available = ver_tuple(latest) > ver_tuple(APP_VERSION)

    result = {
        'update_available': update_available,
        'current_version': APP_VERSION,
        'latest_version': latest,
        'download_url': manifest.get('download_url', ''),
        'release_date': manifest.get('release_date', ''),
        'changelog': manifest.get('changelog', []),
    }
    return jsonify(result)


@app.route('/api/shutdown', methods=['POST'])
def shutdown():
    """Shutdown the Flask server from Electron UI"""
    def wait_and_kill():
        time.sleep(1)
        # sys.exit triggers SystemExit which Flask can handle for cleanup
        # os._exit is used as final fallback only
        try:
            sys.exit(0)
        except SystemExit:
            os._exit(0)
    threading.Thread(target=wait_and_kill, daemon=True).start()
    return jsonify({'success': True, 'message': 'Server shutting down...'})


@app.route('/api/file-info', methods=['POST'])
def get_file_info():
    """Get Excel file statistics"""
    try:
        data = request.json or {}
        file_path = data.get('file_path')

        if not file_path or not os.path.exists(file_path):
            return jsonify({'success': False, 'message': 'File not found'})

        df = pd.read_excel(file_path, engine='openpyxl')

        # Add Status column if not present
        if 'Status' not in df.columns:
            df['Status'] = ''

        # Mark empty as PENDING
        df.loc[df['Status'].isna() | (df['Status'] == ''), 'Status'] = 'PENDING'

        total = len(df)
        success = len(df[df['Status'].str.upper() == 'SUCCESS'])
        failed = len(df[df['Status'].str.upper() == 'FAILED'])
        pending = len(df[df['Status'].str.upper() == 'PENDING'])

        return jsonify({
            'success': True,
            'data': {
                'total': total,
                'success': success,
                'failed': failed,
                'pending': pending
            }
        })

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/start-processing', methods=['POST'])
def start_processing():
    """Start processing accounts"""
    global processing_thread, stop_flag

    try:
        data = request.json or {}
        file_path = data.get('file_path')
        operations = data.get('operations', '1,4,5')
        new_password = data.get('new_password', '')
        recovery_email = data.get('recovery_email', '')
        recovery_phone = data.get('recovery_phone', '')
        num_workers = data.get('num_workers', 5)
        bot_step = int(data.get('bot_step', 2))  # UI explicit payload

        # Multi-step support
        bot_steps = data.get('bot_steps', [bot_step])  # array of step numbers
        linked = data.get('linked', False)              # linked mode toggle
        ops_per_step = data.get('ops_per_step', {})     # per-step operations map

        # Validation
        if not file_path or not os.path.exists(file_path):
            return jsonify({'success': False, 'message': 'File not found'})

        # Check if already processing (thread-safe)
        if not _processing_lock.acquire(blocking=False):
            return jsonify({'success': False, 'message': 'Already processing'})
        try:
            if processing_state['status'] == 'processing':
                return jsonify({'success': False, 'message': 'Already processing'})

            # Reset state — all inside lock to prevent race conditions
            stop_flag.clear()
            processing_state['status'] = 'processing'
            processing_state['current'] = 0
            processing_state['total'] = 0
            processing_state['current_account'] = ''
            processing_state['recent_logs'] = []
            processing_state['file_path'] = file_path
            processing_state['operations'] = operations
        finally:
            _processing_lock.release()

        is_multi = len(bot_steps) > 1
        step_label = '+'.join(str(s) for s in bot_steps)

        # Track step info for reports and UI
        _step_labels = {1: 'Language/Activity', 2: 'Security', 3: 'Maps Reviews', 4: 'Appeals'}
        if is_multi:
            processing_state['step_name'] = '+'.join(f'step{s}' for s in bot_steps)
            processing_state['step_label'] = 'Steps ' + '+'.join(f'{s}' for s in bot_steps)
        else:
            s = bot_steps[0]
            processing_state['step_name'] = f'step{s}'
            processing_state['step_label'] = f'Step {s} - {_step_labels.get(s, "")}'

        add_log('Preparing Excel file with common settings...', 'info')
        add_log(f'[DEBUG] File path: {file_path}', 'info')
        add_log(f'[DEBUG] Operations: {operations}', 'info')
        add_log(f'[DEBUG] Bot steps: {bot_steps} | Linked: {linked}', 'info')

        # Step 1: Prepare Excel
        try:
            add_log('[DEBUG] Calling prepare_excel_with_common_settings...', 'info')
            prepare_excel_with_common_settings(
                file_path,
                operations,
                new_password,
                recovery_email,
                recovery_phone
            )
            add_log(f'[DEBUG] prepare_excel_with_common_settings returned successfully', 'info')
            add_log(f'Excel prepared successfully', 'success')
        except Exception as prep_error:
            import traceback
            traceback.print_exc()
            add_log(f'ERROR preparing Excel: {str(prep_error)}', 'error')
            processing_state['status'] = 'idle'
            return jsonify({'success': False, 'message': f'Excel preparation failed: {str(prep_error)}'})

        add_log(f'[DEBUG] About to start background thread...', 'info')
        if is_multi and linked:
            add_log(f'Running LINKED Steps {step_label} matrix', 'info')
        elif is_multi:
            add_log(f'Running Steps {step_label} SEQUENTIALLY (unlinked)', 'info')
        else:
            add_log(f'Running STEP {bot_steps[0]} matrix', 'info')

        # Step 2: Start processing in background thread
        processing_thread = threading.Thread(
            target=run_processing_worker,
            args=(file_path, num_workers, bot_steps[0] if not is_multi else bot_steps[0]),
            kwargs={
                'bot_steps': bot_steps,
                'linked': linked,
                'ops_per_step': ops_per_step,
            },
            daemon=True,
            name=f'ProcessingWorker-{step_label}'
        )
        processing_thread.start()

        add_log('Background worker thread started successfully', 'success')
        add_log(f'Thread is alive: {processing_thread.is_alive()}', 'info')
        return jsonify({'success': True, 'message': 'Processing started'})

    except Exception as e:
        processing_state['status'] = 'idle'
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/stop-processing', methods=['POST'])
def stop_processing():
    """Stop processing"""
    try:
        stop_flag.set()
        processing_state['status'] = 'stopped'
        add_log('Stop requested - waiting for current account to finish...', 'warning')
        return jsonify({'success': True, 'message': 'Stopping...'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


_last_progress = {}  # Cache last successful progress read to avoid 0-glitch


@app.route('/api/progress', methods=['GET'])
def get_progress():
    """Get current progress - reads real-time from Excel"""
    global _last_progress
    try:
        # Read actual counts from Excel when processing OR just completed/stopped
        if processing_state['status'] in ('processing', 'completed', 'stopped'):
            excel_file = processing_state.get('output_file_path') or processing_state.get('file_path')

            if excel_file:
                try:
                    df = pd.read_excel(excel_file, engine='openpyxl')

                    if 'Status' not in df.columns:
                        df['Status'] = ''

                    total = len(df)
                    success = len(df[df['Status'].str.upper() == 'SUCCESS'])
                    failed = len(df[df['Status'].str.upper() == 'FAILED'])
                    pending = len(df[(df['Status'].isna()) | (df['Status'] == '') | (df['Status'].str.upper() == 'PENDING')])
                    current = success + failed

                    processing_state['total'] = total
                    processing_state['current'] = current

                    # Live Check mode: only when Step 4 A3 is the active operation
                    live_mode = False
                    live_count = 0
                    missing_count = 0
                    cur_ops = processing_state.get('operations', '')
                    is_a3_only = cur_ops.strip() == 'A3'
                    if is_a3_only and 'Live Check Status' in df.columns:
                        lcs = df['Live Check Status'].fillna('').astype(str).str.strip().str.lower()
                        live_count = int((lcs == 'live').sum())
                        missing_count = int(current - live_count)
                        live_mode = True

                    prog = {
                            'status': processing_state['status'],
                            'current': current,
                            'total': total,
                            'success': success if not live_mode else live_count,
                            'failed': failed if not live_mode else missing_count,
                            'pending': pending,
                            'current_account': processing_state.get('current_account', ''),
                            'step_label': processing_state.get('step_label', ''),
                            'logs': processing_state.get('recent_logs', []),
                            'live_mode': live_mode,
                    }
                    _last_progress = prog  # Cache good read
                    return jsonify({'success': True, 'progress': prog})
                except Exception as read_error:
                    # File locked by worker — return cached values instead of 0s
                    if _last_progress:
                        _last_progress['status'] = processing_state['status']
                        _last_progress['current_account'] = processing_state.get('current_account', '')
                        _last_progress['logs'] = processing_state.get('recent_logs', [])
                        return jsonify({'success': True, 'progress': _last_progress})

        # Check if bulk re-login is running — show its progress on dashboard
        try:
            rl = profile_manager.get_bulk_relogin_status()
            if rl.get('running') or rl.get('status') in ('processing', 'completed'):
                total = rl.get('total', 0)
                done = rl.get('done', 0)
                pending = max(0, total - done)
                return jsonify({
                    'success': True,
                    'progress': {
                        'status': rl['status'],
                        'current': done,
                        'total': total,
                        'success': rl.get('success', 0),
                        'failed': rl.get('failed', 0),
                        'pending': pending,
                        'current_account': rl.get('current_account', ''),
                        'step_label': 'Bulk Re-Login',
                        'logs': processing_state.get('recent_logs', [])[-20:],
                        'job_type': 'bulk_relogin',
                        'report_path': rl.get('report_path', ''),
                        'success': rl.get('success', 0),
                        'failed': rl.get('failed', 0),
                    }
                })
        except Exception:
            pass

        # Check if batch login is running — show its progress on dashboard
        try:
            bl = profile_manager.get_batch_login_progress()
            if bl.get('running') or bl.get('status') in ('processing', 'completed'):
                done = bl.get('success', 0) + bl.get('failed', 0)
                return jsonify({
                    'success': True,
                    'progress': {
                        'status': bl['status'],
                        'current': done,
                        'total': bl.get('total', 0),
                        'success': bl.get('success', 0),
                        'failed': bl.get('failed', 0),
                        'pending': bl.get('pending', 0),
                        'current_account': bl.get('current_account', ''),
                        'step_label': 'Batch Login',
                        'logs': processing_state.get('recent_logs', [])[-20:],
                        'job_type': 'batch_login',
                    }
                })
        except Exception:
            pass

        # Default return (idle state)
        return jsonify({
            'success': True,
            'progress': {
                'status': processing_state['status'],
                'current': processing_state.get('current', 0),
                'total': processing_state.get('total', 0),
                'success': 0,
                'failed': 0,
                'pending': 0,
                'current_account': processing_state.get('current_account', ''),
                'step_label': processing_state.get('step_label', ''),
                'logs': processing_state.get('recent_logs', [])[-20:]
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/reports', methods=['GET'])
def get_reports():
    """Get list of generated reports (MailNexus Pro + legacy).
    Only scans the project /output directory."""
    try:
        scan_dirs = _get_report_scan_dirs()

        if not scan_dirs:
            return jsonify({'success': True, 'reports': []})

        # Scan all directories, de-duplicate by file path
        seen_paths = set()
        reports = []
        for dir_path in scan_dirs:
            d = Path(dir_path)
            if not d.exists():
                continue
            for file in d.rglob('*.xlsx'):
                fpath = str(file.resolve())
                if fpath in seen_paths:
                    continue
                seen_paths.add(fpath)
                is_mailnexus = 'Mailnexus' in file.name or 'mailnexus' in file.name.lower()
                reports.append({
                    'name': file.name,
                    'path': str(file),
                    'size': file.stat().st_size,
                    'modified': file.stat().st_mtime,
                    'type': 'mailnexus' if is_mailnexus else 'legacy',
                })

        # Sort by modified time (newest first)
        reports.sort(key=lambda x: x['modified'], reverse=True)

        return jsonify({'success': True, 'reports': reports})

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/reports/generate', methods=['POST'])
def generate_mailnexus_report():
    """Generate MailNexus Pro report.
    - source='profiles': all profiles with appeal + health status
    - source='appeal':   only appeal tracking data
    - source='health':   only health activity tracking data
    - source='':         legacy — from step-processing output Excel file"""
    try:
        data = request.json or {}
        source = data.get('source', '')

        from shared.report_generator import generate_report, generate_from_excel
        from datetime import datetime as _dt

        def _fmt_dt(iso):
            try:
                return _dt.fromisoformat(iso.replace('Z', '+00:00')).strftime('%d %b %Y %H:%M')
            except Exception:
                return iso

        output_dir = profile_manager._get_storage_path() / 'reports'
        output_dir.mkdir(parents=True, exist_ok=True)

        if source in ('profiles', 'appeal', 'health'):
            profiles = profile_manager.list_profiles()
            if not profiles:
                return jsonify({'success': False, 'message': 'No profiles found.'})

            accounts_data = []

            if source == 'appeal':
                for p in profiles:
                    login_status = p.get('status', 'unknown')
                    ok_str   = '✓' if p.get('last_appeal_ok') else ('✗' if p.get('last_appeal_at') else '—')
                    last_run = _fmt_dt(p['last_appeal_at']) if p.get('last_appeal_at') else 'Never'
                    summary  = p.get('last_appeal_summary', '') or ''
                    history  = p.get('appeal_history', [])
                    history_str = ' | '.join(
                        f"{_fmt_dt(h['date'])} {'✓' if h.get('ok') else '✗'}" for h in history
                    ) if history else 'No history'
                    accounts_data.append({
                        'Email':        p.get('email', ''),
                        'Profile Name': p.get('name', ''),
                        'Login Status': login_status.replace('_', ' ').title(),
                        'Result':       ok_str,
                        'Last Appeal':  last_run,
                        'Summary':      summary,
                        'Run History':  history_str,
                    })
                label = 'appeal'

            elif source == 'health':
                for p in profiles:
                    login_status = p.get('status', 'unknown')
                    ok_str   = '✓' if p.get('last_health_ok') else ('✗' if p.get('last_health_at') else '—')
                    last_run = _fmt_dt(p['last_health_at']) if p.get('last_health_at') else 'Never'
                    done     = p.get('last_health_done', 0)
                    total    = p.get('last_health_total', 0)
                    history  = p.get('health_history', [])
                    history_str = ' | '.join(
                        f"{_fmt_dt(h['date'])} {h.get('done',0)}/{h.get('total',0)}" for h in history
                    ) if history else 'No history'
                    accounts_data.append({
                        'Email':           p.get('email', ''),
                        'Profile Name':    p.get('name', ''),
                        'Login Status':    login_status.replace('_', ' ').title(),
                        'Result':          ok_str,
                        'Last Health Run': last_run,
                        'Activities Done': f"{done}/{total}",
                        'Run History':     history_str,
                    })
                label = 'health'

            else:  # 'profiles' combined
                for p in profiles:
                    login_status = p.get('status', 'unknown')
                    status = 'SUCCESS' if login_status == 'logged_in' else (
                        'FAILED' if login_status == 'login_failed' else 'PENDING')
                    appeal_info = (
                        f"{'✓' if p.get('last_appeal_ok') else '✗'} {_fmt_dt(p['last_appeal_at'])}"
                        + (f" — {p.get('last_appeal_summary','')}" if p.get('last_appeal_summary') else '')
                        if p.get('last_appeal_at') else 'Never'
                    )
                    health_info = (
                        f"{'✓' if p.get('last_health_ok') else '✗'} {_fmt_dt(p['last_health_at'])} — {p.get('last_health_done',0)}/{p.get('last_health_total',0)} activities"
                        if p.get('last_health_at') else 'Never'
                    )
                    accounts_data.append({
                        'Email':                p.get('email', ''),
                        'Profile Name':         p.get('name', ''),
                        'Status':               status,
                        'Login':                login_status.replace('_', ' ').title(),
                        'Last Appeal':          appeal_info,
                        'Last Health Activity': health_info,
                        'Operations Done':      f"Appeal: {appeal_info} | Health: {health_info}",
                    })
                label = 'profiles'

            report_path = generate_report(
                output_dir=str(output_dir),
                accounts_data=accounts_data,
                step_name=label,
            )
            return jsonify({
                'success': True,
                'report_path': str(report_path),
                'message': f'{label.title()} report generated: {Path(report_path).name} ({len(accounts_data)} accounts)'
            })

        elif source == 'file':
            # ── Generate from a specific raw output file ──
            file_path = data.get('file_path', '')
            if not file_path or not Path(file_path).exists():
                return jsonify({'success': False, 'message': 'File not found.'})
            report_path = generate_from_excel(file_path)
            return jsonify({
                'success': True,
                'report_path': str(report_path),
                'message': f'Pro report generated: {Path(report_path).name}'
            })

        else:
            # ── Legacy: generate from latest output Excel ──
            output_file = (
                data.get('output_file')
                or processing_state.get('output_file_path')
                or _find_latest_output_file()
            )
            step_name = data.get('step', '')

            if not output_file or not Path(output_file).exists():
                return jsonify({'success': False, 'message': 'No output file found.'})

            report_path = generate_from_excel(output_file, step_name=step_name)
            return jsonify({
                'success': True,
                'report_path': str(report_path),
                'message': f'Pro report generated: {Path(report_path).name}'
            })

    except PermissionError:
        return jsonify({'success': False, 'message': 'Report file is open in Excel. Close it and try again.'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Report generation failed: {str(e)}'})


@app.route('/api/template/generate', methods=['POST'])
def generate_step_template():
    """Generate a blank XLS template for a step showing required input columns."""
    try:
        data = request.json or {}
        step = data.get('step')
        if not step:
            return jsonify({'success': False, 'message': 'Missing "step" parameter (1-4).'})

        step_name = f'step{int(step)}'
        output_dir = str(RESOURCES_PATH / 'templates')

        from shared.report_generator import generate_template
        template_path = generate_template(step_name, output_dir)

        return jsonify({
            'success': True,
            'template_path': str(template_path),
            'message': f'Template generated: {Path(template_path).name}'
        })
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Template generation failed: {str(e)}'})


def _auto_generate_report():
    """Auto-generate MailNexus Pro report from the output file after processing."""
    output_file = processing_state.get('output_file_path')
    if not output_file or not Path(output_file).exists():
        add_log('[REPORT] No output file available for report generation.', 'warning')
        return
    try:
        from shared.report_generator import generate_from_excel
        step_name = processing_state.get('step_name', '')
        # For multi-step (e.g. 'step1+step2'), use first step for column filtering
        if '+' in step_name:
            step_name = step_name.split('+')[0]
        report_path = generate_from_excel(output_file, step_name=step_name)
        add_log(f'[REPORT] MailNexus Pro report: {Path(report_path).name}', 'success')
        add_log(f'[REPORT] Saved to: {report_path}', 'info')

    except Exception as e:
        add_log(f'[REPORT] Auto-report generation failed: {e}', 'error')


def _build_step_cmd(bot_step, file_path, num_workers):
    """Build subprocess command for a single step."""
    import sys
    step_scripts = {
        1: 'gmail_bot_step1.py',
        2: 'gmail_bot_step2.py',
        3: 'gmail_bot_step3.py',
        4: 'gmail_bot_step4.py',
    }
    if getattr(sys, 'frozen', False):
        cmd = [sys.executable, f'--step{int(bot_step)}', file_path, str(num_workers)]
        cmd_display = f'backend.exe --step{int(bot_step)} ...'
    else:
        script_name = step_scripts.get(int(bot_step), 'gmail_bot_step2.py')
        script_path = RESOURCES_PATH / script_name
        if not script_path.exists():
            return None, None, f'Script not found: {script_path}'
        cmd = [sys.executable, str(script_path), file_path, str(num_workers)]
        cmd_display = f'python {script_name} ...'
    return cmd, cmd_display, None


def _build_linked_cmd(file_path, num_workers, steps_json):
    """Build subprocess command for linked multi-step mode."""
    import sys
    if getattr(sys, 'frozen', False):
        cmd = [sys.executable, '--linked', file_path, str(num_workers), steps_json]
        cmd_display = f'backend.exe --linked ...'
    else:
        script_path = RESOURCES_PATH / 'gmail_bot_linked.py'
        if not script_path.exists():
            return None, None, f'Script not found: {script_path}'
        cmd = [sys.executable, str(script_path), file_path, str(num_workers), steps_json]
        cmd_display = f'python gmail_bot_linked.py ...'
    return cmd, cmd_display, None


def _spawn_and_stream(cmd, cmd_display, label=''):
    """Spawn a subprocess and stream its output to the UI log. Returns exit code."""
    import subprocess

    add_log(f'[DEBUG] Command: {cmd_display}', 'info')
    add_log(f'[DEBUG] About to spawn process...', 'info')

    try:
        unbuffered_env = {**os.environ, 'PYTHONUNBUFFERED': '1'}
        process = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
            encoding='utf-8',
            errors='replace',
            env=unbuffered_env,
        )
        add_log(f'[DEBUG] Process spawned with PID: {process.pid}', 'info')
    except Exception as spawn_error:
        add_log(f'[ERROR] Failed to spawn process: {spawn_error}', 'error')
        return -1

    line_count = 0
    for line in iter(process.stdout.readline, ''):
        if stop_flag.is_set():
            process.terminate()
            add_log(f'Processing stopped by user{" (" + label + ")" if label else ""}', 'warning')
            break

        if line:
            line_count += 1
            line = line.strip()

            if '[OUTPUT_FILE]' in line:
                output_path = line.split('[OUTPUT_FILE]')[1].strip()
                # Resolve to absolute path (subprocess may print relative)
                output_path = str(Path(output_path).resolve())
                processing_state['output_file_path'] = output_path
                # Persist to disk so the path survives server restarts
                _persist_output_dirs(
                    output_file_path=output_path,
                    input_file_path=processing_state.get('file_path', ''),
                )
                add_log(f'Output file: {output_path}', 'info')
                continue

            if 'ERROR' in line or 'FAILED' in line:
                add_log(line, 'error')
            elif 'SUCCESS' in line:
                add_log(line, 'success')
            elif 'WARNING' in line:
                add_log(line, 'warning')
            else:
                add_log(line, 'info')

            if '[WORKER' in line and 'Processing:' in line:
                parts = line.split('Processing:')
                if len(parts) > 1:
                    email = parts[1].strip().split()[0]
                    processing_state['current_account'] = email

    add_log(f'[DEBUG] Finished reading output. Total lines: {line_count}', 'info')
    return_code = process.wait()
    add_log(f'[DEBUG] Process exited with code: {return_code}', 'info')
    return return_code


def run_processing_worker(file_path, num_workers, bot_step=2, bot_steps=None,
                          linked=False, ops_per_step=None):
    """Background worker to run production script(s).

    Modes:
      1. Single step        — spawn gmail_bot_stepN.py
      2. Multi-step linked  — spawn gmail_bot_linked.py (one session per account)
      3. Multi-step unlinked — spawn each step's script sequentially
    """
    import json

    if bot_steps is None:
        bot_steps = [bot_step]
    if ops_per_step is None:
        ops_per_step = {}

    is_multi = len(bot_steps) > 1

    try:
        if not Path(file_path).exists():
            add_log(f'ERROR: Excel file not found at {file_path}', 'error')
            processing_state['status'] = 'idle'
            return

        add_log(f'Starting {num_workers} workers...', 'info')
        add_log(f'[DEBUG] Excel file: {file_path}', 'info')

        if is_multi and linked:
            # ── LINKED MODE: single process, all steps in one session ──────
            steps_json = json.dumps({
                'steps': bot_steps,
                'ops_per_step': ops_per_step,
            })
            cmd, cmd_display, err = _build_linked_cmd(file_path, num_workers, steps_json)
            if err:
                add_log(f'ERROR: {err}', 'error')
                processing_state['status'] = 'idle'
                return

            return_code = _spawn_and_stream(cmd, cmd_display, label='linked')

            if return_code == 0:
                add_log('Processing completed successfully', 'success')
                processing_state['status'] = 'completed'
            else:
                add_log(f'Processing failed with exit code: {return_code}', 'error')
                processing_state['status'] = 'idle'

        elif is_multi and not linked:
            # ── UNLINKED MODE: run each step sequentially ──────────────────
            all_ok = True
            for step_num in bot_steps:
                if stop_flag.is_set():
                    add_log('Processing stopped by user before next step', 'warning')
                    all_ok = False
                    break

                add_log(f'--- Starting Step {step_num} ---', 'info')
                cmd, cmd_display, err = _build_step_cmd(step_num, file_path, num_workers)
                if err:
                    add_log(f'ERROR: {err}', 'error')
                    all_ok = False
                    continue

                return_code = _spawn_and_stream(cmd, cmd_display, label=f'Step {step_num}')
                if return_code != 0:
                    add_log(f'Step {step_num} failed with exit code: {return_code}', 'error')
                    all_ok = False
                else:
                    add_log(f'Step {step_num} completed successfully', 'success')

                # Reset statuses for next step so accounts are re-processed
                # (Each step script picks up PENDING rows)

            if all_ok:
                add_log('All steps completed successfully', 'success')
                processing_state['status'] = 'completed'
            else:
                processing_state['status'] = 'idle'

        else:
            # ── SINGLE STEP MODE (original behavior) ──────────────────────
            cmd, cmd_display, err = _build_step_cmd(bot_steps[0], file_path, num_workers)
            if err:
                add_log(f'ERROR: {err}', 'error')
                processing_state['status'] = 'idle'
                return

            return_code = _spawn_and_stream(cmd, cmd_display)

            if return_code == 0:
                add_log('Processing completed successfully', 'success')
                processing_state['status'] = 'completed'
            else:
                add_log(f'Processing failed with exit code: {return_code}', 'error')
                processing_state['status'] = 'idle'

        # ── Auto-generate MailNexus Pro report ────────────────────────────
        _auto_generate_report()

    except Exception as e:
        import traceback
        traceback.print_exc()
        processing_state['status'] = 'idle'
        add_log(f'Processing error: {str(e)}', 'error')
        _auto_generate_report()


@app.route('/api/log-stream')
def log_stream():
    """
    Server-Sent Events endpoint for real-time log streaming.
    The browser connects once; the server pushes each new log entry
    within ~100 ms of it being generated — no polling delay.
    """
    # Only replay logs that arrived AFTER the last clear operation so that
    # a page reload never brings back entries the user already cleared.
    clear_id = processing_state.get('log_clear_id', 0)
    snapshot = [lg for lg in processing_state['recent_logs'] if lg['id'] > clear_id]
    # Use clear_id as the floor so the live-stream loop also ignores old ids
    start_id = snapshot[-1]['id'] if snapshot else clear_id

    def generate():
        # ── Replay logs the client missed before connecting ───────────────
        for log in snapshot:
            yield f"data: {json.dumps(log)}\n\n"

        last_id = start_id
        idle_cycles = 0  # track consecutive empty polls
        max_cycles = 18000  # ~30 min at 100ms intervals → auto-close stale connections

        # ── Stream live logs as they arrive ──────────────────────────────
        while idle_cycles < max_cycles:
            try:
                new_logs = [
                    lg for lg in processing_state['recent_logs']
                    if lg['id'] > last_id
                ]
                for lg in new_logs:
                    yield f"data: {json.dumps(lg)}\n\n"
                    last_id = lg['id']
                    idle_cycles = 0

                if not new_logs:
                    idle_cycles += 1
                    # Send keepalive every ~3 seconds (not every 100ms)
                    if idle_cycles % 30 == 0:
                        yield ": keepalive\n\n"

                time.sleep(0.1)
            except GeneratorExit:
                # Client disconnected — clean exit
                return

    return Response(
        generate(),
        mimetype='text/event-stream',
        headers={
            'Cache-Control': 'no-cache',
            'X-Accel-Buffering': 'no',
            'Connection': 'keep-alive',
        }
    )


@app.route('/api/config', methods=['GET'])
def get_config():
    """Read config/settings.json and config/urls.json"""
    try:
        cfg_path = RESOURCES_PATH / 'config' / 'settings.json'
        url_path = RESOURCES_PATH / 'config' / 'urls.json'

        settings = {}
        urls = {}

        if cfg_path.exists():
            try:
                with open(cfg_path, 'r', encoding='utf-8') as f:
                    settings = json.load(f)
            except (json.JSONDecodeError, ValueError):
                settings = {}

        if url_path.exists():
            try:
                with open(url_path, 'r', encoding='utf-8') as f:
                    urls = json.load(f)
            except (json.JSONDecodeError, ValueError):
                urls = {}

        return jsonify({'success': True, 'settings': settings, 'urls': urls})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/config', methods=['POST'])
def save_config():
    """Write config/settings.json and config/urls.json"""
    try:
        data = request.json or {}
        cfg_path = RESOURCES_PATH / 'config' / 'settings.json'
        url_path = RESOURCES_PATH / 'config' / 'urls.json'

        if 'settings' in data:
            with open(cfg_path, 'w', encoding='utf-8') as f:
                json.dump(data['settings'], f, indent=2)

        if 'urls' in data:
            with open(url_path, 'w', encoding='utf-8') as f:
                json.dump(data['urls'], f, indent=2)

        return jsonify({'success': True, 'message': 'Configuration saved'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/reports/all', methods=['DELETE'])
def delete_all_reports():
    """Delete all .xlsx files inside the app output folders ONLY.
    Never touches user's Downloads or any other directory."""
    try:
        scan_dirs = _get_report_scan_dirs()
        if not scan_dirs:
            return jsonify({'success': True, 'deleted': 0})

        count = 0
        errors = []
        for dir_path in scan_dirs:
            d = Path(dir_path)
            if not d.exists():
                continue
            for f in d.rglob('*.xlsx'):
                # Extra safety: ensure file is actually inside the output dir
                try:
                    f.resolve().relative_to(d.resolve())
                except ValueError:
                    continue
                try:
                    f.unlink()
                    count += 1
                except Exception as e:
                    errors.append(f'{f.name}: {e}')

        if errors:
            return jsonify({'success': False, 'deleted': count,
                            'message': f'Deleted {count}, but {len(errors)} error(s): {"; ".join(errors[:3])}'})
        return jsonify({'success': True, 'deleted': count, 'message': f'Deleted {count} report(s)'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/reports/single', methods=['DELETE'])
def delete_single_report():
    """Delete a single report file by its absolute path.
    Only allows deletion from the project /output folder."""
    try:
        data = request.json or {}
        file_path = data.get('path', '')

        if not file_path:
            return jsonify({'success': False, 'message': 'No file path provided.'})

        fp = Path(file_path)
        if not fp.exists():
            return jsonify({'success': False, 'message': 'File not found.'})

        if not fp.suffix.lower() == '.xlsx':
            return jsonify({'success': False, 'message': 'Only .xlsx files can be deleted.'})

        # Security: only allow deletion from known output folders
        scan_dirs = _get_report_scan_dirs()
        resolved = fp.resolve()
        in_known_dir = any(
            _is_safe_child(resolved, Path(sd))
            for sd in scan_dirs
        )
        if not in_known_dir:
            return jsonify({'success': False, 'message': 'File is not in a known output directory.'})

        fp.unlink()
        return jsonify({'success': True, 'message': f'Deleted: {fp.name}'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/logs/clear', methods=['POST'])
def clear_logs():
    """Clear all in-memory logs and set a clear threshold so SSE never
    replays old entries on page reload."""
    logs = processing_state['recent_logs']
    if logs:
        processing_state['log_clear_id'] = logs[-1]['id']
    processing_state['recent_logs'] = []
    return jsonify({'success': True, 'message': 'Logs cleared',
                    'clear_id': processing_state['log_clear_id']})


@app.route('/api/sms-code', methods=['POST'])
def receive_sms_code():
    """Receive an SMS verification code from the phone relay app."""
    try:
        data = request.get_json(force=True, silent=True) or {}
        code = str(data.get('code', '')).strip()
        if not code:
            return jsonify({'success': False, 'error': 'No code provided'}), 400

        sms_codes.append({
            'code': code,
            'sender': data.get('sender', ''),
            'full_message': data.get('full_message', ''),
            'timestamp': time.time(),
            'used': False
        })
        # Keep only last 20 codes
        if len(sms_codes) > 20:
            sms_codes[:] = sms_codes[-20:]

        add_log(f'[SMS] Code received: {code} from {data.get("sender", "unknown")}', 'info')
        return jsonify({'success': True, 'message': f'Code {code} stored'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/sms-code', methods=['GET'])
def get_sms_code():
    """Get the latest unused SMS verification code.
    ?peek=1 returns without marking as used (for UI display).
    ?max_age=120 limits how old a code can be (seconds)."""
    try:
        max_age = float(request.args.get('max_age', 120))
        peek = request.args.get('peek', '0') == '1'
        now = time.time()
        for entry in reversed(sms_codes):
            if not entry['used'] and (now - entry['timestamp']) < max_age:
                if not peek:
                    entry['used'] = True
                return jsonify({
                    'success': True,
                    'code': entry['code'],
                    'sender': entry.get('sender', ''),
                    'timestamp': entry['timestamp']
                })
        return jsonify({'success': False, 'code': None})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/fingerprint', methods=['GET'])
def get_fingerprint():
    """Read config/fingerprint.json and return fingerprint settings."""
    try:
        fp_path = RESOURCES_PATH / 'config' / 'fingerprint.json'
        if not fp_path.exists():
            return jsonify({'success': True, 'fingerprint': {'os_type': 'random', 'auto_timezone': True}})
        with open(fp_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        return jsonify({'success': True, 'fingerprint': data})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/fingerprint', methods=['POST'])
def save_fingerprint():
    """Write fingerprint settings to config/fingerprint.json."""
    try:
        data = request.get_json(force=True, silent=True) or {}
        _valid_os = ('random', 'windows', 'macos', 'linux', 'android')
        os_type      = str(data.get('os_type', 'random')).lower()
        os_type      = os_type if os_type in _valid_os else 'random'
        auto_timezone = bool(data.get('auto_timezone', True))

        fp_path = RESOURCES_PATH / 'config' / 'fingerprint.json'
        fp_path.parent.mkdir(parents=True, exist_ok=True)
        with open(fp_path, 'w', encoding='utf-8') as f:
            json.dump({'os_type': os_type, 'auto_timezone': auto_timezone}, f, indent=2)

        tz_label = 'Auto from IP (geo-lookup)' if auto_timezone else 'Random pool'
        return jsonify({
            'success': True,
            'message': f'Fingerprint saved — OS: {os_type} | Timezone: {tz_label}'
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/debug/launch', methods=['POST'])
def debug_launch():
    """Launch debug browser(s) with current proxy + fingerprint for manual inspection."""
    try:
        add_log('[DEBUG] Importing debug_launcher...', 'info')
        from shared import debug_launcher
        # Connect debug_launcher logs → UI log panel so user sees everything live
        debug_launcher.set_ui_logger(add_log)

        data         = request.json or {}
        num_browsers = max(1, min(int(data.get('num_browsers', 1)), 10))
        test_url     = str(data.get('test_url', 'https://ipinfo.io')).strip() or 'https://ipinfo.io'

        add_log(f'[DEBUG] Calling launch({num_browsers}, {test_url})...', 'info')
        debug_launcher.launch(num_browsers=num_browsers, test_url=test_url)
        add_log(f'[DEBUG] launch() returned — threads are starting...', 'info')

        return jsonify({
            'success': True,
            'message': f'Launched {num_browsers} debug browser(s). Check logs below for IP/proxy/fingerprint details.',
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        add_log(f'[DEBUG-LAUNCH ERROR] {e}', 'error')
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/debug/close', methods=['POST'])
def debug_close():
    """Close all open debug browsers."""
    try:
        from shared import debug_launcher
        debug_launcher.close_all()
        return jsonify({'success': True, 'message': 'Close signal sent — debug browsers shutting down.'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/debug/status', methods=['GET'])
def debug_status():
    """Return current debug browser status."""
    try:
        from shared import debug_launcher
        s = debug_launcher.status()
        return jsonify({'success': True, **s})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e), 'open': 0, 'total': 0, 'running': False})


def add_log(message, log_type='info'):
    """Add a log entry — visible immediately via SSE stream."""
    log_entry = {
        'id': next(_log_id),          # collision-free incrementing ID
        'message': message,
        'type': log_type,
        'timestamp': time.time()
    }

    processing_state['recent_logs'].append(log_entry)

    # Keep only last 500 logs in memory
    if len(processing_state['recent_logs']) > 500:
        processing_state['recent_logs'] = processing_state['recent_logs'][-500:]

    try:
        print(f"[{log_type.upper()}] {message}")
    except UnicodeEncodeError:
        safe_msg = message.encode(sys.stdout.encoding or 'utf-8', errors='replace').decode(sys.stdout.encoding or 'utf-8')
        print(f"[{log_type.upper()}] {safe_msg}")


# ── Auth Endpoints (licensing removed) ───────────────────────────────────────

@app.route('/api/auth/status', methods=['GET'])
def auth_status():
    return jsonify({'success': True, 'auth_enabled': False, 'license_activated': True})


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# TOOLS API — Screenshots, Auth files, Garbage cleaner
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

@app.route('/api/tools/screenshots', methods=['GET'])
def tools_screenshots():
    """List .png screenshot files with pagination and optional search."""
    try:
        search = request.args.get('search', '').lower()
        page_num = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))

        if not SCREENSHOTS_PATH.exists():
            return jsonify({'success': True, 'files': [], 'total': 0, 'page': page_num})

        all_files = []
        for f in SCREENSHOTS_PATH.iterdir():
            if f.suffix.lower() == '.png':
                if search and search not in f.name.lower():
                    continue
                stat = f.stat()
                all_files.append({
                    'name': f.name,
                    'size': stat.st_size,
                    'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M'),
                    '_sort_time': stat.st_mtime,
                })

        all_files.sort(key=lambda x: x['_sort_time'], reverse=True)
        # Remove internal sort key before sending
        for f in all_files:
            f.pop('_sort_time', None)
        total = len(all_files)
        start = (page_num - 1) * per_page
        page_files = all_files[start:start + per_page]

        return jsonify({
            'success': True, 'files': page_files, 'total': total,
            'page': page_num, 'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page if total > 0 else 0
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/tools/screenshot/<filename>', methods=['GET'])
def tools_screenshot_image(filename):
    """Serve a single screenshot image file."""
    if '..' in filename or '/' in filename or '\\' in filename:
        return jsonify({'success': False, 'message': 'Invalid filename'}), 400
    if not filename.lower().endswith('.png'):
        return jsonify({'success': False, 'message': 'Only PNG files'}), 400
    filepath = SCREENSHOTS_PATH / filename
    if not filepath.exists():
        return jsonify({'success': False, 'message': 'Not found'}), 404
    return send_from_directory(str(SCREENSHOTS_PATH), filename, mimetype='image/png')


@app.route('/api/tools/auth-files', methods=['GET'])
def tools_auth_files():
    """List authenticator_key_*.txt and backup_codes_*.txt files."""
    try:
        search = request.args.get('search', '').lower()
        page_num = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 50))

        if not SCREENSHOTS_PATH.exists():
            return jsonify({'success': True, 'files': [], 'total': 0, 'page': page_num})

        all_files = []
        for f in SCREENSHOTS_PATH.iterdir():
            if f.suffix.lower() == '.txt' and (
                f.name.startswith('authenticator_key_') or
                f.name.startswith('backup_codes_')
            ):
                if search and search not in f.name.lower():
                    continue
                stat = f.stat()
                all_files.append({
                    'name': f.name,
                    'size': stat.st_size,
                    'modified': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M'),
                    '_sort_time': stat.st_mtime,
                    'type': 'authenticator' if f.name.startswith('authenticator_key_') else 'backup',
                })

        all_files.sort(key=lambda x: x['_sort_time'], reverse=True)
        for f in all_files:
            f.pop('_sort_time', None)
        total = len(all_files)
        start = (page_num - 1) * per_page
        page_files = all_files[start:start + per_page]

        return jsonify({
            'success': True, 'files': page_files, 'total': total,
            'page': page_num, 'per_page': per_page,
            'total_pages': (total + per_page - 1) // per_page if total > 0 else 0
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/tools/auth-file/<filename>', methods=['GET'])
def tools_auth_file_content(filename):
    """Read and return content of an authenticator/backup text file."""
    if '..' in filename or '/' in filename or '\\' in filename:
        return jsonify({'success': False, 'message': 'Invalid filename'}), 400
    if not filename.lower().endswith('.txt'):
        return jsonify({'success': False, 'message': 'Only TXT files'}), 400
    filepath = SCREENSHOTS_PATH / filename
    if not filepath.exists():
        return jsonify({'success': False, 'message': 'Not found'}), 404
    try:
        content = filepath.read_text(encoding='utf-8')
        return jsonify({'success': True, 'filename': filename, 'content': content})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/tools/storage-stats', methods=['GET'])
def tools_storage_stats():
    """Return storage statistics for screenshots, txt files, and backend log."""
    try:
        stats = {
            'screenshots': {'count': 0, 'total_size': 0},
            'authenticator': {'count': 0, 'total_size': 0},
            'backup_codes': {'count': 0, 'total_size': 0},
            'log': {'count': 0, 'total_size': 0},
        }

        if SCREENSHOTS_PATH.exists():
            for f in SCREENSHOTS_PATH.iterdir():
                try:
                    sz = f.stat().st_size
                except OSError:
                    continue
                if f.suffix.lower() == '.png':
                    stats['screenshots']['count'] += 1
                    stats['screenshots']['total_size'] += sz
                elif f.name.startswith('authenticator_key_') and f.suffix == '.txt':
                    stats['authenticator']['count'] += 1
                    stats['authenticator']['total_size'] += sz
                elif f.name.startswith('backup_codes_') and f.suffix == '.txt':
                    stats['backup_codes']['count'] += 1
                    stats['backup_codes']['total_size'] += sz

        # Backend log in AppData
        log_path = Path(os.environ.get('APPDATA', '')) / 'gmail-bot-pro' / 'backend.log'
        if log_path.exists():
            stats['log']['count'] = 1
            stats['log']['total_size'] = log_path.stat().st_size

        return jsonify({'success': True, 'stats': stats})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/tools/cleanup', methods=['POST'])
def tools_cleanup():
    """Delete files by category: screenshots, authenticator, backup_codes, log."""
    try:
        data = request.get_json(force=True, silent=True) or {}
        category = data.get('category', '')
        deleted = 0
        freed = 0

        if category == 'screenshots' and SCREENSHOTS_PATH.exists():
            for f in list(SCREENSHOTS_PATH.iterdir()):
                if f.suffix.lower() == '.png':
                    freed += f.stat().st_size
                    f.unlink()
                    deleted += 1

        elif category == 'authenticator' and SCREENSHOTS_PATH.exists():
            for f in list(SCREENSHOTS_PATH.iterdir()):
                if f.name.startswith('authenticator_key_') and f.suffix == '.txt':
                    freed += f.stat().st_size
                    f.unlink()
                    deleted += 1

        elif category == 'backup_codes' and SCREENSHOTS_PATH.exists():
            for f in list(SCREENSHOTS_PATH.iterdir()):
                if f.name.startswith('backup_codes_') and f.suffix == '.txt':
                    freed += f.stat().st_size
                    f.unlink()
                    deleted += 1

        elif category == 'log':
            log_path = Path(os.environ.get('APPDATA', '')) / 'gmail-bot-pro' / 'backend.log'
            if log_path.exists():
                freed += log_path.stat().st_size
                log_path.unlink()
                deleted += 1

        else:
            return jsonify({'success': False, 'message': f'Unknown category: {category}'})

        return jsonify({
            'success': True,
            'message': f'Deleted {deleted} file(s)',
            'deleted': deleted,
            'freed_bytes': freed
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PROFILE MANAGER API
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

from shared import nexus_profile_manager as profile_manager, recovery_tracker
profile_manager.init(RESOURCES_PATH)
profile_manager.set_ui_logger(add_log)
recovery_tracker.init(RESOURCES_PATH)

# Safety net: close all managed browsers when backend exits
# This ensures Chrome processes we launched don't become orphans.
# Does NOT touch the user's own Chrome browser.
def _cleanup_browsers_on_exit():
    try:
        profile_manager.close_all_profiles()
    except Exception:
        pass
atexit.register(_cleanup_browsers_on_exit)


@app.route('/api/profiles/groups', methods=['GET'])
def profiles_list_groups():
    """Return all unique profile group names with counts."""
    profiles = profile_manager.list_profiles()
    from collections import Counter
    counts = Counter()
    for p in profiles:
        for g in profile_manager._get_groups(p):
            counts[g] += 1
    groups = sorted(counts.keys())
    return jsonify({'success': True, 'groups': groups, 'counts': dict(counts)})


@app.route('/api/profiles/bulk-assign-group', methods=['POST'])
def profiles_bulk_assign_group():
    data = request.get_json() or {}
    ids = data.get('ids', [])
    group = (data.get('group') or 'default').strip()
    mode = data.get('mode', 'add')  # 'add' or 'set'
    note = (data.get('note') or '').strip()
    if not ids:
        return jsonify({'success': False, 'message': 'No profiles selected'})
    updated = profile_manager.bulk_assign_group(ids, group, mode=mode)
    notes_updated = 0
    if note:
        for pid in ids:
            try:
                profile_manager.update_profile(pid, notes=note)
                notes_updated += 1
            except Exception:
                pass
    return jsonify({'success': True, 'updated': updated, 'notes_updated': notes_updated})


@app.route('/api/profiles/bulk-remove-group', methods=['POST'])
def profiles_bulk_remove_group():
    data = request.get_json() or {}
    ids = data.get('ids', [])
    group = (data.get('group') or '').strip()
    note = (data.get('note') or '').strip()
    if not ids or not group:
        return jsonify({'success': False, 'message': 'ids and group required'})
    updated = profile_manager.remove_profile_from_group(ids, group)
    notes_updated = 0
    if note:
        for pid in ids:
            try:
                profile_manager.update_profile(pid, notes=note)
                notes_updated += 1
            except Exception:
                pass
    return jsonify({'success': True, 'updated': updated, 'notes_updated': notes_updated})


@app.route('/api/profiles/bulk-update-notes', methods=['POST'])
def profiles_bulk_update_notes():
    """Update notes on multiple profiles at once without changing groups."""
    data = request.get_json() or {}
    ids = data.get('ids', [])
    note = (data.get('note') or '').strip()
    if not ids:
        return jsonify({'success': False, 'message': 'No profiles selected'})
    if not note:
        return jsonify({'success': False, 'message': 'Note is empty'})
    updated = 0
    for pid in ids:
        try:
            profile_manager.update_profile(pid, notes=note)
            updated += 1
        except Exception:
            pass
    return jsonify({'success': True, 'updated': updated})


@app.route('/api/profiles/bulk-update-proxy', methods=['POST'])
def profiles_bulk_update_proxy():
    """Update proxy username and/or password for multiple profiles at once."""
    data = request.get_json() or {}
    ids = data.get('ids', [])
    proxy_user = (data.get('proxy_user') or '').strip()
    proxy_pass = (data.get('proxy_pass') or '').strip()
    if not ids:
        return jsonify({'success': False, 'message': 'No profiles selected'})
    if not proxy_user and not proxy_pass:
        return jsonify({'success': False, 'message': 'Provide at least proxy user or password'})
    updated = 0
    for pid in ids:
        try:
            p = profile_manager.get_profile(pid)
            if not p:
                continue
            proxy = dict(p.get('proxy') or {})
            if proxy_user:
                proxy['username'] = proxy_user
            if proxy_pass:
                proxy['password'] = proxy_pass
            profile_manager.update_profile(pid, proxy=proxy)
            updated += 1
        except Exception:
            pass
    return jsonify({'success': True, 'updated': updated})


@app.route('/api/profiles/groups/rename', methods=['POST'])
def profiles_rename_group():
    data = request.get_json() or {}
    old_name = (data.get('old_name') or '').strip()
    new_name = (data.get('new_name') or '').strip()
    if not old_name or not new_name:
        return jsonify({'success': False, 'message': 'old_name and new_name are required'})
    updated = profile_manager.rename_group(old_name, new_name)
    return jsonify({'success': True, 'updated': updated})


@app.route('/api/profiles/groups/<path:group_name>', methods=['DELETE'])
def profiles_delete_group(group_name):
    data = request.get_json() or {}
    reassign_to = (data.get('reassign_to') or 'default').strip()
    updated = profile_manager.delete_group(group_name, reassign_to)
    return jsonify({'success': True, 'updated': updated})


@app.route('/api/profiles/export-excel', methods=['POST'])
def profiles_export_excel():
    """Export selected profiles to Excel — full credentials + Write Review columns."""
    import io
    import json as _json
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    data = request.get_json(force=True, silent=True) or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'No profile IDs provided'})

    # Use _read_profiles() directly so password/totp/backup_codes are included
    all_profiles = profile_manager._read_profiles()
    id_set = set(ids)
    selected = [p for p in all_profiles if p.get('id') in id_set]
    if not selected:
        return jsonify({'success': False, 'message': 'No matching profiles found'})

    wb = openpyxl.Workbook()

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Sheet 1 — Credentials
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws = wb.active
    ws.title = 'Credentials'

    thin       = Side(style='thin', color='CBD5E1')
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font   = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    hdr_fill   = PatternFill('solid', fgColor='1E3A5F')
    hdr_align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(horizontal='left', vertical='center')

    # Divider style — separates credential block from write-review block
    div_fill   = PatternFill('solid', fgColor='1E3A5F')
    div_font   = Font(name='Calibri', bold=True, color='FCD34D', size=10)

    status_colors = {
        'logged_in':    'D1FAE5',
        'login_failed': 'FEE2E2',
        'pending':      'FEF9C3',
    }

    # Column layout: (header, width)
    cred_cols = [
        ('Name',          20),
        ('Email',         30),
        ('Password',      22),
        ('TOTP Secret',   28),
        ('Backup Codes',  45),
        ('Login Status',  15),
        ('Groups',        18),
        ('Engine',        10),
        ('Proxy',         38),
        ('Address',       35),
        ('Notes',         30),
        ('Created At',    18),
    ]
    wr_cols = [
        ('GMB URL',      40),
        ('Review Text',  50),
        ('Review Stars', 14),
    ]
    all_cols = cred_cols + [('', 3)] + wr_cols  # empty column as visual divider

    for col_idx, (h, w) in enumerate(all_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        ws.column_dimensions[get_column_letter(col_idx)].width = w
        if h == '':
            # Divider column — dark fill, no text
            cell.fill = PatternFill('solid', fgColor='334155')
            continue
        if col_idx <= len(cred_cols):
            cell.font = hdr_font
            cell.fill = hdr_fill
        else:
            # Write Review columns — amber header
            cell.font = Font(name='Calibri', bold=True, color='1C1917', size=11)
            cell.fill = PatternFill('solid', fgColor='FCD34D')
        cell.alignment = hdr_align
        cell.border = border

    ws.row_dimensions[1].height = 24

    # ── Data rows ─────────────────────────────────────────────────────────
    for row_idx, p in enumerate(selected, 2):
        status = p.get('status', '')
        groups = ', '.join(profile_manager._get_groups(p))

        # Proxy → readable string  host:port (user:pass)
        proxy = p.get('proxy') or {}
        if proxy and proxy.get('host'):
            host = proxy.get('host', '')
            port = proxy.get('port', '')
            pu   = proxy.get('username', '')
            pp   = proxy.get('password', '')
            ptype = proxy.get('type', 'http')
            proxy_str = f"{ptype}://{host}:{port}"
            if pu or pp:
                proxy_str += f"  ({pu}:{pp})"
        else:
            proxy_str = ''

        # Backup codes → one per line
        backup_codes = p.get('backup_codes') or []
        if isinstance(backup_codes, list):
            codes_str = '\n'.join(str(c) for c in backup_codes if c)
        else:
            codes_str = str(backup_codes)

        row_data = [
            p.get('name', ''),
            p.get('email', ''),
            p.get('password', ''),
            p.get('totp_secret', ''),
            codes_str,
            status,
            groups,
            (p.get('engine', 'nexus') or 'nexus').upper(),
            proxy_str,
            p.get('address', ''),
            p.get('notes', ''),
            p.get('created_at', ''),
            # divider column
            '',
            # Write Review (blank — user fills)
            '',
            '',
            '',
        ]

        row_fill_color = status_colors.get(status, 'FFFFFF')
        row_fill = PatternFill('solid', fgColor=row_fill_color)

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            # Divider column
            if col_idx == len(cred_cols) + 1:
                cell.fill = PatternFill('solid', fgColor='334155')
                continue
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            if col_idx <= len(cred_cols):
                cell.fill = row_fill
            # Backup codes column — slightly larger row height handled below

        # Row height — taller if there are backup codes
        n_codes = len(backup_codes) if isinstance(backup_codes, list) else 1
        ws.row_dimensions[row_idx].height = max(18, min(14 * n_codes, 90))

    ws.freeze_panes = 'B2'  # freeze Name column + header

    # ── Section label above Write Review columns ──────────────────────────
    wr_label_col = len(cred_cols) + 2  # first WR column
    label_cell = ws.cell(row=1, column=wr_label_col)
    # Already set above, add a small note below header in row 0 — skip, header is enough

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Sheet 2 — Write Review (email + WR cols only, ready to use directly)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    wr = wb.create_sheet('Write Review')
    wr_headers = [('Email', 30), ('GMB URL', 40), ('Review Text', 50), ('Review Stars', 14)]
    for col_idx, (h, w) in enumerate(wr_headers, 1):
        cell = wr.cell(row=1, column=col_idx, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
        wr.column_dimensions[get_column_letter(col_idx)].width = w
    wr.row_dimensions[1].height = 24

    for row_idx, p in enumerate(selected, 2):
        for col_idx, value in enumerate([p.get('email', ''), '', '', ''], 1):
            cell = wr.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = border
        wr.row_dimensions[row_idx].height = 18

    wr.freeze_panes = 'A2'

    # ── Save ─────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from datetime import datetime as _dt
    date_str = _dt.now().strftime('%Y%m%d_%H%M')
    filename = f'profiles_export_{len(selected)}accs_{date_str}.xlsx'

    from flask import send_file
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/profiles', methods=['GET'])
def profiles_list():
    """List profiles with pagination support."""
    search = request.args.get('search', '').lower()
    status_filter = request.args.get('filter', 'all').lower()
    group_filter = request.args.get('group', '').lower()
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))
    per_page = max(10, min(per_page, 10000))  # allow loading all profiles at once

    profiles = profile_manager.list_profiles()

    # Search filter — also search notes, proxy, and group
    if search:
        search_by = request.args.get('search_by', 'name').lower()
        if search_by == 'email':
            profiles = [p for p in profiles if search in p.get('email', '').lower()]
        elif search_by == 'notes':
            profiles = [p for p in profiles if search in p.get('note', '').lower()]
        elif search_by == 'proxy':
            profiles = [p for p in profiles if search in str(p.get('proxy', {}).get('host', '')).lower()
                        or search in str(p.get('proxy', {}).get('server', '')).lower()]
        elif search_by == 'group':
            profiles = [p for p in profiles if any(search in g.lower() for g in profile_manager._get_groups(p))]
        else:
            profiles = [p for p in profiles if search in p.get('name', '').lower()
                        or search in p.get('email', '').lower()]

    if group_filter:
        profiles = [p for p in profiles if group_filter in [g.lower() for g in profile_manager._get_groups(p)]]

    # Server-side status filter
    if status_filter == 'running':
        profiles = [p for p in profiles if p.get('browser_open') == 'running']
    elif status_filter == 'logged_in':
        profiles = [p for p in profiles if p.get('status') == 'logged_in']
    elif status_filter == 'not_logged_in':
        profiles = [p for p in profiles if p.get('status') not in ('logged_in', 'login_failed')]
    elif status_filter == 'login_failed':
        profiles = [p for p in profiles if p.get('status') == 'login_failed']
    elif status_filter == 'nst':
        profiles = [p for p in profiles if p.get('engine', 'nexus') == 'nst']
    elif status_filter == 'nexus':
        profiles = [p for p in profiles if p.get('engine', 'nexus') == 'nexus']

    total = len(profiles)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page
    end = start + per_page

    return jsonify({
        'success': True,
        'profiles': profiles[start:end],
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': total_pages,
    })


@app.route('/api/profiles/generate-fingerprint', methods=['POST'])
def profiles_generate_fingerprint():
    """Generate a fingerprint for preview (without creating a profile).

    In NST mode, fingerprints are managed by NST Browser — we return
    placeholder values indicating NST will handle it.
    """
    data = request.get_json(force=True, silent=True) or {}
    os_type = data.get('os', 'windows')

    # Check if NST mode is active
    use_nst = False
    try:
        import json as _json
        bj = RESOURCES_PATH / 'config' / 'browser.json'
        if bj.exists():
            bcfg = _json.loads(bj.read_text('utf-8'))
            use_nst = bcfg.get('use_nst', False)
    except Exception:
        pass

    if use_nst:
        # NST Browser handles fingerprints — return defaults with NST marker
        os_map = {'windows': 'Windows 11', 'macos': 'macOS 14', 'linux': 'Linux'}
        fp = {
            'overview': {
                'os': os_type,
                'os_version': os_map.get(os_type, 'Windows 11'),
                'device_type': 'desktop',
                'browser_kernel': 'nstbrowser',
                'kernel_version': 133,
                'user_agent': '(managed by NST Browser)',
                'platform': 'nst',
                'startup_urls': [],
            },
            'hardware': {
                'webgl': 'noise', 'webgl_metadata': 'masked',
                'webgl_vendor': '(managed by NST Browser)',
                'webgl_renderer': '(managed by NST Browser)',
                'canvas': 'noise', 'canvas_seed': 0,
                'audio_context': 'noise', 'audio_seed': 0,
                'client_rects': 'real', 'speech_voice': 'masked',
                'media_devices': {'mode': 'custom', 'video_inputs': 0,
                                  'audio_inputs': 1, 'audio_outputs': 1},
                'battery': 'masked',
                'hardware_concurrency': 4, 'device_memory': 8,
                'device_name': '', 'mac_address': '',
                'hardware_acceleration': True,
            },
            'advanced': {
                'language': 'based_on_ip', 'language_value': '',
                'timezone': 'based_on_ip', 'timezone_value': '',
                'geolocation_prompt': 'prompt', 'geolocation_source': 'based_on_ip',
                'webrtc': 'masked',
                'screen_resolution': 'custom',
                'screen_width': 1920, 'screen_height': 1080,
                'fonts': 'masked', 'do_not_track': False,
                'port_scan_protection': 'disabled',
                'disable_image_loading': False, 'save_tabs': True,
                'launch_args': '',
            },
        }
        return jsonify({'success': True, 'fingerprint': fp, 'nst_mode': True})

    # NST mode not active — return empty placeholder
    return jsonify({'success': False, 'message': 'NST Browser required for fingerprint generation'}), 400


@app.route('/api/profiles', methods=['POST'])
def profiles_create():
    """Create a new profile."""
    data = request.get_json(force=True, silent=True) or {}
    name = data.get('name', '').strip()
    email = data.get('email', '').strip()
    if not name:
        return jsonify({'success': False, 'message': 'Profile name is required'}), 400
    proxy = data.get('proxy')
    notes = data.get('notes', '')
    fingerprint_prefs = data.get('fingerprint_prefs', {})
    password = data.get('password', '')
    totp_secret = data.get('totp_secret', '')
    backup_codes = data.get('backup_codes', [])
    # Pass full overview/hardware/advanced from frontend (if user edited them)
    frontend_sections = {}
    if data.get('overview'):
        frontend_sections['overview'] = data['overview']
    if data.get('hardware'):
        frontend_sections['hardware'] = data['hardware']
    if data.get('advanced'):
        frontend_sections['advanced'] = data['advanced']
    engine = data.get('engine', 'nexus')
    if engine not in ('nst', 'nexus'):
        engine = 'nexus'
    group = (data.get('group', '') or 'default').strip()
    profile = profile_manager.create_profile(
        name, email, proxy=proxy, notes=notes,
        fingerprint_prefs=fingerprint_prefs,
        password=password, totp_secret=totp_secret, backup_codes=backup_codes,
        frontend_sections=frontend_sections,
        engine=engine,
    )
    # Set group (not accepted by create_profile directly)
    if group:
        profile_manager.update_profile(profile['id'], group=group)
        profile['group'] = group
    nst_err = profile.pop('_nst_create_error', None)
    if nst_err:
        return jsonify({'success': False, 'message': nst_err, 'profile': profile})
    return jsonify({'success': True, 'profile': profile})


@app.route('/api/profiles/<profile_id>', methods=['GET'])
def profiles_get(profile_id):
    """Get a single profile."""
    profile = profile_manager.get_profile(profile_id)
    if not profile:
        return jsonify({'success': False, 'message': 'Profile not found'}), 404
    return jsonify({'success': True, 'profile': profile})


@app.route('/api/profiles/<profile_id>', methods=['PUT'])
def profiles_update(profile_id):
    """Update a profile."""
    data = request.get_json(force=True, silent=True) or {}
    profile = profile_manager.update_profile(profile_id, **data)
    if not profile:
        return jsonify({'success': False, 'message': 'Profile not found'}), 404
    return jsonify({'success': True, 'profile': profile})


@app.route('/api/profiles/delete-all', methods=['DELETE'])
def profiles_delete_all():
    """Delete ALL profiles and their data directories."""
    try:
        profiles = profile_manager.list_profiles()
        count = len(profiles)
        for p in profiles:
            try:
                profile_manager.delete_profile(p['id'])
            except Exception:
                pass
        return jsonify({'success': True, 'deleted': count, 'message': f'Deleted {count} profiles'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/profiles/delete-by-engine/<engine>', methods=['DELETE'])
def profiles_delete_by_engine(engine):
    """Delete all profiles matching a specific engine (nst or nexus)."""
    if engine not in ('nst', 'nexus'):
        return jsonify({'success': False, 'message': 'Engine must be nst or nexus'}), 400
    try:
        count = profile_manager.delete_all_by_engine(engine)
        label = 'NST' if engine == 'nst' else 'Local'
        return jsonify({'success': True, 'deleted': count, 'message': f'Deleted {count} {label} profiles'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


_bulk_delete_progress = {
    'running': False, 'status': 'idle',
    'total': 0, 'deleted': 0, 'failed': 0, 'pending': 0,
    'current_profile': '',
}

@app.route('/api/profiles/delete-bulk', methods=['DELETE'])
def profiles_delete_bulk():
    """Delete multiple profiles by IDs with progress tracking."""
    global _bulk_delete_progress
    data = request.get_json() or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'success': False, 'message': 'No profile IDs provided'}), 400

    if _bulk_delete_progress.get('running'):
        return jsonify({'success': False, 'message': 'Bulk delete already running'}), 409

    # Set progress BEFORE spawning thread
    _bulk_delete_progress.update({
        'running': True, 'status': 'processing',
        'total': len(ids), 'deleted': 0, 'failed': 0, 'pending': len(ids),
        'current_profile': '',
    })

    def _worker():
        global _bulk_delete_progress
        deleted = 0
        failed = 0
        for pid in ids:
            # Get profile name for progress display
            try:
                p = profile_manager.get_profile(pid)
                name = p.get('name', p.get('email', pid)) if p else pid
            except Exception:
                name = pid
            _bulk_delete_progress['current_profile'] = name
            try:
                if profile_manager.delete_profile(pid):
                    deleted += 1
                else:
                    failed += 1
            except Exception:
                failed += 1
            _bulk_delete_progress.update({
                'deleted': deleted, 'failed': failed,
                'pending': max(0, len(ids) - deleted - failed),
            })
        _bulk_delete_progress.update({
            'running': False, 'status': 'completed',
            'deleted': deleted, 'failed': failed, 'pending': 0,
            'current_profile': '',
        })

    import threading
    threading.Thread(target=_worker, daemon=True, name='bulk-delete').start()
    return jsonify({'success': True, 'total': len(ids), 'message': 'Bulk delete started'})


@app.route('/api/profiles/delete-bulk-status', methods=['GET'])
def profiles_delete_bulk_status():
    """Return current bulk delete progress."""
    return jsonify({'success': True, 'progress': dict(_bulk_delete_progress)})


@app.route('/api/profiles/<profile_id>', methods=['DELETE'])
def profiles_delete(profile_id):
    """Delete a profile and its data directory."""
    ok = profile_manager.delete_profile(profile_id)
    if not ok:
        return jsonify({'success': False, 'message': 'Profile not found'}), 404
    return jsonify({'success': True, 'message': 'Profile deleted'})


@app.route('/api/profiles/<profile_id>/launch', methods=['POST'])
def profiles_launch(profile_id):
    """Launch a profile browser."""
    result = profile_manager.launch_profile(profile_id)
    if not result['success']:
        return jsonify(result), 400
    return jsonify(result)


@app.route('/api/profiles/<profile_id>/close', methods=['POST'])
def profiles_close(profile_id):
    """Close a profile browser."""
    ok = profile_manager.close_profile(profile_id)
    return jsonify({'success': ok})


@app.route('/api/profiles/close-all', methods=['POST'])
def profiles_close_all():
    """Close all open profile browsers."""
    profile_manager.close_all_profiles()
    return jsonify({'success': True})


@app.route('/api/profiles/cleanup', methods=['POST'])
def profiles_cleanup():
    """Delete orphan profile folders not in profiles.json."""
    result = profile_manager.cleanup_orphans()
    return jsonify({'success': True, **result})


@app.route('/api/profiles/<profile_id>/status', methods=['GET'])
def profiles_status(profile_id):
    """Get browser status for a profile."""
    return jsonify(profile_manager.profile_status(profile_id))


@app.route('/api/profiles/status', methods=['GET'])
def profiles_all_status():
    """Get counts of open/total profile browsers."""
    return jsonify(profile_manager.all_status())


@app.route('/api/profiles/batch-login-preview', methods=['POST'])
def profiles_batch_login_preview():
    """Read Excel and return count of valid accounts without running login."""
    data = request.get_json(force=True, silent=True) or {}
    file_path = data.get('file_path', '').strip()
    if not file_path or not os.path.isfile(file_path):
        return jsonify({'success': False, 'message': 'File not found'})
    try:
        df = pd.read_excel(file_path)
        total = len(df)
        valid = 0
        for _, row in df.iterrows():
            email = str(row.get('Email', '')).strip()
            password = str(row.get('Password', '')).strip()
            if email and password and email.lower() != 'nan' and password.lower() != 'nan':
                valid += 1
        cols = list(df.columns)
        return jsonify({'success': True, 'total': total, 'valid': valid, 'columns': cols})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/profiles/batch-login', methods=['POST'])
def profiles_batch_login():
    """Start batch login from Excel file."""
    data = request.get_json(force=True, silent=True) or {}
    file_path = data.get('file_path', '').strip()
    num_workers = int(data.get('workers', 3))
    engine = data.get('engine', 'nexus')
    os_type = data.get('os_type', 'random')
    group = data.get('group', 'default') or 'default'
    if not file_path:
        return jsonify({'success': False, 'message': 'File path is required'}), 400
    result = profile_manager.batch_login(file_path, num_workers, engine=engine, os_type=os_type, group=group)
    return jsonify(result)


@app.route('/api/profiles/config', methods=['GET'])
def profiles_config_get():
    """Get profile storage config."""
    return jsonify({'success': True, 'config': profile_manager.get_config()})


@app.route('/api/profiles/config', methods=['POST'])
def profiles_config_set():
    """Set profile storage path."""
    data = request.get_json(force=True, silent=True) or {}
    storage_path = data.get('storage_path', '')
    try:
        config = profile_manager.set_storage_path(storage_path)
        return jsonify({'success': True, 'config': config})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400


# ── Profile Manager: Do All Appeal ────────────────────────────────────────────

@app.route('/api/profiles/do-all-appeal', methods=['POST'])
def profiles_do_all_appeal():
    """Start Do All Appeal for selected profiles."""
    data = request.get_json(force=True, silent=True) or {}
    num_workers = int(data.get('num_workers', 5))
    profile_ids = data.get('profile_ids', [])
    result = profile_manager.do_all_appeal_profiles(num_workers=num_workers, profile_ids=profile_ids)
    if not result['success']:
        return jsonify(result), 400
    return jsonify(result)


@app.route('/api/profiles/appeal-status', methods=['GET'])
def profiles_appeal_status():
    """Get status of running Do All Appeal operation."""
    return jsonify(profile_manager.get_appeal_status())


@app.route('/api/profiles/appeal-match-excel', methods=['POST'])
def profiles_appeal_match_excel():
    """Read an Excel file, extract emails, and return matched profile IDs."""
    data = request.get_json(force=True, silent=True) or {}
    file_path = data.get('file_path', '')
    if not file_path or not os.path.isfile(file_path):
        return jsonify({'success': False, 'message': 'File not found'})
    try:
        df = pd.read_excel(file_path, engine='openpyxl')
        # Find the email column (case-insensitive)
        email_col = None
        for c in df.columns:
            if str(c).strip().lower() == 'email':
                email_col = c
                break
        if email_col is None:
            return jsonify({'success': False, 'message': 'No "Email" column found in the Excel file'})
        # Collect unique emails from Excel
        emails_in_excel = set()
        for _, row in df.iterrows():
            e = str(row.get(email_col, '')).strip().lower()
            if e and e != 'nan':
                emails_in_excel.add(e)
        if not emails_in_excel:
            return jsonify({'success': False, 'message': 'No emails found in the file'})
        # Match against existing profiles
        all_profiles = profile_manager.list_profiles()
        matched = []
        not_found = []
        for email in emails_in_excel:
            found = False
            for p in all_profiles:
                if (p.get('email') or '').strip().lower() == email:
                    matched.append({'id': p['id'], 'email': p.get('email', '')})
                    found = True
                    break
            if not found:
                not_found.append(email)
        return jsonify({
            'success': True,
            'total_emails': len(emails_in_excel),
            'matched': matched,
            'matched_count': len(matched),
            'not_found': not_found,
            'not_found_count': len(not_found),
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400


@app.route('/api/profiles/<profile_id>/relogin', methods=['POST'])
def profiles_relogin(profile_id):
    """Re-login a single profile using its saved credentials."""
    try:
        result = profile_manager.relogin_profile(profile_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/profiles/bulk-relogin', methods=['POST'])
def profiles_bulk_relogin():
    """Re-login multiple selected profiles with worker control."""
    try:
        data = request.get_json(force=True, silent=True) or {}
        ids = data.get('ids', [])
        num_workers = max(1, min(int(data.get('workers', 2)), 10))
        if not ids:
            return jsonify({'success': False, 'error': 'No profiles selected'})
        result = profile_manager.bulk_relogin_profiles(ids, num_workers)
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/profiles/bulk-relogin-status', methods=['GET'])
def profiles_bulk_relogin_status():
    try:
        return jsonify(profile_manager.get_bulk_relogin_status())
    except Exception:
        return jsonify({'running': False, 'status': 'idle'})


@app.route('/api/profiles/stop-appeal', methods=['POST'])
def profiles_stop_appeal():
    """Stop running appeal operation."""
    return jsonify(profile_manager.stop_appeal())


@app.route('/api/profiles/stop-health', methods=['POST'])
def profiles_stop_health():
    """Stop running health activity."""
    return jsonify(profile_manager.stop_health())


# ── Profile Manager: Run Operations (Step 1/2) ──────────────────────────────

@app.route('/api/profiles/run-operations', methods=['POST'])
def profiles_run_operations():
    """Run Step 1/2 operations on all logged-in profiles."""
    data = request.get_json(force=True, silent=True) or {}
    operations = data.get('operations', '')
    num_workers = int(data.get('num_workers', 5))
    params = {
        'new_password': data.get('new_password', ''),
        'recovery_email': data.get('recovery_email', ''),
        'recovery_phone': data.get('recovery_phone', ''),
        'name_country': data.get('name_country', 'US'),
        'first_name': data.get('first_name', ''),
        'last_name': data.get('last_name', ''),
    }
    result = profile_manager.run_operations_on_profiles(
        operations=operations, num_workers=num_workers, params=params
    )
    if not result['success']:
        return jsonify(result), 400
    return jsonify(result)


@app.route('/api/profiles/ops-status', methods=['GET'])
def profiles_ops_status():
    """Get status of running operations."""
    return jsonify(profile_manager.get_ops_status())


@app.route('/api/profiles/run-ops', methods=['POST'])
def profiles_run_ops():
    """Run operations on selected profiles."""
    data = request.get_json(force=True, silent=True) or {}
    profile_ids = data.get('profile_ids', [])
    operations = data.get('operations', '')
    params = data.get('params', {})
    num_workers = max(1, min(int(data.get('num_workers', 5)), 20))

    if not profile_ids:
        return jsonify({'success': False, 'error': 'No profiles selected'})
    if not operations:
        return jsonify({'success': False, 'error': 'No operations selected'})

    try:
        result = profile_manager.run_operations_on_profiles(
            operations=operations,
            num_workers=num_workers,
            params=params,
            profile_ids=profile_ids,
        )
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/profiles/run-health', methods=['POST'])
def profiles_run_health():
    """Start health activity on selected profiles with specific activities."""
    data = request.get_json(force=True, silent=True) or {}
    num_workers = int(data.get('num_workers', 3))
    activities = data.get('activities', [])
    profile_ids = data.get('profile_ids', [])
    country = data.get('country', 'US')
    rounds = int(data.get('rounds', 1))
    duration_minutes = int(data.get('duration_minutes', 0))
    gmb_name = data.get('gmb_name', '')
    gmb_address = data.get('gmb_address', '')
    result = profile_manager.run_health_activity(
        num_workers=num_workers,
        activities=activities,
        profile_ids=profile_ids,
        country=country,
        rounds=rounds,
        duration_minutes=duration_minutes,
        gmb_name=gmb_name,
        gmb_address=gmb_address,
    )
    return jsonify(result)


@app.route('/api/profiles/health-status', methods=['GET'])
def profiles_health_status():
    """Get status of running health activity."""
    return jsonify(profile_manager.get_health_status())


@app.route('/api/profiles/do-write-review', methods=['POST'])
def profiles_do_write_review():
    """Start Write Review operation from Excel file for matched profiles."""
    data = request.get_json(force=True, silent=True) or {}
    excel_file = data.get('excel_file', '').strip()
    num_workers = int(data.get('num_workers', 3))
    profile_ids = data.get('profile_ids') or None
    if not excel_file:
        return jsonify({'success': False, 'message': 'excel_file path is required'}), 400
    if not os.path.isfile(excel_file):
        return jsonify({'success': False, 'message': f'File not found: {excel_file}'}), 400
    result = profile_manager.do_write_review_profiles(
        excel_file=excel_file,
        num_workers=num_workers,
        profile_ids=profile_ids,
    )
    if not result.get('success'):
        return jsonify(result), 400
    return jsonify(result)


@app.route('/api/profiles/review-status', methods=['GET'])
def profiles_review_status():
    """Get status of running Write Review operation."""
    return jsonify(profile_manager.get_review_status())


@app.route('/api/profiles/write-review-template', methods=['GET'])
def profiles_write_review_template():
    """Return a pre-filled Excel template showing how to prepare the Write Review sheet."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Sheet 1 — Reviews (the actual data sheet)
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    ws = wb.active
    ws.title = 'Reviews'

    thin = Side(style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header style
    hdr_font  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    hdr_fill  = PatternFill('solid', fgColor='1E3A5F')
    hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Column definitions: (header, width, note)
    columns = [
        ('Email',        30),
        ('Review URL',   55),
        ('GMB URL',      45),
        ('Review Text',  55),
        ('Review Stars', 14),
    ]
    for col_idx, (h, w) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 24

    # Example rows  (Review URL, GMB URL, Review Text, Stars)
    examples = [
        ('john123@gmail.com',  'https://www.google.com/maps/place//data=!4m3!3m2!1s0x1234:0x5678!12e1', 'https://maps.google.com/maps?cid=123456789012345', 'Amazing place! Highly recommend to everyone. Very professional and friendly staff.', 5),
        ('mary456@gmail.com',  'https://www.google.com/maps/place//data=!4m3!3m2!1s0xAAAA:0xBBBB!12e1', 'https://maps.google.com/maps?cid=987654321098765', 'Great experience overall. Clean, organised and the team is very helpful.', 4),
        ('alex789@gmail.com',  'https://www.google.com/maps/place//data=!4m3!3m2!1s0xCCCC:0xDDDD!12e1', 'https://maps.google.com/maps?cid=111222333444555', 'Excellent service and quality. Will definitely visit again. Truly outstanding!', 5),
        ('sara001@gmail.com',  'https://www.google.com/maps/place//data=!4m3!3m2!1s0xEEEE:0xFFFF!12e1', 'https://maps.google.com/maps?cid=666777888999000', '',                                                                                  5),
        ('test002@gmail.com',  'https://www.google.com/maps/place//data=!4m3!3m2!1s0x1111:0x2222!12e1', 'https://maps.google.com/maps?cid=222333444555666', 'Very good. Satisfied with the service provided by the team here.',                 4),
    ]

    row_fills = ['EFF6FF', 'F0FDF4', 'EFF6FF', 'FEF9C3', 'F0FDF4']
    cell_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for row_idx, (email, gmb, text, stars) in enumerate(examples, 2):
        fill = PatternFill('solid', fgColor=row_fills[row_idx - 2])
        for col_idx, value in enumerate([email, gmb, text, stars], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = border
            cell.fill = fill
        ws.row_dimensions[row_idx].height = 32

    ws.freeze_panes = 'A2'

    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    # Sheet 2 — Instructions
    # ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    wi = wb.create_sheet('Instructions')
    wi.sheet_view.showGridLines = False
    wi.column_dimensions['A'].width = 2
    wi.column_dimensions['B'].width = 22
    wi.column_dimensions['C'].width = 60

    title_font  = Font(name='Calibri', bold=True, size=16, color='1E3A5F')
    h2_font     = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
    h2_fill     = PatternFill('solid', fgColor='1E3A5F')
    label_font  = Font(name='Calibri', bold=True, size=11, color='1E40AF')
    body_font   = Font(name='Calibri', size=11, color='374151')
    note_font   = Font(name='Calibri', italic=True, size=10, color='6B7280')
    ok_font     = Font(name='Calibri', bold=True, size=11, color='065F46')
    warn_font   = Font(name='Calibri', bold=True, size=11, color='991B1B')
    green_fill  = PatternFill('solid', fgColor='D1FAE5')
    red_fill    = PatternFill('solid', fgColor='FEE2E2')

    def _irow(row, col_b='', col_c='', bfont=None, cfont=None, bfill=None, cfill=None, height=20, merge_bc=False):
        wi.row_dimensions[row].height = height
        if col_b:
            cb = wi.cell(row=row, column=2, value=col_b)
            if bfont: cb.font = bfont
            if bfill: cb.fill = bfill
            cb.alignment = Alignment(vertical='center', horizontal='left', indent=1)
        if col_c:
            cc = wi.cell(row=row, column=3, value=col_c)
            if cfont: cc.font = cfont
            if cfill: cc.fill = cfill
            cc.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)
        if merge_bc:
            wi.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)

    r = 1
    _irow(r, 'WRITE REVIEW — Excel Sheet Guide', '', bfont=title_font, height=36, merge_bc=True); r += 1
    _irow(r, height=10); r += 1

    # Required columns section
    _irow(r, '  REQUIRED COLUMNS', '', bfont=h2_font, bfill=h2_fill, height=24, merge_bc=True); r += 1
    _irow(r, 'Email', 'The Gmail account email address. Must match exactly what is saved in the profile.', bfont=label_font, cfont=body_font, height=28); r += 1
    _irow(r, 'Review URL', 'Direct review link that opens the review popup instantly. Use "GMB → Review URL" tool to generate these. RECOMMENDED — much faster than GMB URL.', bfont=label_font, cfont=body_font, height=40); r += 1
    _irow(r, height=8); r += 1

    # Optional columns section
    _irow(r, '  OPTIONAL COLUMNS', '', bfont=h2_font, bfill=h2_fill, height=24, merge_bc=True); r += 1
    _irow(r, 'GMB URL', 'Fallback: Google Maps business page URL. Only used if Review URL is empty for that row. Slower — requires clicking "Write a review" button.', bfont=label_font, cfont=body_font, height=40); r += 1
    _irow(r, 'Review Text', 'The review text to post. Leave blank to post stars only (no text review).', bfont=label_font, cfont=body_font, height=28); r += 1
    _irow(r, 'Review Stars', 'Number from 1 to 5. If left blank or missing, defaults to 5 stars.', bfont=label_font, cfont=body_font, height=28); r += 1
    _irow(r, height=8); r += 1

    # Rules section
    _irow(r, '  RULES', '', bfont=h2_font, bfill=h2_fill, height=24, merge_bc=True); r += 1
    rules = [
        '1.  Column headers must be EXACTLY: Email, Review URL, GMB URL, Review Text, Review Stars',
        '2.  Spelling and capitalisation matters — "email" or "review url" will NOT work',
        '3.  One account per row — do not merge or duplicate rows',
        '4.  Rows with empty Email or no URL (both Review URL and GMB URL empty) are skipped',
        '5.  Stars must be a plain number: 1, 2, 3, 4 or 5 — not "5 stars" or "five"',
        '6.  The system matches each row to a profile by email — no manual selection needed',
        '7.  If an email is in the sheet but not saved as a profile, that row is skipped',
        '8.  Review URL is preferred — if both Review URL and GMB URL are filled, Review URL is used',
    ]
    for rule in rules:
        _irow(r, '', rule, cfont=body_font, height=22); r += 1
    _irow(r, height=8); r += 1

    # GMB URL section
    _irow(r, '  HOW TO GET GMB URL', '', bfont=h2_font, bfill=h2_fill, height=24, merge_bc=True); r += 1
    gmb_steps = [
        'Step 1 →  Open Google Maps (maps.google.com)',
        'Step 2 →  Search for the business name',
        'Step 3 →  Click on the business listing',
        'Step 4 →  Copy the full URL from the browser address bar',
        'Step 5 →  Paste it into the GMB URL column',
        '',
        'Example URL:  https://www.google.com/maps/place/Business+Name/@lat,lng,zoom/...',
        'Also works:   https://maps.google.com/maps?cid=123456789012345',
    ]
    for step in gmb_steps:
        _irow(r, '', step, cfont=body_font, height=22); r += 1
    _irow(r, height=8); r += 1

    # Do / Don't
    _irow(r, '  ✅  CORRECT EXAMPLES', '', bfont=Font(name='Calibri', bold=True, size=12, color='065F46'), bfill=green_fill, height=24, merge_bc=True); r += 1
    goods = [
        'john@gmail.com  |  Review URL: https://google.com/maps/place//data=...  |  Great service!  |  5  (fastest)',
        'mary@gmail.com  |  Review URL: (blank)  |  GMB URL: https://maps.google.com/maps?cid=123  |  4  (fallback)',
    ]
    for g in goods:
        _irow(r, '', g, cfont=ok_font, height=22); r += 1
    _irow(r, height=8); r += 1

    _irow(r, '  ❌  WRONG EXAMPLES (will be skipped or cause errors)', '', bfont=Font(name='Calibri', bold=True, size=12, color='991B1B'), bfill=red_fill, height=24, merge_bc=True); r += 1
    bads = [
        'john123  |  (missing @gmail.com — won\'t match any profile)',
        '(blank)  |  https://google.com/maps/place/...  |  Review  |  5  — row skipped: no email',
        'mary@gmail.com  |  (both Review URL and GMB URL blank)  — row skipped: no URL',
        'john@gmail.com  |  Review URL  |  text  |  5 stars  — stars must be a number',
    ]
    for b in bads:
        _irow(r, '', b, cfont=warn_font, height=22); r += 1

    # ── Save ─────────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from flask import send_file
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='WriteReview_Template.xlsx'
    )


@app.route('/api/profiles/write-review-preview', methods=['POST'])
def profiles_write_review_preview():
    """Preview Excel file for Write Review — returns matched profile count."""
    data = request.get_json(force=True, silent=True) or {}
    excel_file = data.get('excel_file', '').strip()
    if not excel_file or not os.path.isfile(excel_file):
        return jsonify({'success': False, 'message': 'File not found'})
    try:
        df = pd.read_excel(excel_file)
        cols = list(df.columns)
        has_review_url = 'Review URL' in cols
        has_gmb_url = 'GMB URL' in cols
        has_required = 'Email' in cols and (has_review_url or has_gmb_url)
        valid = 0
        if has_required:
            for _, row in df.iterrows():
                email = str(row.get('Email', '')).strip()
                review_url = str(row.get('Review URL', '')).strip() if has_review_url else ''
                gmb = str(row.get('GMB URL', '')).strip() if has_gmb_url else ''
                if review_url.lower() == 'nan': review_url = ''
                if gmb.lower() == 'nan': gmb = ''
                if email and email.lower() != 'nan' and (review_url or gmb):
                    valid += 1
        # Count how many profiles match the emails
        emails_in_excel = set()
        if has_required:
            for _, row in df.iterrows():
                e = str(row.get('Email', '')).strip().lower()
                if e and e != 'nan': emails_in_excel.add(e)
        all_profiles = profile_manager.list_profiles()
        matched = sum(1 for p in all_profiles if (p.get('email') or '').strip().lower() in emails_in_excel)
        return jsonify({
            'success': True, 'total_rows': len(df), 'valid_rows': valid,
            'matched_profiles': matched, 'columns': cols,
            'has_review_text': 'Review Text' in cols,
            'has_stars': 'Review Stars' in cols,
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


@app.route('/api/gmb-to-review/preview', methods=['POST'])
def gmb_to_review_preview():
    """Preview Excel file for GMB → Review URL — returns row count and GMB URL count."""
    data = request.get_json(force=True, silent=True) or {}
    file_path = data.get('file_path', '').strip()
    if not file_path or not os.path.isfile(file_path):
        return jsonify({'success': False, 'message': 'File not found'})
    try:
        df = pd.read_excel(file_path)
        cols = list(df.columns)
        if 'GMB URL' not in cols:
            return jsonify({'success': False, 'message': 'Missing required column: "GMB URL"'})
        gmb_count = int(df['GMB URL'].dropna().astype(str).str.strip().loc[lambda s: (s != '') & (s.str.lower() != 'nan')].shape[0])
        return jsonify({
            'success': True,
            'total_rows': len(df),
            'gmb_url_count': gmb_count,
            'columns': cols,
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})


_gmb_review_progress = {
    'running': False, 'total': 0, 'done': 0, 'success': 0, 'failed': 0,
    'current_url': '', 'results': [], 'report_path': None,
}

def _gmb_review_worker(file_path):
    """Background thread: resolve GMB URLs and build Review URLs."""
    import re
    import requests as req_lib
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    global _gmb_review_progress
    try:
        df = pd.read_excel(file_path)
        rows = []
        for _, row in df.iterrows():
            gmb_url = str(row.get('GMB URL', '')).strip()
            if gmb_url and gmb_url.lower() != 'nan':
                rows.append((row, gmb_url))
            else:
                rows.append((row, ''))

        _gmb_review_progress['total'] = len([r for r in rows if r[1]])
        _gmb_review_progress['done'] = 0
        _gmb_review_progress['success'] = 0
        _gmb_review_progress['failed'] = 0
        _gmb_review_progress['results'] = []

        review_urls = []
        for row_data, gmb_url in rows:
            if not gmb_url:
                review_urls.append('')
                continue
            _gmb_review_progress['current_url'] = gmb_url
            try:
                r = req_lib.head(gmb_url, allow_redirects=True, timeout=15)
                match = re.search(r'(?:!1s|ftid=)(0x[0-9a-fA-F]+:0x[0-9a-fA-F]+)', r.url)
                if match:
                    hex_cid = match.group(1)
                    review_url = f"https://www.google.com/maps/place//data=!4m3!3m2!1s{hex_cid}!12e1"
                    review_urls.append(review_url)
                    _gmb_review_progress['success'] += 1
                    _gmb_review_progress['results'].append({'url': gmb_url, 'status': 'success', 'review_url': review_url})
                else:
                    review_urls.append('ERROR: CID not found')
                    _gmb_review_progress['failed'] += 1
                    _gmb_review_progress['results'].append({'url': gmb_url, 'status': 'failed', 'error': 'CID not found'})
            except Exception as e:
                review_urls.append(f'ERROR: {e}')
                _gmb_review_progress['failed'] += 1
                _gmb_review_progress['results'].append({'url': gmb_url, 'status': 'failed', 'error': str(e)})
            _gmb_review_progress['done'] += 1

        df['Review URL'] = review_urls

        # Build styled Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Review URLs'

        thin = Side(style='thin', color='CBD5E1')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        hdr_fill = PatternFill('solid', fgColor='1E3A5F')
        hdr_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        headers = list(df.columns)
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = border

        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            for c_idx, col in enumerate(headers, 1):
                val = row[col]
                if pd.isna(val):
                    val = ''
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
                if col == 'Review URL':
                    if str(val).startswith('ERROR'):
                        cell.font = Font(color='DC2626')
                    elif val:
                        cell.font = Font(color='059669')

        for c_idx, col in enumerate(headers, 1):
            max_len = len(str(col))
            for row in ws.iter_rows(min_row=2, min_col=c_idx, max_col=c_idx):
                for cell in row:
                    if cell.value:
                        max_len = max(max_len, min(len(str(cell.value)), 60))
            ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = max_len + 4

        ws.freeze_panes = 'A2'

        # Save to output directory so it appears in Reports tab
        output_dir = str((Path(__file__).parent.parent / 'output').resolve())
        os.makedirs(output_dir, exist_ok=True)
        src_name = os.path.splitext(os.path.basename(file_path))[0]
        from datetime import datetime
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'{src_name}_ReviewURLs_{ts}.xlsx'
        out_path = os.path.join(output_dir, filename)
        wb.save(out_path)
        _gmb_review_progress['report_path'] = out_path

    except Exception as e:
        _gmb_review_progress['results'].append({'url': '', 'status': 'failed', 'error': str(e)})
    finally:
        _gmb_review_progress['running'] = False
        _gmb_review_progress['current_url'] = ''


@app.route('/api/gmb-to-review/process', methods=['POST'])
def gmb_to_review_process():
    """Start background thread to resolve GMB URLs and build Review URLs."""
    import threading

    global _gmb_review_progress
    if _gmb_review_progress.get('running'):
        return jsonify({'success': False, 'message': 'Already processing'}), 400

    data = request.get_json(force=True, silent=True) or {}
    file_path = data.get('file_path', '').strip()
    if not file_path or not os.path.isfile(file_path):
        return jsonify({'success': False, 'message': 'File not found'}), 400

    try:
        df = pd.read_excel(file_path)
        if 'GMB URL' not in df.columns:
            return jsonify({'success': False, 'message': 'Missing required column: "GMB URL"'}), 400
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 400

    _gmb_review_progress = {
        'running': True, 'total': 0, 'done': 0, 'success': 0, 'failed': 0,
        'current_url': '', 'results': [], 'report_path': None,
    }

    t = threading.Thread(target=_gmb_review_worker, args=(file_path,), daemon=True)
    t.start()

    return jsonify({'success': True, 'message': 'Processing started'})


@app.route('/api/gmb-to-review/status', methods=['GET'])
def gmb_to_review_status():
    """Return current progress of GMB → Review URL processing."""
    return jsonify({
        'running': _gmb_review_progress.get('running', False),
        'total': _gmb_review_progress.get('total', 0),
        'done': _gmb_review_progress.get('done', 0),
        'success': _gmb_review_progress.get('success', 0),
        'failed': _gmb_review_progress.get('failed', 0),
        'current_url': _gmb_review_progress.get('current_url', ''),
        'report_path': _gmb_review_progress.get('report_path'),
    })


@app.route('/api/gmb-to-review/download', methods=['GET'])
def gmb_to_review_download():
    """Download the generated Review URLs Excel file."""
    report_path = _gmb_review_progress.get('report_path')
    if not report_path or not os.path.isfile(report_path):
        return jsonify({'success': False, 'message': 'No report file available'}), 404
    return send_file(
        report_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=os.path.basename(report_path)
    )


@app.route('/api/nst/status', methods=['GET'])
def nst_status():
    """Check NST Browser API connectivity."""
    try:
        from shared.nexus_profile_manager import _nst_check, _nst_api_base
        connected = _nst_check()
        return jsonify({
            'success': True,
            'connected': connected,
            'api_base': _nst_api_base,
        })
    except Exception as e:
        return jsonify({'success': False, 'connected': False, 'message': str(e)})


@app.route('/api/nst/config', methods=['GET'])
def nst_config_get():
    """Get NST Browser config (browser.json)."""
    try:
        import json as _json
        bj = RESOURCES_PATH / 'config' / 'browser.json'
        if bj.exists():
            cfg = _json.loads(bj.read_text('utf-8'))
            # Mask key for display — only show last 8 chars
            key = cfg.get('nst_api_key', '')
            masked = ('*' * max(0, len(key) - 8) + key[-8:]) if len(key) > 8 else key
            return jsonify({
                'success': True,
                'nst_api_key': key,
                'nst_api_key_masked': masked,
                'nst_api_base': cfg.get('nst_api_base', 'http://localhost:8848/api/v2'),
            })
        return jsonify({'success': True, 'nst_api_key': '', 'nst_api_base': 'http://localhost:8848/api/v2'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/nst/config', methods=['POST'])
def nst_config_save():
    """Save NST Browser config (API key + base URL) to browser.json and reload."""
    try:
        import json as _json
        data = request.get_json(force=True)
        bj = RESOURCES_PATH / 'config' / 'browser.json'

        # Load existing config
        cfg = {}
        if bj.exists():
            cfg = _json.loads(bj.read_text('utf-8'))

        # Update NST fields
        new_key = data.get('nst_api_key', '').strip()
        new_base = data.get('nst_api_base', '').strip()
        if new_key:
            cfg['nst_api_key'] = new_key
        if new_base:
            cfg['nst_api_base'] = new_base

        # Ensure NST mode is on
        cfg['use_nst'] = True

        # Save
        bj.write_text(_json.dumps(cfg, indent=4), 'utf-8')

        # Reload in profile manager
        try:
            import shared.nexus_profile_manager as npm
            npm._nst_api_key = cfg.get('nst_api_key', '')
            npm._nst_api_base = cfg.get('nst_api_base', 'http://localhost:8848/api/v2')
            connected = npm._nst_check()
        except Exception:
            connected = False

        return jsonify({'success': True, 'connected': connected, 'message': 'NST config saved'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/nexus-api-key', methods=['GET'])
def nexus_api_key_get():
    """Get the Nexus API key from browser.json."""
    try:
        bj = RESOURCES_PATH / 'config' / 'browser.json'
        if bj.exists():
            cfg = json.loads(bj.read_text('utf-8'))
            return jsonify({'success': True, 'nexus_api_key': cfg.get('nexus_api_key', '')})
        return jsonify({'success': True, 'nexus_api_key': ''})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/nexus-api-key', methods=['POST'])
def nexus_api_key_save():
    """Save the Nexus API key to browser.json."""
    try:
        data = request.get_json(force=True, silent=True) or {}
        bj = RESOURCES_PATH / 'config' / 'browser.json'
        cfg = {}
        if bj.exists():
            cfg = json.loads(bj.read_text('utf-8'))
        cfg['nexus_api_key'] = data.get('nexus_api_key', '').strip()
        bj.write_text(json.dumps(cfg, indent=4), 'utf-8')
        return jsonify({'success': True, 'message': 'Nexus API key saved'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/name-countries', methods=['GET'])
def get_name_countries():
    """Get available countries for random name generation."""
    from shared.random_names import get_available_countries
    return jsonify(get_available_countries())


# ── VPN Control Setup ─────────────────────────────────────────────────────────

@app.route('/api/vpn/status', methods=['GET'])
def vpn_status():
    """Check if VPN kill task is set up and current IP."""
    try:
        from shared.vpn_controller import is_vpn_task_setup, get_public_ip
        task_ready = is_vpn_task_setup()
        current_ip = get_public_ip()
        return jsonify({
            'success': True,
            'task_setup': task_ready,
            'current_ip': current_ip or 'unknown',
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/vpn/setup', methods=['POST'])
def vpn_setup():
    """Create the admin scheduled task for VPN kill (one-time, shows UAC)."""
    try:
        from shared.vpn_controller import setup_vpn_task
        result = setup_vpn_task()
        return jsonify(result)
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/vpn/test-kill', methods=['POST'])
def vpn_test_kill():
    """Test VPN kill (disconnect) using the scheduled task."""
    try:
        from shared.vpn_controller import disconnect_vpn
        result = disconnect_vpn()
        return jsonify({'success': True, **result})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/vpn/test-connect', methods=['POST'])
def vpn_test_connect():
    """Test VPN connect cycle."""
    try:
        vpn_path = (request.get_json(force=True, silent=True) or {}).get(
            'vpn_path', r'C:\Program Files\Privax\HMA VPN\Vpn.exe'
        )
        from shared.vpn_controller import connect_vpn
        result = connect_vpn(vpn_path=vpn_path, max_retries=3)
        return jsonify({'success': result['connected'], **result})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# NEXUS API v2 — External REST API (NST-compatible response format)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

nexus_api = Blueprint('nexus_api', __name__, url_prefix='/api/nexus/v2')


def _napi(data=None, msg='success', err=False, status=200):
    """NST-compatible JSON response wrapper."""
    return jsonify({'err': err, 'msg': msg, 'data': data}), status


def _load_nexus_api_key():
    """Load nexus_api_key from browser.json."""
    try:
        bj = RESOURCES_PATH / 'config' / 'browser.json'
        if bj.exists():
            cfg = json.loads(bj.read_text('utf-8'))
            return cfg.get('nexus_api_key', '')
    except Exception:
        pass
    return ''


@nexus_api.before_request
def _nexus_api_check_key():
    """Validate x-api-key header if nexus_api_key is configured."""
    expected = _load_nexus_api_key()
    if not expected:
        return None  # No key set — allow all requests
    provided = request.headers.get('x-api-key', '')
    if provided != expected:
        return _napi(msg='Invalid or missing API key', err=True, status=401)


# ── Profile CRUD ─────────────────────────────────────────────────────────────

@nexus_api.route('/profiles/groups', methods=['GET'])
def napi_list_groups():
    """Return all unique profile group names."""
    profiles = profile_manager.list_profiles()
    all_groups = set()
    for p in profiles:
        for g in profile_manager._get_groups(p):
            all_groups.add(g)
    return _napi({'groups': sorted(all_groups)})


@nexus_api.route('/profiles', methods=['GET'])
def napi_list_profiles():
    """List profiles with optional filtering and pagination."""
    search = request.args.get('search', '').lower()
    filt = request.args.get('filter', 'all').lower()
    group_filter = request.args.get('group', '').lower()
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 20))
    per_page = max(1, min(per_page, 10000))

    profiles = profile_manager.list_profiles()
    if search:
        profiles = [p for p in profiles if search in p.get('name', '').lower()
                    or search in p.get('email', '').lower()]
    if group_filter:
        profiles = [p for p in profiles if group_filter in [g.lower() for g in profile_manager._get_groups(p)]]
    if filt == 'running':
        profiles = [p for p in profiles if p.get('browser_open') == 'running']
    elif filt == 'logged_in':
        profiles = [p for p in profiles if p.get('status') == 'logged_in']
    elif filt == 'not_logged_in':
        profiles = [p for p in profiles if p.get('status') not in ('logged_in', 'login_failed')]
    elif filt == 'login_failed':
        profiles = [p for p in profiles if p.get('status') == 'login_failed']

    total = len(profiles)
    total_pages = max(1, (total + per_page - 1) // per_page)
    page = max(1, min(page, total_pages))
    start = (page - 1) * per_page

    return _napi({
        'profiles': profiles[start:start + per_page],
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': total_pages,
    })


@nexus_api.route('/profiles/<profile_id>', methods=['GET'])
def napi_get_profile(profile_id):
    """Get a single profile."""
    p = profile_manager.get_profile(profile_id)
    if not p:
        return _napi(msg='Profile not found', err=True, status=404)
    return _napi(p)


@nexus_api.route('/profiles', methods=['POST'])
def napi_create_profile():
    """Create a new profile."""
    data = request.get_json(force=True, silent=True) or {}
    name = data.get('name', f'Profile {secrets.token_hex(3)}')
    engine = data.get('engine', 'nexus')
    os_type = data.get('os_type', 'windows')

    # Parse proxy from string if given as string
    proxy = data.get('proxy')
    if isinstance(proxy, str) and proxy:
        from shared.nexus_proxy_manager import parse_proxy
        proxy = parse_proxy(proxy)

    try:
        profile = profile_manager.create_profile(
            name=name,
            email=data.get('email', ''),
            proxy=proxy,
            notes=data.get('notes', ''),
            fingerprint_prefs={'os_type': os_type},
            password=data.get('password', ''),
            totp_secret=data.get('totp_secret', ''),
            backup_codes=data.get('backup_codes', []),
            engine=engine,
            frontend_sections={
                'overview': data.get('overview', {}),
                'advanced': data.get('advanced', {}),
            },
        )
        nst_err = profile.pop('_nst_create_error', None)
        if nst_err:
            return _napi(profile, msg=nst_err, err=True, status=400)
        return _napi(profile, msg='Profile created', status=201)
    except Exception as e:
        return _napi(msg=str(e), err=True, status=500)


@nexus_api.route('/profiles/<profile_id>', methods=['PUT'])
def napi_update_profile(profile_id):
    """Update profile fields."""
    data = request.get_json(force=True, silent=True) or {}
    # Parse proxy from string if given as string
    if 'proxy' in data and isinstance(data['proxy'], str) and data['proxy']:
        from shared.nexus_proxy_manager import parse_proxy
        data['proxy'] = parse_proxy(data['proxy'])
    try:
        result = profile_manager.update_profile(profile_id, **data)
        if result:
            return _napi(result, msg='Profile updated')
        return _napi(msg='Profile not found', err=True, status=404)
    except Exception as e:
        return _napi(msg=str(e), err=True, status=500)


@nexus_api.route('/profiles/<profile_id>', methods=['DELETE'])
def napi_delete_profile(profile_id):
    """Delete a single profile."""
    ok = profile_manager.delete_profile(profile_id)
    if ok:
        return _napi(msg='Profile deleted')
    return _napi(msg='Profile not found', err=True, status=404)


@nexus_api.route('/profiles', methods=['DELETE'])
def napi_delete_all_profiles():
    """Delete ALL profiles."""
    profile_manager.delete_all_profiles()
    return _napi(msg='All profiles deleted')


# ── Browser Control ──────────────────────────────────────────────────────────

@nexus_api.route('/browsers/<profile_id>', methods=['POST'])
def napi_launch_browser(profile_id):
    """Launch browser and return CDP WebSocket URL."""
    try:
        ws = profile_manager.launch_and_connect(profile_id)
        return _napi({'webSocketDebuggerUrl': ws}, msg='Browser launched')
    except Exception as e:
        return _napi(msg=str(e), err=True, status=500)


@nexus_api.route('/browsers/<profile_id>', methods=['DELETE'])
def napi_stop_browser(profile_id):
    """Stop browser for a profile."""
    profile_manager.stop_nst_browser(profile_id)
    return _napi(msg='Browser stopped')


@nexus_api.route('/browsers/<profile_id>/launch', methods=['POST'])
def napi_launch_profile(profile_id):
    """Launch browser for UI viewing (Play button equivalent)."""
    result = profile_manager.launch_profile(profile_id)
    if result.get('success'):
        return _napi(result, msg='Browser launched')
    return _napi(msg=result.get('error', 'Launch failed'), err=True, status=500)


@nexus_api.route('/browsers/<profile_id>/close', methods=['POST'])
def napi_close_profile(profile_id):
    """Close browser for a profile."""
    ok = profile_manager.close_profile(profile_id)
    if ok:
        return _napi(msg='Browser closed')
    return _napi(msg='Browser not running', err=True, status=404)


@nexus_api.route('/browsers/close-all', methods=['POST'])
def napi_close_all():
    """Close all running browsers."""
    profile_manager.close_all_profiles()
    return _napi(msg='All browsers closed')


# ── Status ───────────────────────────────────────────────────────────────────

@nexus_api.route('/browsers/<profile_id>/status', methods=['GET'])
def napi_profile_status(profile_id):
    """Get browser status for a single profile."""
    return _napi(profile_manager.profile_status(profile_id))


@nexus_api.route('/browsers/status', methods=['GET'])
def napi_all_status():
    """Get status of all profiles."""
    profiles = profile_manager.list_profiles()
    statuses = {}
    for p in profiles:
        statuses[p['id']] = {
            'browser_open': p.get('browser_open', 'stopped'),
            'engine': p.get('engine', 'nst'),
        }
    running = sum(1 for s in statuses.values() if s['browser_open'] == 'running')
    return _napi({
        'profiles': statuses,
        'running_count': running,
        'total_count': len(statuses),
    })


# ── Batch Operations ─────────────────────────────────────────────────────────

@nexus_api.route('/profiles/batch-create', methods=['POST'])
def napi_batch_create():
    """Create multiple profiles at once."""
    data = request.get_json(force=True, silent=True) or {}
    count = int(data.get('count', 1))
    engine = data.get('engine', 'nexus')
    os_type = data.get('os_type', 'windows')
    proxy_list = data.get('proxy_list', [])

    created = []
    for i in range(count):
        proxy = None
        if proxy_list:
            raw = proxy_list[i % len(proxy_list)]
            if isinstance(raw, str):
                from shared.nexus_proxy_manager import parse_proxy
                proxy = parse_proxy(raw)
            else:
                proxy = raw
        try:
            p = profile_manager.create_profile(
                name=f'Profile {i + 1}',
                fingerprint_prefs={'os_type': os_type},
                proxy=proxy,
                engine=engine,
            )
            created.append(p)
        except Exception:
            pass

    return _napi({'created': len(created), 'profiles': created}, msg=f'{len(created)} profiles created')


@nexus_api.route('/profiles/batch-login', methods=['POST'])
def napi_batch_login():
    """Batch login from Excel file."""
    data = request.get_json(force=True, silent=True) or {}
    file_path = data.get('file_path', '')
    workers = int(data.get('num_workers', 3))
    engine = data.get('engine', 'nexus')
    os_type = data.get('os_type', 'random')
    group = data.get('group', 'default') or 'default'
    if not file_path:
        return _napi(msg='file_path required', err=True, status=400)
    try:
        result = profile_manager.batch_login(file_path, workers, engine=engine, os_type=os_type, group=group)
        return _napi(result, msg='Batch login started')
    except Exception as e:
        return _napi(msg=str(e), err=True, status=500)


@nexus_api.route('/profiles/run-operations', methods=['POST'])
def napi_run_operations():
    """Run bot operations on profiles."""
    data = request.get_json(force=True, silent=True) or {}
    operations = data.get('operations', '')
    workers = int(data.get('num_workers', 5))
    params = data.get('params', {})
    try:
        result = profile_manager.run_operations_on_profiles(operations, workers, params)
        return _napi(result, msg='Operations started')
    except Exception as e:
        return _napi(msg=str(e), err=True, status=500)


@nexus_api.route('/profiles/ops-status', methods=['GET'])
def napi_ops_status():
    """Get current operations progress."""
    return _napi(profile_manager.get_ops_status())


# ── Config ───────────────────────────────────────────────────────────────────

@nexus_api.route('/config', methods=['GET'])
def napi_config_get():
    """Get profile manager config."""
    cfg = profile_manager.get_config()
    return _napi(cfg)


@nexus_api.route('/config', methods=['POST'])
def napi_config_set():
    """Update profile manager config."""
    data = request.get_json(force=True, silent=True) or {}
    if 'storage_path' in data:
        result = profile_manager.set_storage_path(data['storage_path'])
        return _napi(result, msg='Config updated')
    return _napi(msg='No config fields to update', err=True, status=400)


@nexus_api.route('/profiles/export', methods=['POST'])
def napi_export():
    """Export profiles to JSON."""
    data = request.get_json(force=True, silent=True) or {}
    ids = data.get('profile_ids', [])
    result = profile_manager.export_profiles(ids)
    return _napi(result, msg='Exported')


# Register the Nexus API blueprint
app.register_blueprint(nexus_api)


def run_app():
    """Called by main_entry.py (frozen mode) to start the Flask server."""
    print("=" * 60)
    print("Gmail Bot Backend Server")
    print("=" * 60)
    print("Server starting on http://localhost:5000")
    print("[AUTH] No license required — open access")
    print("Server started - Ready to accept requests")
    print("=" * 60)
    app.run(host='127.0.0.1', port=5000, debug=False, threaded=True, use_reloader=False)


if __name__ == '__main__':
    run_app()
